from openpyxl.styles import Alignment

from modules.cailiaodingyi.funcs.funcs_pdf_change import update_element_name_data, \
    get_design_params_by_product_id, update_guankou_param_flex_db, query_guankou_affiliation, resolve_gasket_dimensions, query_guankou_codes,\
    invalidate_caches_for_product
from modules.cailiaodingyi.controllers.check_dianpian import clear_all_pn_user_input_for_product, force_recompute_and_update_pn
from modules.cailiaodingyi.funcs.funcs_pdf_input import query_all_guankou_categories
# from modules.cailiaodingyi.funcs.funcs_pdf_change import update_element_name_data, \
#     get_design_params_by_product_id, update_guankou_param_flex_db, query_guankou_affiliation, resolve_gasket_dimensions
from modules.condition_input.funcs.db_cnt import get_connection
from PyQt5.QtWidgets import (QTableWidgetItem, QTableWidget, QHeaderView, QWidget,
                             QMessageBox, QUndoStack, QFileDialog, QComboBox, QStyledItemDelegate, QShortcut,
                             QTabWidget, QStackedWidget)
from PyQt5.QtCore import Qt, QTimer
from PyQt5.QtGui import QColor, QStandardItemModel, QStandardItem, QBrush, QKeySequence
import re
import ast
import os
import math
import pandas as pd
from openpyxl.cell.cell import MergedCell
from openpyxl import load_workbook
from modules.chanpinguanli import bianl
from modules.condition_input.funcs.undo_command import CellEditCommand
from modules.condition_input.funcs.funcs_def_check import check_dn, check_work_pressure, check_work_temp_in, \
    check_work_temp_out, check_work_pressure_max, check_tubeplate_design_pressure_gap, \
    check_design_pressure, check_design_temp_max, check_design_temp_min, \
    check_in_out_pressure_gap, check_trail_stand_pressure_medium_density, check_insulation_layer_thickness, \
    check_insulation_material_density, check_def_trail_stand_pressure_lying, check_def_trail_stand_pressure_stand, \
    check_trail_stand_pressure_type, check_pressure_test_temp, check_avg_tube_metal_temp, check_avg_shell_metal_temp

#数据库连接
db_config_1 = {
    'host': 'localhost',
    'port': 3306,
    'user': 'root',
    'password': '123456',
    'database': '产品条件库'
}

#1206新修改-外径、外径系列、是否已外径为基准、公称直径联动
# === 旧版：外径自动填充映射表（已废弃，现从user_config的2.2.11.1读取） ===

#1206新修改-外径、外径系列、是否已外径为基准、公称直径联动
def _parse_int_safe(text: str):
    """从字符串中提取整数（忽略非数字），失败返回 None。"""
    try:
        import re
        m = re.findall(r"\d+", str(text))
        if not m:
            return None
        return int(m[0])
    except Exception:
        return None

#1206新修改-外径、外径系列、是否已外径为基准、公称直径联动
def _warn_once(viewer: QWidget, message: str, key: str, window_ms: int = 1500):
    """在给定时间窗口内仅弹一次同类提示（以 key 区分）。"""
    try:
        import time
        store = getattr(viewer, "_outer_warn_times", None)
        if store is None:
            store = {}
            setattr(viewer, "_outer_warn_times", store)
        now = int(time.time() * 1000)
        last = store.get(key, 0)
        if now - last < window_ms:
            return
        store[key] = now
        QMessageBox.warning(viewer, "提示", message)
    except Exception:
        pass

#1206新修改-外径、外径系列、是否已外径为基准、公称直径联动
def _choose_dn_with_prompt(viewer: QWidget):
    """
    检查“公称直径*”壳程/管程是否填写完整：
    - 任一侧缺失则弹窗点名提示；
    - 返回用于外径计算的 DN（两侧都空返回 None；两侧都有且不一致优先管程）。
    """
    table = getattr(viewer, "tableWidget_design_data", None)
    if table is None:
        return None
    target_row = None
    for r in range(table.rowCount()):
        it = table.item(r, 1)
        if it and it.text().strip() == "公称直径*":
            target_row = r
            break
    if target_row is None:
        return None

    shell_text = table.item(target_row, 3).text().strip() if table.item(target_row, 3) else ""
    tube_text  = table.item(target_row, 4).text().strip() if table.item(target_row, 4) else ""
    shell_dn = _parse_int_safe(shell_text) if shell_text else None
    tube_dn  = _parse_int_safe(tube_text) if tube_text else None

    # 仅提示一次：若任一侧缺失，仅在首次发现时提示；当两侧都补齐时重置提示标记
    missing = []
    if shell_dn is None:
        missing.append("壳程")
    if tube_dn is None:
        missing.append("管程")

    if missing:
        warned = getattr(viewer, "_dn_missing_warned", False)
        if not warned and not getattr(viewer, "_is_loading_data", False) and getattr(viewer, "_outer_autofill_ready", False):
            _warn_once(viewer, f"{'/'.join(missing)}公称直径未输入，请核对！", key="dn_missing")
            try:
                setattr(viewer, "_dn_missing_warned", True)
            except Exception:
                pass
            # 在首次确认提示后：切换到“设计数据”页签（tab_design_data），并定位缺失单元格
            try:
                if getattr(viewer, "_is_loading_data", False):
                    raise Exception("skip during loading")
                # 先确定目标页面：优先使用命名的 tab_design_data，否则回退到设计数据表所属页面
                page = getattr(viewer, "tab_design_data", None)
                # 如果没有提供页面对象，尝试从表控件向上找到归属页面
                if page is None:
                    try:
                        tbl = getattr(viewer, "tableWidget_design_data", None)
                        p = tbl.parent() if tbl is not None else None
                        while p is not None and not isinstance(p, (QTabWidget, QStackedWidget)):
                            page = p  # 记录可能的页面部件
                            p = p.parent()
                    except Exception:
                        pass

                switched = False
                # 优先切换 QTabWidget，这将同步更新顶部“选中”标签状态
                try:
                    for tw in viewer.findChildren(QTabWidget):
                        idx = tw.indexOf(page) if page is not None else -1
                        if idx == -1 and page is not None:
                            # 若 page 不是直接子项，尝试沿 page 的父链寻找 tw 的子页
                            pp = page
                            while pp is not None and idx == -1:
                                idx = tw.indexOf(pp)
                                pp = pp.parent()
                        if idx != -1:
                            tw.setCurrentIndex(idx)
                            switched = True
                            break
                except Exception:
                    pass

                # 若未找到 QTabWidget，再回退切换 QStackedWidget（无标签栏，仅内容切换）
                if not switched:
                    try:
                        for sw in viewer.findChildren(QStackedWidget):
                            idx = sw.indexOf(page) if page is not None else -1
                            if idx == -1 and page is not None:
                                pp = page
                                while pp is not None and idx == -1:
                                    idx = sw.indexOf(pp)
                                    pp = pp.parent()
                            if idx != -1:
                                sw.setCurrentIndex(idx)
                                switched = True
                                break
                    except Exception:
                        pass

                # 优先定位到第一个缺失侧
                target_col = 3 if ("壳程" in missing and "管程" not in missing) else 4 if ("管程" in missing and "壳程" not in missing) else (3 if "壳程" in missing else 4)
                if table is not None and target_row is not None and target_row >= 0:
                    table.setCurrentCell(target_row, target_col)
                    table.scrollToItem(table.item(target_row, target_col))
                    table.setFocus()
            except Exception:
                pass
    else:
        # 两侧DN都已补齐，清除一次性提示标记
        if getattr(viewer, "_dn_missing_warned", False):
            try:
                setattr(viewer, "_dn_missing_warned", False)
            except Exception:
                pass

    # 选择用于计算的 DN
    if shell_dn is None and tube_dn is None:
        return None
    if shell_dn is not None and tube_dn is not None:
        return tube_dn if tube_dn != shell_dn else shell_dn
    return tube_dn if tube_dn is not None else shell_dn

#1206新修改-外径、外径系列、是否已外径为基准、公称直径联动
def _get_product_type_safe(viewer: QWidget) -> str:
    """安全获取产品类型；无 product_id 或查询失败则返回空串。"""
    try:
        pid = getattr(viewer, "product_id", None)
        if not pid:
            return ""
        return get_product_type_from_db(pid) or ""
    except Exception:
        return ""

#1206新修改-外径、外径系列、是否已外径为基准、公称直径联动
def _is_shell_and_tube(viewer: QWidget) -> bool:
    """仅对管壳式热交换器启用外径相关逻辑。"""
    return _get_product_type_safe(viewer) == "管壳式热交换器"

#1206新修改-外径、外径系列、是否已外径为基准、公称直径联动
def _get_dn_from_design(viewer: QWidget):
    """读取设计数据表中的“公称直径*”，按优先级选择 DN 整数。"""
    table = getattr(viewer, "tableWidget_design_data", None)
    if table is None:
        return None
    target_row = None
    for r in range(table.rowCount()):
        it = table.item(r, 1)
        if it and it.text().strip() == "公称直径*":
            target_row = r
            break
    if target_row is None:
        return None
    shell_text = table.item(target_row, 3).text().strip() if table.item(target_row, 3) else ""
    tube_text  = table.item(target_row, 4).text().strip() if table.item(target_row, 4) else ""
    shell_dn = _parse_int_safe(shell_text) if shell_text else None
    tube_dn  = _parse_int_safe(tube_text) if tube_text else None
    if shell_dn is not None and tube_dn is not None:
        if shell_dn == tube_dn:
            return shell_dn
        # 不一致时优先管程
        return tube_dn
    return tube_dn if tube_dn is not None else shell_dn

# 0221新修改-配置库-外径系列-外径
# === 新增：从user_config读取配置的函数 ===
def _get_user_config_value(config_id: str):
    """
    从user_config表读取配置值
    :param config_id: 配置ID，如 "2.2.11.1"
    :return: 配置值（字符串），失败返回None
    """
    try:
        conn = get_connection(**db_config_config)
        try:
            with conn.cursor() as cursor:
                cursor.execute("SELECT value FROM user_config WHERE id = %s", (config_id,))
                result = cursor.fetchone()
                if result:
                    return result.get('value') if isinstance(result, dict) else result[0] if result else None
                return None
        finally:
            conn.close()
    except Exception as e:
        print(f"[读取user_config失败] id={config_id}, 错误: {e}")
        return None
# 0221新修改-配置库-外径系列-外径
# === 新增：判断外径系列（根据2.2.11.2和2.2.11.3） ===
def _determine_diameter_series():
    """
    根据user_config判断外径系列
    :return: "英制系列" 或 "公制系列"，失败返回None
    """
    try:
        val_2_2_11_2 = _get_user_config_value("2.2.11.2")
        val_2_2_11_3 = _get_user_config_value("2.2.11.3")
        
        if val_2_2_11_2 is None or val_2_2_11_3 is None:
            return None
        
        # 转换为布尔值（处理字符串"true"/"false"或布尔值）
        def to_bool(v):
            if isinstance(v, bool):
                return v
            if isinstance(v, str):
                v_lower = v.lower().strip()
                if v_lower in ('true', '1', 'yes'):
                    return True
                elif v_lower in ('false', '0', 'no'):
                    return False
            return bool(v)
        
        bool_2_2_11_2 = to_bool(val_2_2_11_2)
        bool_2_2_11_3 = to_bool(val_2_2_11_3)
        
        # 2.2.11.2为false，2.2.11.3为true → 公制系列
        if not bool_2_2_11_2 and bool_2_2_11_3:
            return "公制系列"
        # 2.2.11.2为true，2.2.11.3为false → 英制系列
        elif bool_2_2_11_2 and not bool_2_2_11_3:
            return "英制系列"
        else:
            # 其他情况，返回None
            return None
    except Exception as e:
        print(f"[判断外径系列失败] 错误: {e}")
        return None
# 0221新修改-配置库-外径系列-外径
# === 新增：解析映射表并查找外径值 ===
def _get_outer_diameter_from_mapping(dn: int, series: str):
    """
    从user_config的2.2.11.1映射表中查找外径值
    :param dn: 公称直径（整数）
    :param series: 外径系列（"英制系列"或"公制系列"）
    :return: 外径值（字符串），不存在返回None
    """
    try:
        mapping_data = _get_user_config_value("2.2.11.1")
        if not mapping_data:
            return None
        
        # 解析数据（可能是JSON字符串或已经是列表）
        import json
        if isinstance(mapping_data, str):
            try:
                mapping_list = json.loads(mapping_data)
            except:
                # 如果不是JSON，尝试用ast.literal_eval
                try:
                    mapping_list = ast.literal_eval(mapping_data)
                except:
                    return None
        else:
            mapping_list = mapping_data
        
        if not isinstance(mapping_list, list) or len(mapping_list) < 2:
            return None
        
        # 第一行是表头：["DN/OD", "inch", "metric"]
        # 确定系列对应的列索引
        header = mapping_list[0]
        if not isinstance(header, list) or len(header) < 3:
            return None
        
        # 根据系列确定列索引
        if series == "英制系列":
            col_index = 1  # "inch"列
        elif series == "公制系列":
            col_index = 2  # "metric"列
        else:
            return None
        
        # 查找匹配的DN值
        dn_str = str(dn)
        for row in mapping_list[1:]:  # 跳过表头
            if not isinstance(row, list) or len(row) < 3:
                continue
            if str(row[0]).strip() == dn_str:
                # 找到匹配的DN，返回对应系列的外径值
                outer_d = str(row[col_index]).strip()
                return outer_d if outer_d else None
        
        return None
    except Exception as e:
        print(f"[从映射表查找外径值失败] DN={dn}, series={series}, 错误: {e}")
        return None

#1206新修改-外径、外径系列、是否已外径为基准、公称直径联动
def _get_series_from_general(viewer: QWidget):
    table = getattr(viewer, "tableWidget_general_data", None)
    if table is None:
        return None
    for r in range(table.rowCount()):
        it = table.item(r, 1)
        if it and it.text().strip() == "外径系列":
            v_item = table.item(r, 3)
            return v_item.text().strip() if v_item else ""
    return None

#1206新修改-外径、外径系列、是否已外径为基准、公称直径联动
def _is_outer_by_diameter_enabled(viewer: QWidget) -> bool:
    """读取“是否以外径为基准*”是否为“是”。"""
    table = getattr(viewer, "tableWidget_general_data", None)
    if table is None:
        return False
    for r in range(table.rowCount()):
        it = table.item(r, 1)
        if it and it.text().strip() == "是否以外径为基准*":
            v_item = table.item(r, 3)
            val = v_item.text().strip() if v_item else ""
            return val == "是"
    return False

#1206新修改-外径、外径系列、是否已外径为基准、公称直径联动
def _set_general_outer_diameter(viewer: QWidget, text_val: str):
    """把值写入通用数据表“外径”的“数值”列，带撤销与防递归。"""
    table = getattr(viewer, "tableWidget_general_data", None)
    if table is None:
        return
    target_row = None
    for r in range(table.rowCount()):
        it = table.item(r, 1)
        if it and it.text().strip() == "外径":
            target_row = r
            break
    if target_row is None:
        return
    item = table.item(target_row, 3)
    if item is None:
        item = QTableWidgetItem()
        table.setItem(target_row, 3, item)
    old_val = item.text()
    if old_val == text_val:
        return
    try:
        table.blockSignals(True)
        item.setText(text_val)
    finally:
        table.blockSignals(False)
    undo_stack = getattr(viewer, "undo_stack", None)
    if undo_stack is not None:
        try:
            undo_stack.push(CellEditCommand(table, target_row, 3, old_val, text_val))
        except Exception:
            pass

# 0221新修改-配置库-外径系列-外径
# === 新增：设置外径系列到通用数据表 ===
def _set_general_outer_diameter_series(viewer: QWidget, series: str):
    """把外径系列值写入通用数据表"外径系列"的"数值"列"""
    table = getattr(viewer, "tableWidget_general_data", None)
    if table is None:
        return
    target_row = None
    for r in range(table.rowCount()):
        it = table.item(r, 1)
        if it and it.text().strip() == "外径系列":
            target_row = r
            break
    if target_row is None:
        return
    item = table.item(target_row, 3)
    if item is None:
        item = QTableWidgetItem()
        table.setItem(target_row, 3, item)
    old_val = item.text()
    if old_val == series:
        return
    try:
        table.blockSignals(True)
        item.setText(series)
    finally:
        table.blockSignals(False)

#1206新修改-外径、外径系列、是否已外径为基准、公称直径联动
def autofill_outer_diameter(viewer: QWidget):
    """
    基于 设计数据表"公称直径*" 与外径系列，自动填充通用数据表中的"外径"：
    1. 外径系列优先使用通用数据表当前值；
       若当前为空，再根据user_config(2.2.11.2 / 2.2.11.3)给一个默认系列；
    2. 外径值优先从user_config的2.2.11.1映射表查找；
    3. 映射表没有对应DN时，按公式 round(25.4 * DN / 25, 0) 计算，
       界面显示"—"，但计算值会保存到数据库。
    """
    # 未就绪时不进行自动填充（避免界面进入前弹窗）
    if not getattr(viewer, "_outer_autofill_ready", False):
        return
    # 仅针对管壳式热交换器
    if not _is_shell_and_tube(viewer):
        return
    if not _is_outer_by_diameter_enabled(viewer):
        return
    # 防重入：避免一次用户操作导致多次调用
    if getattr(viewer, "_outer_autofill_lock", False):
        return
    setattr(viewer, "_outer_autofill_lock", True)
    try:
        dn = _choose_dn_with_prompt(viewer)

        # 若DN缺失，直接写入"—"
        if dn is None:
            _set_general_outer_diameter(viewer, "—")
            # 清除计算值缓存
            if hasattr(viewer, "_calculated_outer_diameter"):
                viewer._calculated_outer_diameter = None
            setattr(viewer, "_outer_last_pair", (None, None))
            return

        # === 步骤1：确定外径系列（优先使用界面值，其次配置库默认） ===
        # ① 先从通用数据表当前值读取外径系列
        series = _get_series_from_general(viewer)

        # ② 若当前没有外径系列，再从user_config读取默认系列
        if not series:
            series_from_cfg = _determine_diameter_series()
            if series_from_cfg:
                _set_general_outer_diameter_series(viewer, series_from_cfg)
                series = series_from_cfg

        # ③ 依然没有系列信息，则无法计算外径，只把界面置为"—"
        if not series:
            _set_general_outer_diameter(viewer, "—")
            if hasattr(viewer, "_calculated_outer_diameter"):
                viewer._calculated_outer_diameter = None
            setattr(viewer, "_outer_last_pair", (dn, None))
            return

        # 若与上一次处理的 (dn, series) 完全一致，直接返回，避免重复计算
        last_pair = getattr(viewer, "_outer_last_pair", None)
        cur_pair = (dn, series)
        if last_pair == cur_pair:
            return

        # === 步骤2：从映射表查找外径值 ===
        outer_d_from_mapping = _get_outer_diameter_from_mapping(dn, series)

        if outer_d_from_mapping:
            # 映射表中存在，直接使用映射值
            _set_general_outer_diameter(viewer, outer_d_from_mapping)
            # 清除计算值缓存（因为使用的是映射值）
            if hasattr(viewer, "_calculated_outer_diameter"):
                viewer._calculated_outer_diameter = None
        else:
            # 映射表中不存在，使用公式计算
            # 公式：外径值 = round(25.4 * DN / 25, 0)
            calculated_value = round(25.4 * dn / 25, 0)
            calculated_value_str = str(int(calculated_value))
            
            # 界面显示"—"
            _set_general_outer_diameter(viewer, "—")
            
            # 保存计算值到缓存（用于保存到数据库）
            if not hasattr(viewer, "_calculated_outer_diameter"):
                viewer._calculated_outer_diameter = {}
            viewer._calculated_outer_diameter = calculated_value_str
            print(f"[外径计算] DN={dn}, 系列={series}, 计算值={calculated_value_str}, 界面显示=—")
        
        setattr(viewer, "_outer_last_pair", cur_pair)
    finally:
        setattr(viewer, "_outer_autofill_lock", False)

db_config_2 = {
    'host': 'localhost',
    'port': 3306,
    'user': 'root',
    'password': '123456',
    'database': '产品设计活动库'
}

db_config_3 = {
    'host': 'localhost',
    'port': 3306,
    'user': 'root',
    'password': '123456',
    'database': '产品需求库'
}

db_config_4 = {
    'host': 'localhost',
    'port': 3306,
    'user': 'root',
    'password': '123456',
    'database': '项目需求库'
}
# 0221新修改-配置库-外径系列-外径
# 配置库连接配置（用于读取user_config表）
db_config_config = {
    'host': 'localhost',
    'port': 3306,
    'user': 'root',
    'password': '123456',
    'database': '配置库'
}

# === 新增：模式顺序相关工具  开始 ========================================= 新增
# 1) 我们假设“参数顺序模板表”位于【产品条件库】数据库，有字段：
#    - 模板名称 (varchar)
#    - 参数顺序 (varchar，形如 "[1,6,9,10,...]" 的字符串)
# 2) 仅用于“界面显示顺序”的重排；数据库与Excel保存一律按默认顺序。

def fetch_all_mode_orders():
    """
    读取【产品条件库.参数顺序模板表】 -> 返回 {模板名称: [id1,id2,...]}。
    若表不存在或解析失败，返回空dict。
    """
    try:
        conn = get_connection(**db_config_1)
        with conn.cursor() as cur:
            cur.execute("SELECT 模板名称, 参数顺序 FROM 设计数据参数顺序模板表")
            rows = cur.fetchall()
        conn.close()
        result = {}
        for r in rows:
            name = (r.get("模板名称") or "").strip()
            raw = (r.get("参数顺序") or "").strip()
            try:
                # 支持 "[1,2,3]" 或 "1,2,3" 两种
                if raw.startswith("["):
                    seq = ast.literal_eval(raw)
                else:
                    seq = [int(x) for x in re.split(r"[，,]\s*", raw) if x.strip()]
                seq = [int(x) for x in seq]
            except Exception:
                seq = []
            if name and seq:
                result[name] = seq
                print(result)
        return result
    except Exception:
        return {}

def _read_row_as_list(table_widget, row):
    """把一行所有列的 QTableWidgetItem 文本读取为 list[str]；空位返回''。"""
    cols = table_widget.columnCount()
    values = []
    for c in range(cols):
        item = table_widget.item(row, c)
        values.append(item.text() if item else "")
    return values

def _write_row_from_list(table_widget, row, values, header_userroles=None):
    """把 list[str] 写回到指定行；尽可能维持原对齐/可编辑属性的简化版。"""
    cols = table_widget.columnCount()
    for c in range(cols):
        val = values[c] if c < len(values) else ""
        item = QTableWidgetItem(val)
        # 名称列/单位列按原 fill_table_widget 约定设置 flags & 对齐
        header_item = table_widget.horizontalHeaderItem(c)
        header_text = header_item.text() if header_item else ""
        # 默认：可编辑
        if header_text in ("序号",):
            # 序号列：不可编辑，居中
            item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
            item.setTextAlignment(Qt.AlignCenter)
        elif header_text in ("参数名称", "规范/标准名称", "用途", "细类"):
            item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)  # 名称等不允许改
            item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter if header_text in ("参数名称","规范/标准名称") else Qt.AlignCenter)
        elif header_text in ("参数单位",):
            item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
            item.setTextAlignment(Qt.AlignCenter)
        else:
            item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled | Qt.ItemIsEditable)
            item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter if header_text in ("规范/标准代号",) else Qt.AlignCenter)
        if header_userroles and header_userroles.get(c) is not None:
            item.setData(Qt.UserRole, header_userroles[c])
        table_widget.setItem(row, c, item)

def capture_default_order(table_widget):
    """
    记录该表“默认顺序”的行号索引（以‘参数ID/序号’列的值为依据）。
    要求第0列是‘序号’（真实字段为 参数IDxxx）。
    """
    ids = []
    for r in range(table_widget.rowCount()):
        item = table_widget.item(r, 0)
        ids.append(int(item.text())) if (item and item.text().strip().isdigit()) else ids.append(None)
    # 保存两份：id列表 与 id->原始行号映射
    table_widget._default_param_ids = ids[:]  # 可能有空行 None
    id2row = {}
    for r, pid in enumerate(ids):
        if pid is not None and pid not in id2row:
            id2row[pid] = r
    table_widget._default_id2row = id2row

def restore_default_order(table_widget):
    # 0103新修改
    """
    恢复表格到默认顺序（严格按照 _default_param_ids 的顺序，不做任何分组处理）。
    用于从工作模式切换回设计模式时恢复原始顺序。
    """
    if table_widget.rowCount() == 0 or table_widget.columnCount() == 0:
        return
    
    # 获取默认顺序
    default_ids = getattr(table_widget, "_default_param_ids", None)
    if not default_ids:
        return
    
    # 🔒 关键：关闭排序 & 冻结刷新
    was_sorting = table_widget.isSortingEnabled()
    if was_sorting:
        table_widget.setSortingEnabled(False)
    table_widget.setUpdatesEnabled(False)
    
    # 读取当前所有行
    all_rows = [_read_row_as_list(table_widget, r) for r in range(table_widget.rowCount())]
    
    # 建立当前行的 ID 到行的映射（每个ID只映射一次，取第一个匹配的行）
    id_to_row = {}
    none_rows = []  # 存储ID为None的行
    
    for idx, row_data in enumerate(all_rows):
        if len(row_data) > 0 and row_data[0]:  # 第0列是参数ID
            try:
                param_id = int(row_data[0].strip()) if row_data[0].strip().isdigit() else None
                if param_id is not None:
                    if param_id not in id_to_row:
                        id_to_row[param_id] = row_data
                else:
                    none_rows.append(row_data)
            except Exception:
                none_rows.append(row_data)
        else:
            none_rows.append(row_data)
    
    # 按照 default_ids 的顺序重建行
    new_rows = []
    used_ids = set()
    
    # 第一步：按照 default_ids 的顺序添加行（严格按顺序）
    for param_id in default_ids:
        if param_id is not None:
            if param_id in id_to_row and param_id not in used_ids:
                new_rows.append(id_to_row[param_id])
                used_ids.add(param_id)
        else:
            # 如果是None，从none_rows中取一个（如果有的话）
            if none_rows:
                new_rows.append(none_rows.pop(0))
    
    # 第二步：添加 default_ids 中没有但表格中存在的行（这些行可能是新增的，保留在末尾）
    for param_id, row_data in id_to_row.items():
        if param_id not in used_ids:
            new_rows.append(row_data)
    
    # 第三步：添加剩余的ID为None的行
    new_rows.extend(none_rows)
    
    # 清空旧内容并写入新顺序
    table_widget.clearContents()
    table_widget.setRowCount(len(new_rows))
    for r, row_vals in enumerate(new_rows):
        _write_row_from_list(table_widget, r, row_vals)
    
    # 恢复刷新
    table_widget.setUpdatesEnabled(True)
    table_widget.viewport().update()

def apply_mode_param_order(table_widget, target_id_seq):
    # 1226新修改_工作模式不同产品参数显示顺序调整
    """
    按照 target_id_seq 对表格进行"界面行顺序"的重排（不改单元格内容）。
    特殊处理：工作模式下，所有带*的必填项参数必须优先显示在不带*的参数之前。
    仅处理第0列可解析为 int 的行；其余行保持在末尾原顺序。
    """
    if table_widget.rowCount() == 0 or table_widget.columnCount() == 0:
        return

    # 🔒 关键：关闭排序 & 冻结刷新
    was_sorting = table_widget.isSortingEnabled()
    if was_sorting:
        table_widget.setSortingEnabled(False)
    table_widget.setUpdatesEnabled(False)

    # 取当前整表文稿
    all_rows = [_read_row_as_list(table_widget, r) for r in range(table_widget.rowCount())]

    # 当前每行的 参数ID（来自第0列）和参数名称（第1列）
    cur_ids = []
    param_names = []
    for r in range(table_widget.rowCount()):
        it = table_widget.item(r, 0)
        try:
            cur_ids.append(int(it.text().strip()) if it and it.text().strip() else None)
        except Exception:
            cur_ids.append(None)
        # 获取参数名称（第1列），用于判断是否为必填项（带*）
        name_item = table_widget.item(r, 1)
        param_names.append(name_item.text().strip() if name_item else "")

    id2rows = {}
    id_to_required = {}  # 记录每个ID对应的参数是否为必填项（带*）
    others = []
    for idx, pid in enumerate(cur_ids):
        if pid is None:
            others.append(all_rows[idx])
        else:
            id2rows[pid] = all_rows[idx]
            # 判断该参数是否为必填项（参数名称包含*）
            is_required = "*" in param_names[idx]
            id_to_required[pid] = is_required

    new_rows = []
    
    # ✅ 第一步：按照 target_id_seq 的顺序，收集所有表格中实际存在的必填项（带*）
    for pid in target_id_seq:
        if pid in id2rows and id_to_required.get(pid, False):
            new_rows.append(id2rows[pid])
            id2rows.pop(pid)  # 从字典中移除，标记为已处理
    
    # ✅ 第二步：收集剩余的必填项（不在模板顺序中，但在表格中存在的带*参数）
    # 按原始顺序排列剩余的必填项
    remaining_required_ids = [pid for pid in cur_ids if pid is not None and pid in id2rows and id_to_required.get(pid, False)]
    for pid in remaining_required_ids:
        new_rows.append(id2rows[pid])
        id2rows.pop(pid)
    
    # ✅ 第三步和第四步合并：收集所有剩余的非必填项，按ID从小到大排序
    remaining_non_required_ids = [pid for pid in id2rows.keys() if not id_to_required.get(pid, False)]
    remaining_non_required_ids.sort()  # 按ID从小到大排序
    for pid in remaining_non_required_ids:
        new_rows.append(id2rows[pid])
        id2rows.pop(pid)
    
    # ✅ 第五步：处理其他特殊情况（理论上此时id2rows应该为空，但保留此逻辑以防万一）
    remaining_ids = [pid for pid in cur_ids if pid is not None and pid in id2rows]
    for pid in remaining_ids:
        new_rows.append(id2rows[pid])

    new_rows.extend(others)

    # ⚠️ 清空旧内容再写入，避免残留
    table_widget.clearContents()
    table_widget.setRowCount(len(new_rows))
    for r, row_vals in enumerate(new_rows):
        _write_row_from_list(table_widget, r, row_vals)

    # ✅ 恢复刷新（但不要恢复排序）
    table_widget.setUpdatesEnabled(True)
    table_widget.viewport().update()

def get_row_index_order_for_default_write(table_widget):
    """
    保存/导出Excel时使用：返回“默认顺序”的写出索引序列。
    默认顺序来自 capture_default_order() 记录的 _default_param_ids。
    若没记录，则退化为 [0..n-1]。
    """
    n = table_widget.rowCount()
    if not hasattr(table_widget, "_default_param_ids"):
        return list(range(n))

    ids = table_widget._default_param_ids
    id_rows = [(pid, idx) for idx, pid in enumerate(ids) if pid is not None]
    # 按“默认ID出现顺序”排序（就是初始载入时的顺序），None 行放后面保持原序
    ordered_indices = [idx for (_, idx) in id_rows]
    none_indices = [idx for idx, pid in enumerate(ids) if pid is None]
    return ordered_indices + none_indices
# === 新增：模式顺序相关工具  结束 =========================================


"""导入数据库数据表相关函数"""
def make_header_item(text):
    """
    创建一个“仿真表头”项：
    - 居中对齐
    - 加粗字体
    - 可选中（点击高亮列）
    - 不设置背景颜色（保留原始白色）
    """
    item = QTableWidgetItem(text)
    item.setTextAlignment(Qt.AlignCenter)

    # ✅ 可选中 + 不可编辑（用户可以点击高亮，但不能修改内容）
    item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)

    # ✅ 设置加粗字体
    font = item.font()
    font.setBold(True)
    item.setFont(font)

    return item

def load_design_data_if_exists(product_id, product_form="all"):
    """
    给定产品ID，从设计活动库优先加载5张数据表，如不存在则退回产品条件库模板表。
    【新增】当从模板库加载时，会根据 product_form 决定是否加载特定产品形式的参数。
    """
    design_tables = {
        "产品标准": "产品设计活动表_产品标准数据表",
        "设计数据": "产品设计活动表_设计数据表",
        "通用数据": "产品设计活动表_通用数据表",
        "检测数据": "产品设计活动表_无损检测数据表",
        "涂漆数据": "产品设计活动表_涂漆数据表"
    }

    template_tables = {
        "产品标准": "产品标准数据模板表",
        "设计数据": "设计数据模板表",
        "通用数据": "通用数据模板表",
        "检测数据": "无损检测数据模板表",
        "涂漆数据": "涂漆数据模板表"
    }

    result = {"数据": {}}  # 确保返回数据时有 "数据" 键

    # 判断设计库中是否有记录
    design_data_exists = False
    connection = get_connection(**db_config_2)
    try:
        with connection.cursor() as cursor:
            cursor.execute(f"SELECT 1 FROM {design_tables['产品标准']} WHERE 产品ID = %s LIMIT 1", (product_id,))
            design_data_exists = bool(cursor.fetchone())
    finally:
        connection.close()

    # 逐表加载（优先设计库，后退模板库）
    for key in design_tables:
        db_used = db_config_2 if design_data_exists else db_config_1
        table_name = design_tables[key] if design_data_exists else template_tables[key]
        # ▼▼▼【诊断点 1】: 在循环开始时打印当前处理的表和产品信息 ▼▼▼
        print("\n" + "="*50)
        print(f"开始处理: [ {key} ] | 产品形式: '{product_form}' | 数据源: {'设计活动库' if design_data_exists else '模板库'}")
        print(f"目标表名: `{table_name}`")
        print("="*50)
        connection = get_connection(**db_used)
        try:
            with connection.cursor() as cursor:
                # 获取字段名，按表类型决定是否保留 参数ID 字段
                cursor.execute(f"DESCRIBE {table_name}")
                columns = cursor.fetchall()
                internal_columns_to_hide = ['所属类型', '所属型式']

                # 确定要查询的列名
                column_names = [
                    col['Field'] for col in columns
                    if '产品ID' not in col['Field'] and '更改状态' not in col['Field']

                ]

                field_str = ', '.join([f"`{col}`" for col in column_names])
                # 构造查询语句
                sql_query = f"SELECT {field_str} FROM `{table_name}`"
                params = []

                if design_data_exists:#从设计活动库加载
                    # cursor.execute(f"SELECT {field_str} FROM {table_name} WHERE 产品ID = %s", (product_id,))
                    sql_query += " WHERE `产品ID` = %s"
                    params.append(product_id)
                else:#从模板库加载 用产品形式过滤
                    # ▼▼▼【核心修改点 2】: 从模板库加载时，应用产品形式过滤 ▼▼▼
                    # 检查是否是受影响的表，并且模板表里真的有所属型式 列
                    is_form_dependent_table = key in ["设计数据"] # 只影响设计数据表
                    form_column_name = '所属型式'
                    has_form_column = any(col['Field'] == form_column_name for col in columns)

                    if is_form_dependent_table and has_form_column:
                        # 1216新修改-bem也显示两个金属温度的参数
                        # 2026-01: AEM或后续的产品型式需要显示（直接在数据库表里加上产品型式）
                        # 用 FIND_IN_SET 做“精确匹配”，避免 LIKE 子串误命中。
                        if product_form and product_form != "all":
                            sql_query += (
                                f" WHERE `{form_column_name}` = %s"
                                f" OR FIND_IN_SET(%s, `{form_column_name}`)"
                            )
                            params.extend(['all', product_form])
                        else:
                            sql_query += f" WHERE `{form_column_name}` = %s"
                            params.append('all')
                    # cursor.execute(f"SELECT {field_str} FROM {table_name} WHERE 所属形式 = %s",(product_form,))
                # ▼▼▼【诊断点 2】: 打印最终要执行的 SQL 和参数 (最关键！) ▼▼▼
                print(f"[SQL诊断] 最终执行的查询语句:\n    {sql_query}")
                print(f"[SQL诊断] 最终绑定的参数:\n    {tuple(params)}")
                # 执行最终构建的查询
                cursor.execute(sql_query, tuple(params))
                rows = cursor.fetchall()
                # ▼▼▼【诊断点 3】: 打印查询结果的数量 ▼▼▼
                print(f"[SQL诊断] 查询返回的行数: {len(rows)}")
                if len(rows) > 0:
                    print(f"[SQL诊断] 查询结果第一行示例: {rows[0]}")
                print("-" * 50)

                # ▼▼▼【核心修改点 3】: 从最终结果中移除用于过滤的列 ▼▼▼
                # 无论从哪里加载，都不希望在UI上看到 '所属型式' 这一列
                display_column_names = [
                    name for name in column_names
                    if name not in internal_columns_to_hide # 可以在这里添加更多不想显示的内部列
                ]

                # 清洗空值
                for row in rows:
                    for k in row:
                        if row[k] is None:
                            row[k] = ""

                data = {
                    "headers": display_column_names,
                    "rows": rows,
                    "count": len(rows)
                }

                # 设置界面用的“序号列”字段名（实际用于表格第0列）
                # if preserve_param_id:
                #     data["prepend_index_header"] = column_names[0]
                if display_column_names and display_column_names[0].endswith("参数ID"):
                    data["prepend_index_header"] = display_column_names[0]

                if key == "检测数据":
                    data["格式化"] = format_trail_table(display_column_names, rows)
                if key == "涂漆数据":
                    data["格式化"] = format_coating_table(display_column_names, rows)

                result["数据"][key] = data  # 存表格数据

        finally:
            connection.close()

    # 设置数据来源状态和导入状态
    result["data_source_status"] = "设计活动库" if design_data_exists else "条件模板"
    # result["import_status"] = True if design_data_exists or len(rows) > 0 else False
    # 只要任何一个表有数据，就认为导入成功
    result["import_status"] = any(d["count"] > 0 for d in result["数据"].values())
    return result

def format_trail_table(headers, rows):
    # 将检测数据表按“接头种类”字段进行分组（用于合并同类行显示）
    grouped = {}
    for row in rows:
        接头种类 = row['接头种类']
        if 接头种类 not in grouped:
            grouped[接头种类] = []
        grouped[接头种类].append(row)
    return grouped

def format_coating_table(headers, rows):
    """
    将涂漆数据按“用途”字段进行分组
    并将“用途”字段中的复合值进行拆分（提取出 细类：底漆、中间漆、面漆）
    如：'内涂漆（壳程）_底漆' -> 用途='内涂漆（壳程）', 细类='底漆'
    """
    grouped = {}
    for row in rows:
        用途字段 = row['用途']
        if '）_' in 用途字段:
            左, 右 = 用途字段.split('）_')
            用途 = 左 + '）'     # 例：'内涂漆（壳程）'
            涂层 = 右           # 例：'底漆'
        else:
            用途 = 用途字段
            涂层 = ""

        row['_细类'] = 涂层  # ✅ 注意是临时字段
        if 用途 not in grouped:
            grouped[用途] = []
        grouped[用途].append(row)
    return grouped

def render_grouped_table(table_widget, grouped_data, headers, group_key_column=0):
    header_rows = 2
    total_rows = sum(len(v) for v in grouped_data.values())
    table_widget.setRowCount(total_rows + header_rows)
    table_widget.setColumnCount(len(headers))
    table_widget.setHorizontalHeaderLabels(headers)

    current_row = header_rows
    for group_key, row_list in grouped_data.items():
        span_start = current_row
        for row in row_list:
            jt_type = row.get("接头种类", "").strip()  # ✅ 每行记住接头种类

            for col_idx, key in enumerate(headers):
                if col_idx == group_key_column:
                    # ✅ 给合并列每行都建一个 item（保证 UserRole 存在）
                    group_item = QTableWidgetItem(group_key)
                    group_item.setTextAlignment(Qt.AlignCenter)
                    group_item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                    group_item.setData(Qt.UserRole, jt_type)  # 存 jt_type
                    table_widget.setItem(current_row, col_idx, group_item)
                    continue

                val = str(row.get(key, ""))
                item = QTableWidgetItem(val)
                item.setTextAlignment(Qt.AlignCenter)
                item.setData(Qt.UserRole, jt_type)  # ✅ 每个单元格都存接头种类

                # 第2列（检测方法）不可编辑
                if col_idx == 1:
                    item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                else:
                    item.setFlags(Qt.ItemIsEditable | Qt.ItemIsSelectable | Qt.ItemIsEnabled)

                detect_method = row.get("检测方法", "").strip()
                # 技术等级为 '/' → 不可编辑
                if detect_method in ["M.T.", "P.T.", "M.T.[FB]"] and key in ["壳程_技术等级", "管程_技术等级"]:
                    item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)

                # 合格级别为 M.T./P.T. → 不可编辑
                if detect_method in ["M.T.", "P.T.", "M.T.[FB]"] and key in ["壳程_合格级别", "管程_合格级别"]:
                    item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)

                # 特殊逻辑：T（管头） → 壳程三列不可编辑，且默认填 "/"
                if jt_type == "T（管头）" and key in ["壳程_技术等级", "壳程_合格级别", "壳程_检测比例"]:
                    item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                    if not val.strip():
                        item.setText("/")

                # 默认值作为校验基准
                if key.endswith("技术等级") or key.endswith("合格级别"):
                    side = "壳程" if "壳程" in key else "管程"
                    ratio = str(row.get(f"{side}_检测比例", "")).strip()
                    field_type = "技术等级" if "技术等级" in key else "合格级别"
                    from .funcs_cdt_input import compute_trail_default_grade
                    default_val = compute_trail_default_grade(detect_method, ratio, field_type)
                    if default_val:
                        item.setData(Qt.UserRole + 2, default_val)

                table_widget.setItem(current_row, col_idx, item)

            current_row += 1

        # ✅ 设置合并列的视觉效果（显示只在起始行）
        table_widget.setSpan(span_start, group_key_column, len(row_list), 1)

    table_widget.resizeColumnsToContents()


def set_multilevel_headers(table_widget: QTableWidget, top_headers: list, sub_headers: list, span_map: list):
    """
    设置 QTableWidget 的两级表头结构（不破坏数据内容）。
    - top_headers：一级标题（支持横向合并、纵向合并）
    - sub_headers：二级字段名
    - span_map：格式如 [(start, span)]，表示从哪列开始、合并几列
    """

    col_count = sum(span for _, span in span_map)
    header_rows = 2

    # 创建表头：先扩展一张空表，仅用于设置头部结构（内容之后渲染）
    table_widget.setColumnCount(col_count)
    table_widget.setRowCount(header_rows)  # 只设置前2行用于表头

    # 设置一级标题（带纵向合并）
    for i, (start, span) in enumerate(span_map):
        header_text = top_headers[i] if top_headers[i].strip() else " "
        item = make_header_item(header_text)

        if span == 1:
            table_widget.setSpan(0, start, 2, 1)  # 垂直合并2行
            table_widget.setItem(0, start, item)
        else:
            table_widget.setSpan(0, start, 1, span)  # 水平合并
            table_widget.setItem(0, start, item)

    # 设置子标题
    sub_col = 0
    for i, (start, span) in enumerate(span_map):
        if span > 1:
            for offset in range(span):
                item = make_header_item(sub_headers[sub_col])
                table_widget.setItem(1, start + offset, item)
                sub_col += 1
        else:
            sub_col += 1  # 跳过

    # 不设置内容行，让调用者单独设置数据内容行（从第2行开始）
    table_widget.verticalHeader().setVisible(False)
    table_widget.horizontalHeader().setVisible(False)

def render_coating_table(table_widget: QTableWidget, grouped_data: dict, exec_std_value: str = ""):
    headers = ["用途", "细类", "油漆类别", "颜色", "干膜厚度（μm）", "涂漆面积", "备注"]
    total_data_rows = sum(len(rows) for rows in grouped_data.values())
    table_widget.setRowCount(2 + total_data_rows)
    table_widget.setColumnCount(len(headers))

    all_rows = [row for group in grouped_data.values() for row in group]
    std_value = exec_std_value

    table_widget.verticalHeader().setVisible(False)
    table_widget.horizontalHeader().setVisible(False)

    # ✅ 第一行：执行标准/规范
    table_widget.setSpan(0, 0, 1, 2)
    table_widget.setItem(0, 0, make_header_item("执行标准/规范"))
    std_item = QTableWidgetItem(std_value)
    std_item.setTextAlignment(Qt.AlignCenter)
    std_item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
    table_widget.setSpan(0, 2, 1, len(headers) - 2)
    table_widget.setItem(0, 2, std_item)

    # ✅ 第二行：表头
    table_widget.setSpan(1, 0, 1, 2)
    table_widget.setItem(1, 0, make_header_item("用途"))
    for col, header in enumerate(headers[2:], start=2):
        table_widget.setItem(1, col, make_header_item(header))

    current_row = 2
    for group_key, row_list in grouped_data.items():
        span_start = current_row
        merge_data = {"涂漆面积": "", "备注": ""}

        for idx, row in enumerate(row_list):
            values = [
                group_key,
                row.get("_细类", ""),
                row.get("油漆类别", ""),
                row.get("颜色", ""),
                row.get("干膜厚度（μm）", ""),
                row.get("涂漆面积", ""),
                row.get("备注", "")
            ]
            for col, val in enumerate(values):
                val = "" if val is None else str(val)
                item = QTableWidgetItem(val)
                item.setTextAlignment(Qt.AlignCenter)

                # ✅ 设置可编辑性（只用途/细类列是只读）
                if col in (0, 1):
                    item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                else:
                    item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled | Qt.ItemIsEditable)

                table_widget.setItem(current_row, col, item)

            if idx == 0:
                merge_data["涂漆面积"] = str(row.get("涂漆面积", "") or "")
                merge_data["备注"] = str(row.get("备注", "") or "")

            current_row += 1

        row_count = len(row_list)

        # ✅ 合并用途列
        item = QTableWidgetItem(group_key)
        item.setTextAlignment(Qt.AlignCenter)
        item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
        table_widget.setSpan(span_start, 0, row_count, 1)
        table_widget.setItem(span_start, 0, item)

        # ✅ 合并涂漆面积
        area_item = QTableWidgetItem(merge_data["涂漆面积"])
        area_item.setTextAlignment(Qt.AlignCenter)
        area_item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled | Qt.ItemIsEditable)
        table_widget.setSpan(span_start, 5, row_count, 1)
        table_widget.setItem(span_start, 5, area_item)

        # ✅ 合并备注
        comment_item = QTableWidgetItem(merge_data["备注"])
        comment_item.setTextAlignment(Qt.AlignCenter)
        comment_item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled | Qt.ItemIsEditable)
        table_widget.setSpan(span_start, 6, row_count, 1)
        table_widget.setItem(span_start, 6, comment_item)

    table_widget.resizeColumnsToContents()

    # ✅ 设置 logical_headers：确保校验函数能获取正确列名
    table_widget.logical_headers = [
        "用途", "细类", "油漆类别", "颜色", "干膜厚度（μm）", "涂漆面积", "备注"
    ]

"""表格显示样式"""
def get_merged_cell_start(table_widget, row, col):
    """返回 (row, col) 所属合并单元格的起始行"""
    for r in range(table_widget.rowCount()):
        rowspan = table_widget.rowSpan(r, col)
        if rowspan > 1 and r <= row < r + rowspan:
            return r
    return row

def highlight_entire_row(table_widget):
    selected_indexes = table_widget.selectedIndexes()
    if not selected_indexes:
        return

    selected_rows = {i.row() for i in selected_indexes}
    selected_cols = {i.column() for i in selected_indexes}

    # ✅ 只在真正点击了表头时跳过整行高亮
    row_count = table_widget.rowCount()
    is_full_column_selected = (
        len(selected_cols) == 1 and
        len(selected_rows) >= row_count and
        all(table_widget.model().index(r, list(selected_cols)[0]) in selected_indexes for r in range(row_count))
    )
    if is_full_column_selected:
        return

    # ✅ 清除旧高亮（保持缺失项不动）
    for row in range(table_widget.rowCount()):
        for col in range(table_widget.columnCount()):
            item = table_widget.item(row, col)
            if item:
                if item.data(Qt.UserRole + 1) == "missing":
                    continue
                if row % 2 == 0:
                    item.setBackground(QColor("#ffffff"))
                else:
                    item.setBackground(QColor("#f0f0f0"))
                item.setForeground(QBrush())

    # ✅ 单独处理：合并单元格块（只高亮合并区域）
    for index in selected_indexes:
        row, col = index.row(), index.column()
        rowspan = table_widget.rowSpan(row, col)
        colspan = table_widget.columnSpan(row, col)

        if rowspan > 1 or colspan > 1:
            for r in range(row, row + rowspan):
                for c in range(col, col + colspan):
                    item = table_widget.item(r, c)
                    if item and item.data(Qt.UserRole + 1) != "missing":
                        item.setBackground(QColor("#d0e7ff"))
                        item.setForeground(QBrush(Qt.black))

    # ✅ 收集所有普通格所在的行（跳过合并起始格）
    rows_to_highlight = set()
    for index in selected_indexes:
        row, col = index.row(), index.column()
        rowspan = table_widget.rowSpan(row, col)
        colspan = table_widget.columnSpan(row, col)
        if rowspan == 1 and colspan == 1:
            rows_to_highlight.add(row)

    # ✅ 普通整行高亮（非合并格）
    for row in rows_to_highlight:
        for col in range(table_widget.columnCount()):
            if table_widget.rowSpan(row, col) > 1 or table_widget.columnSpan(row, col) > 1:
                continue  # 跳过合并格
            item = table_widget.item(row, col)
            if item and item.data(Qt.UserRole + 1) != "missing":
                item.setBackground(QColor("#d0e7ff"))
                item.setForeground(QBrush(Qt.black))

def apply_table_style(table_widget):
    table_widget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
    table_widget.verticalHeader().setVisible(False)
    table_widget.setAlternatingRowColors(True)
    table_widget.setSelectionBehavior(table_widget.SelectItems)

    # ✅ 为表头加上四边边框线 已修改
    table_widget.horizontalHeader().setStyleSheet("""
        QHeaderView::section {
            border: 1px solid #D8D8D8;        /* 更细更柔和的边框 */
            background-color: white;         /* 白色背景 */
            color: black;                    /* 黑色字体 */
            padding: 4px;                    /* 内边距让文字不挤 */
            font-weight: bold;               /* 加粗字体 */
        }
    """)


#新增
def shrink_index_column(table_widget, width: int = 100):
    """
    将第 0 列（默认是“序号”列）设为较小宽度
    """
    header = table_widget.horizontalHeader()
    header.setSectionResizeMode(0, QHeaderView.Fixed)
    table_widget.setColumnWidth(0, width)
#新增
def shrink_unit_column(table_widget, width: int = 300):
    """
    将第 2 列（默认是“参数单位”列）设为较小宽度
    """
    header = table_widget.horizontalHeader()
    header.setSectionResizeMode(2, QHeaderView.Fixed)
    table_widget.setColumnWidth(2, width)


"""存入数据库相关函数"""

def get_table_header_columns(table_widget):
    headers = []
    for col in range(table_widget.columnCount()):
        item = table_widget.horizontalHeaderItem(col)
        if item:
            true_field = item.data(Qt.UserRole)
            headers.append(true_field if true_field else item.text())
    return headers

def get_table_data(table_widget):
    """
    提取表格所有行数据为结构化列表，每行是一个 dict（包含第0列）
    """
    headers = get_table_header_columns(table_widget)
    data = []

    for row in range(table_widget.rowCount()):
        row_data = {}
        for col_index, header in enumerate(headers):
            item = table_widget.item(row, col_index)
            value = item.text() if item else ""
            row_data[header] = value
        data.append(row_data)

    return data

def save_data_to_database(data, product_id, table_name, table_widget, is_from_design_lib=True, viewer=None):
    """
    将表格数据保存至数据库：
    - 无论是 INSERT 还是 UPDATE，统一先对比模板表字段值，判断更改状态；
    - 更改状态字段统一标记；
    - 特殊处理：如果外径显示为"—"但有计算值缓存，保存计算值到数据库。
    """
    connection = get_connection(**db_config_2)
    try:
        with connection.cursor() as cursor:
            header_columns = get_table_header_columns(table_widget)

            # 获取数据库字段结构
            cursor.execute(f"DESCRIBE {table_name}")
            table_columns = cursor.fetchall()
            db_fields = [col['Field'] for col in table_columns]

            # 获取"更改状态"字段名
            change_status_column = None
            for col in table_columns:
                if re.search(r'更改状态$', col['Field']):
                    change_status_column = col['Field']
                    break
            if not change_status_column:
                raise ValueError("未找到更改状态字段")

            # 确定"参数名称"字段
            name_column = "规范/标准名称" if "产品标准" in table_name else "参数名称"

            # === 参数ID字段映射（避免用"序号"）===
            id_field_mapping = {
                "产品设计活动表_产品标准数据表": "产品标准参数ID",
                "产品设计活动表_设计数据表": "设计数据参数ID",
                "产品设计活动表_通用数据表": "通用数据参数ID"
            }
            param_id_field = id_field_mapping.get(table_name, table_columns[0]['Field'])

            # UI 表头第0列（序号）→ 实际数据库的参数ID字段
            param_id_column = header_columns[0]   # UI显示是"序号"

            # 匹配模板表名
            template_table_mapping = {
                "产品设计活动表_产品标准数据表": "产品标准数据模板表",
                "产品设计活动表_设计数据表": "设计数据模板表",
                "产品设计活动表_通用数据表": "通用数据模板表"
            }
            template_table_name = template_table_mapping.get(table_name, "")

            # 获取模板字段列表（用于对比）
            template_compare_fields = []
            if template_table_name:
                cursor.execute(f"DESCRIBE 产品条件库.{template_table_name}")
                template_compare_fields = [col['Field'] for col in cursor.fetchall()]
                print(f"[DEBUG] 模板表字段={template_compare_fields}")

            # === 特殊处理：获取viewer实例以访问计算值缓存 ===
            calculated_outer_d = None
            if table_name == "产品设计活动表_通用数据表" and viewer:
                # 如果传入了viewer且有计算值缓存
                try:
                    if hasattr(viewer, "_calculated_outer_diameter"):
                        calculated_outer_d = viewer._calculated_outer_diameter
                except Exception as e:
                    print(f"[保存] 获取计算值缓存失败: {e}")

            for row_idx, row in enumerate(data):
                param_name = row.get(name_column)
                if not param_name:
                    print(f"[DEBUG] 跳过：没有{name_column}")
                    continue

                # === 特殊处理：外径参数 ===
                # 如果当前行是"外径"参数，且界面显示为"—"，但有计算值缓存，则使用计算值
                if param_name == "外径" and calculated_outer_d:
                    # 查找"数值"列
                    value_column = None
                    for col_name in header_columns:
                        if "数值" in col_name or col_name == "数值":
                            value_column = col_name
                            break
                    if value_column:
                        current_value = str(row.get(value_column, "")).strip()
                        if current_value == "—" and calculated_outer_d:
                            # 使用计算值替换"—"
                            row[value_column] = calculated_outer_d
                            print(f"[保存外径] 界面显示=—, 保存计算值={calculated_outer_d}")

                # 获取模板数据行
                template = None
                if template_table_name:
                    cursor.execute(
                        f"SELECT * FROM 产品条件库.{template_table_name} WHERE `{name_column}` = %s",
                        (param_name,)
                    )
                    template = cursor.fetchone()
                # 判断是否与模板数据有差异（更改状态）
                def is_changed(template_row, current_row):
                    if not template_row:
                        return True
                    for key in header_columns:
                        if key not in template_compare_fields:
                            continue
                        cur_val = str(current_row.get(key, "")).strip()
                        tpl_val = str(template_row.get(key, "")).strip()
                        if cur_val != tpl_val:
                            return True
                    return False

                change_detected = is_changed(template, row)

                if is_from_design_lib:
                    # UPDATE 操作
                    cursor.execute(
                        f"SELECT * FROM {table_name} WHERE 产品ID = %s AND `{name_column}` = %s",
                        (product_id, param_name)
                    )
                    existing = cursor.fetchone()
                    if existing:
                        update_values = {}
                        for key in header_columns:
                            new_val = row.get(key, "")
                            old_val = existing.get(key, "")
                            if str(new_val) != str(old_val):
                                update_values[key] = new_val
                        if update_values:
                            update_values[change_status_column] = change_detected
                            update_set = ', '.join([f"`{k}` = %s" for k in update_values])
                            sql = f"UPDATE {table_name} SET {update_set} WHERE 产品ID = %s AND `{name_column}` = %s"
                            cursor.execute(sql, tuple(update_values.values()) + (product_id, param_name))
                else:
                    # INSERT 操作
                    insert_row = {}
                    for field in [col['Field'] for col in table_columns if col['Extra'] != "auto_increment"]:
                        if field == "产品ID":
                            insert_row[field] = product_id
                        elif field == param_id_field:
                            # ⚠️ 注意：这里要看 row 里到底有没有"序号"
                            insert_row[field] = row.get(param_id_column, "")
                        elif field == change_status_column:
                            insert_row[field] = change_detected
                        else:
                            insert_row[field] = row.get(field, "")

                    print(f"[DEBUG] insert_row={insert_row}")

                    columns = ', '.join(f"`{k}`" for k in insert_row)
                    placeholders = ', '.join(['%s'] * len(insert_row))
                    # 当相同 (产品ID + 参数ID/名称) 已存在时，执行更新，避免 PRIMARY KEY 冲突
                    update_set = ', '.join([f"`{k}`=VALUES(`{k}`)" for k in insert_row.keys() if k not in ("产品ID", param_id_field)])
                    sql = (
                        f"INSERT INTO {table_name} ({columns}) VALUES ({placeholders}) "
                        f"ON DUPLICATE KEY UPDATE {update_set}"
                    )
                    cursor.execute(sql, tuple(insert_row.values()))

        connection.commit()

    finally:
        connection.close()

def save_coating_table_to_database(table_widget: QTableWidget, table_name, product_id: int, source_status: str):
    """
    保存涂漆数据至【产品设计活动表_涂漆数据表】
    - 如果数据来源为条件模板，则执行 INSERT
    - 如果来源为设计活动库，则执行 UPDATE（根据 产品ID + 参数ID 匹配）
    """
    connection = get_connection(**db_config_2)

    try:
        with connection.cursor() as cursor:
            # ✅ 获取执行标准/规范（表格第0行第2列）
            exec_std_item = table_widget.item(0, 2)
            exec_std = exec_std_item.text().strip() if exec_std_item else ""

            id_counter = 1  # 参数ID，从1开始

            row_count = table_widget.rowCount()
            current_row = 2

            while current_row < row_count:
                # ✅ 当前组用途
                usage_item = table_widget.item(current_row, 0)
                current_usage = usage_item.text().strip() if usage_item else ""

                # ✅ 合并列提取：面积、备注
                paint_area_item = table_widget.item(current_row, 5)
                comment_item = table_widget.item(current_row, 6)
                group_paint_area = paint_area_item.text().strip() if paint_area_item else ""
                group_comment = comment_item.text().strip() if comment_item else ""

                sub_row = current_row
                while sub_row < row_count:
                    usage_item_sub = table_widget.item(sub_row, 0)
                    sub_usage = usage_item_sub.text().strip() if usage_item_sub else ""
                    if sub_row != current_row and sub_usage != current_usage:
                        break  # 下一组开始

                    # ✅ 各字段
                    subtype = table_widget.item(sub_row, 1).text().strip() if table_widget.item(sub_row, 1) else ""
                    category = table_widget.item(sub_row, 2).text().strip() if table_widget.item(sub_row, 2) else ""
                    color = table_widget.item(sub_row, 3).text().strip() if table_widget.item(sub_row, 3) else ""
                    thickness = table_widget.item(sub_row, 4).text().strip() if table_widget.item(sub_row, 4) else ""
                    full_usage = f"{current_usage}_{subtype}" if subtype else current_usage

                    if source_status == "条件模板":
                        # ✅ 插入或更新，避免重复主键导致失败
                        cursor.execute(f"""
                            INSERT INTO {table_name} (
                                `涂漆数据参数ID`, `产品ID`, `用途`, `油漆类别`, `颜色`,
                                `干膜厚度（μm）`, `涂漆面积`, `备注`
                            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                            ON DUPLICATE KEY UPDATE
                                `用途`=VALUES(`用途`),
                                `油漆类别`=VALUES(`油漆类别`),
                                `颜色`=VALUES(`颜色`),
                                `干膜厚度（μm）`=VALUES(`干膜厚度（μm）`),
                                `涂漆面积`=VALUES(`涂漆面积`),
                                `备注`=VALUES(`备注`)
                        """, (
                            id_counter,
                            product_id,
                            full_usage,
                            category,
                            color,
                            thickness,
                            group_paint_area,
                            group_comment
                        ))

                    else:  # 来源为“设计活动库” → UPDATE
                        cursor.execute(f"""
                            UPDATE {table_name}
                            SET `用途` = %s,
                                `油漆类别` = %s,
                                `颜色` = %s,
                                `干膜厚度（μm）` = %s,
                                `涂漆面积` = %s,
                                `备注` = %s
                            WHERE `涂漆数据参数ID` = %s AND `产品ID` = %s
                        """, (
                            full_usage,
                            category,
                            color,
                            thickness,
                            group_paint_area,
                            group_comment,
                            id_counter,
                            product_id
                        ))

                    id_counter += 1
                    sub_row += 1

                current_row = sub_row

        connection.commit()

    finally:
        connection.close()

def save_trail_table_to_database(table_widget: QTableWidget, table_name: str, product_id: int, source_status: str):
    """
    保存无损检测数据至【产品设计活动表_无损检测数据表】
    - 支持条件模板插入 or 设计活动库更新
    - 接头种类为合并分组列（需展开）
    - 表格格式为：检测方法、壳程（3列）、管程（3列）
    """
    connection = get_connection(**db_config_2)

    try:
        with connection.cursor() as cursor:
            # ✅ 递增参数id
            id_counter = 1

            row_count = table_widget.rowCount()
            current_row = 2  # 数据从第2行开始（前2行为表头）

            while current_row < row_count:
                # ✅ 获取分组字段：接头种类（合并项）
                joint_type_item = table_widget.item(current_row, 0)
                current_joint_type = joint_type_item.text().strip() if joint_type_item else ""

                sub_row = current_row
                while sub_row < row_count:
                    # 判断是否是新组
                    if sub_row != current_row:
                        joint_type_check = table_widget.item(sub_row, 0)
                        if joint_type_check and joint_type_check.text().strip():
                            break

                    # ✅ 提取每一行字段
                    detect_method = table_widget.item(sub_row, 1).text().strip() if table_widget.item(sub_row, 1) else ""

                    shell_tech = table_widget.item(sub_row, 2).text().strip() if table_widget.item(sub_row, 2) else ""
                    shell_ratio = table_widget.item(sub_row, 3).text().strip() if table_widget.item(sub_row, 3) else ""
                    shell_level = table_widget.item(sub_row, 4).text().strip() if table_widget.item(sub_row, 4) else ""

                    tube_tech = table_widget.item(sub_row, 5).text().strip() if table_widget.item(sub_row, 5) else ""
                    tube_ratio = table_widget.item(sub_row, 6).text().strip() if table_widget.item(sub_row, 6) else ""
                    tube_level = table_widget.item(sub_row, 7).text().strip() if table_widget.item(sub_row, 7) else ""

                    if source_status == "条件模板":
                        # ✅ INSERT ... ON DUPLICATE KEY UPDATE（避免重复主键报错）
                        cursor.execute(f"""
                            INSERT INTO {table_name} (
                                `无损检测数据参数ID`, `产品ID`, `接头种类`, `检测方法`,
                                `壳程_技术等级`, `壳程_检测比例`, `壳程_合格级别`,
                                `管程_技术等级`, `管程_检测比例`, `管程_合格级别`
                            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                            ON DUPLICATE KEY UPDATE
                                `接头种类`=VALUES(`接头种类`),
                                `检测方法`=VALUES(`检测方法`),
                                `壳程_技术等级`=VALUES(`壳程_技术等级`),
                                `壳程_检测比例`=VALUES(`壳程_检测比例`),
                                `壳程_合格级别`=VALUES(`壳程_合格级别`),
                                `管程_技术等级`=VALUES(`管程_技术等级`),
                                `管程_检测比例`=VALUES(`管程_检测比例`),
                                `管程_合格级别`=VALUES(`管程_合格级别`)
                        """, (
                            id_counter,
                            product_id,
                            current_joint_type,
                            detect_method,
                            shell_tech, shell_ratio, shell_level,
                            tube_tech, tube_ratio, tube_level
                        ))
                    else:
                        # ✅ UPDATE 更新
                        cursor.execute(f"""
                            UPDATE {table_name}
                            SET `接头种类` = %s,
                                `检测方法` = %s,
                                `壳程_技术等级` = %s,
                                `壳程_检测比例` = %s,
                                `壳程_合格级别` = %s,
                                `管程_技术等级` = %s,
                                `管程_检测比例` = %s,
                                `管程_合格级别` = %s
                            WHERE `无损检测数据参数ID` = %s AND `产品ID` = %s
                        """, (
                            current_joint_type,
                            detect_method,
                            shell_tech, shell_ratio, shell_level,
                            tube_tech, tube_ratio, tube_level,
                            id_counter,
                            product_id
                        ))

                    id_counter += 1
                    sub_row += 1

                current_row = sub_row

        connection.commit()

    finally:
        connection.close()


def sync_design_params_to_element_params(product_id):

    # ✅ 1. 获取腐蚀裕量
    conn = get_connection(**db_config_2)
    try:
        with conn.cursor() as cursor:
            cursor.execute("""
                SELECT 参数名称, 管程数值, 壳程数值
                FROM 产品设计活动表_设计数据表
                WHERE 产品ID = %s
            """, (product_id,))
            rows = cursor.fetchall()
            ca_map = {row["参数名称"].strip(): row for row in rows}
    finally:
        conn.close()

    tube_ca = ca_map.get("腐蚀裕量*", {}).get("管程数值", "")
    shell_ca = ca_map.get("腐蚀裕量*", {}).get("壳程数值", "")

    # ✅ 2. 判断当前产品是否有元件材料
    conn1 = get_connection(**db_config_2)
    try:
        with conn1.cursor() as cur:
            cur.execute(
                "SELECT 1 FROM 产品设计活动表_元件附加参数表 WHERE 产品ID=%s LIMIT 1",
                (product_id,)
            )
            exists = cur.fetchone() is not None
    finally:
        conn1.close()

    if not exists:
        # 这个产品还没有元件附加参数，不做批量写回
        return

    # ✅ 3. 搜集当前产品下的垫片
    conn1 = get_connection(**db_config_2)
    try:
        with conn1.cursor() as cur:
            cur.execute("""
                    SELECT 元件名称, 参数名称, 参数值
                    FROM 产品设计活动表_元件附加参数表
                    WHERE 产品ID = %s
                      AND 元件名称 LIKE %s
                      AND 参数名称 IN ('垫片标准','垫片类型')
                """, (product_id, "%垫片%"))
            rows = cur.fetchall() or []
    finally:
        conn1.close()

    gaskets = {}
    for r in rows:
        en = (r.get("元件名称") or "").strip()
        pnam = (r.get("参数名称") or "").strip()
        pval = (r.get("参数值") or "").strip()
        if not en:
            continue
        info = gaskets.setdefault(en, {"name": "", "standard": "", "type": ""})
        if pnam == "垫片名称" and pval:
            info["name"] = pval
        elif pnam == "垫片标准" and pval:
            info["standard"] = pval
        elif pnam == "垫片类型" and pval:
            info["type"] = pval

    # ✅ 4. 垫片数值写回数据库
    def _norm_out(v: str) -> str:
        """字段值兜底：空/None -> '程序推荐'"""
        if v is None: return "程序推荐"
        s = str(v).strip()
        return s if s else "程序推荐"

    for element_name, meta in gaskets.items():
        gasket_name = meta["name"] or element_name  # 名称缺省用元件名
        gasket_standard = meta["standard"]
        gasket_type = meta["type"]

        try:
            spec = resolve_gasket_dimensions(
                product_id=product_id,
                gasket_name=gasket_name,
                gasket_standard=gasket_standard,
                gasket_type=gasket_type
            )
        except Exception as e:
            # 任何异常均写“程序推荐”
            update_element_name_data(product_id, element_name, "垫片名义外径D2n", "程序推荐")
            update_element_name_data(product_id, element_name, "垫片名义内径D1n", "程序推荐")
            update_element_name_data(product_id, element_name, "环内径d1", "程序推荐")
            continue

        # 命中情况下，有些字段可能仍为空 -> 单字段兜底为“程序推荐”
        if not spec.get("nonstd", False):
            update_element_name_data(product_id, element_name, "垫片名义外径D2n", _norm_out(spec.get("外直径D")))
            update_element_name_data(product_id, element_name, "垫片名义内径D1n", _norm_out(spec.get("内直径d")))
            update_element_name_data(product_id, element_name, "环内径d1", _norm_out(spec.get("环内径d1")))
        else:
            update_element_name_data(product_id, element_name, "垫片名义外径D2n", "程序推荐")
            update_element_name_data(product_id, element_name, "垫片名义内径D1n", "程序推荐")
            update_element_name_data(product_id, element_name, "环内径d1", "程序推荐")


    # ✅ 5. 腐蚀裕量写入数据库
    if tube_ca:
        update_element_name_data(product_id, "固定管板", "管程侧腐蚀裕量", str(tube_ca))
        update_element_name_data(product_id, "浮动管板", "管程侧腐蚀裕量", str(tube_ca))
        update_element_name_data(product_id, "球冠形封头", "管程侧腐蚀裕量", str(tube_ca))

    if shell_ca:
        update_element_name_data(product_id, "固定管板", "壳程侧腐蚀裕量", str(shell_ca))
        update_element_name_data(product_id, "浮动管板", "壳程侧腐蚀裕量", str(shell_ca))
        update_element_name_data(product_id, "球冠形封头", "壳程侧腐蚀裕量", str(shell_ca))



def sync_corrosion_to_guankou_param(product_id, guankou_codes, category_label=None):
    """
    将条件输入（设计数据表）的腐蚀裕量同步到管口参数：
    - case1: 当管程/壳程腐蚀裕量数值相同 → 用该值填写管口3列默认值
    - case2: 如果管口号都属于管程或壳程 → 用对应的腐蚀裕量值填写
    - case3: 以上两种情况都不满足时 → 不填，保持为空
    """

    # 1. 从条件输入获取腐蚀裕量
    ca_map = get_design_params_by_product_id(product_id)
    tube_ca = ca_map.get("腐蚀裕量*", {}).get("管程数值", "")
    shell_ca = ca_map.get("腐蚀裕量*", {}).get("壳程数值", "")

    # 如果没有腐蚀裕量则跳过
    if not tube_ca and not shell_ca:
        print("[跳过] 条件输入没有腐蚀裕量")
        return

    print(f"[调试] tube_ca={tube_ca} ({type(tube_ca)}), shell_ca={shell_ca} ({type(shell_ca)})")

    # === case1: 管壳程腐蚀裕量相同 ===
    if tube_ca and shell_ca and str(tube_ca) == str(shell_ca):
        update_guankou_param_flex_db(product_id, "接管腐蚀裕量", str(tube_ca), tab_name=category_label)
        print(f"[case1] 管壳程腐蚀裕量相同，写入默认值 {tube_ca}")
        return

    # === case2: 管壳程腐蚀裕量不同 ===
    if guankou_codes:
        # 获取所有管口所属（管程/壳程）
        affiliations = [query_guankou_affiliation(product_id, code) for code in guankou_codes]

        # 如果所有管口号都属于管程
        if all(a == "管程" for a in affiliations if a):
            if tube_ca:
                update_guankou_param_flex_db(product_id, "接管腐蚀裕量", str(tube_ca), tab_name=category_label)
                print(f"[case2] {guankou_codes} 都属于管程，写入 {tube_ca}")
            else:
                update_guankou_param_flex_db(product_id, "接管腐蚀裕量", "", tab_name=category_label)
                print(f"[case2] {guankou_codes} 管程腐蚀裕量为空，留空")

        # 如果所有管口号都属于壳程
        elif all(a == "壳程" for a in affiliations if a):
            if shell_ca:
                update_guankou_param_flex_db(product_id, "接管腐蚀裕量", str(shell_ca), tab_name=category_label)
                print(f"[case2] {guankou_codes} 都属于壳程，写入 {shell_ca}")
            else:
                update_guankou_param_flex_db(product_id, "接管腐蚀裕量", "", tab_name=category_label)
                print(f"[case2] {guankou_codes} 壳程腐蚀裕量为空，留空")

        # 如果管口号的所属元件既有管程也有壳程
        else:
            update_guankou_param_flex_db(product_id, "接管腐蚀裕量", "", tab_name=category_label)
            print(f"[case3] {guankou_codes} 管口号所属元件不同，留空")

    else:
        # 如果没有管口号且腐蚀裕量不同，保持为空
        update_guankou_param_flex_db(product_id, "接管腐蚀裕量", "", tab_name=category_label)
        print("[case3] 没有管口号且腐蚀裕量不同，留空")

def save_all_tables(viewer, product_id):
    """
    保存所有表格数据（标准、设计、通用、涂漆、无损检测）至数据库
    """
    try:
        if not product_id:
            QMessageBox.warning(viewer, "产品ID无效", "产品ID不能为空")
            return

        is_from_design_lib = viewer.design_data_source == "设计活动库"

        # 提取数据并保存到各自表
        save_data_to_database(
            get_table_data(viewer.tableWidget_product_std),
            product_id,
            "产品设计活动表_产品标准数据表",
            viewer.tableWidget_product_std,
            is_from_design_lib
        )

        save_data_to_database(
            get_table_data(viewer.tableWidget_design_data),
            product_id,
            "产品设计活动表_设计数据表",
            viewer.tableWidget_design_data,
            is_from_design_lib
        )

        sync_design_params_to_element_params(product_id)

        # 1124新修改-保存时增加元件定义腐蚀余量同步
        try:
            labels = query_all_guankou_categories(product_id) or ["管口材料分类1"]
            for label in labels:
                codes = query_guankou_codes(product_id, label) or []
                sync_corrosion_to_guankou_param(product_id, codes, label)
        except Exception as e:
            print(f"[警告] 设计数据保存后的腐蚀裕量同步失败: {e}")

        save_data_to_database(
            get_table_data(viewer.tableWidget_general_data),
            product_id,
            "产品设计活动表_通用数据表",
            viewer.tableWidget_general_data,
            is_from_design_lib,
            viewer=viewer  # 传递viewer实例以便访问计算值缓存
        )

        save_coating_table_to_database(
            viewer.tableWidget_coating_data,
            "产品设计活动表_涂漆数据表",
            product_id,
            viewer.design_data_source
        )

        save_trail_table_to_database(
            viewer.tableWidget_trail_data,
            "产品设计活动表_无损检测数据表",
            product_id,
            viewer.design_data_source
        )
        viewer.design_data_source = "设计活动库"
        try:
            invalidate_caches_for_product(product_id)
        except Exception as e:
            print(f"[警告] 条件输入保存后的缓存失效失败: {e}")
        try:
            clear_all_pn_user_input_for_product(product_id)
            force_recompute_and_update_pn(product_id)
        except Exception as e:
            print(f"[警告] 条件输入保存后的PN刷新失败: {e}")
    except Exception as e:
        QMessageBox.critical(viewer, "保存失败", f"保存数据时发生错误：{str(e)}")

"""保存前检查必填项"""
def validate_required_fields(table_widget, mode="设计数据"):
    """
    检查带星号的“参数名称”对应的必填字段是否为空
    - mode="设计数据"：要求壳程数值、管程数值必须填写
    - mode="通用数据"：要求参数值必须填写
    - 特殊强制：进、出口压力差 的管程数值为必填
    """
    required_col_name = {
        "设计数据": ["壳程数值", "管程数值"],
        "通用数据": ["数值"]
    }

    header_map = {}
    for col in range(table_widget.columnCount()):
        item = table_widget.horizontalHeaderItem(col)
        if item:
            header_map[item.text()] = col

    name_col = header_map.get("参数名称")
    if name_col is None:
        return False, []

    required_cols = [header_map.get(cn) for cn in required_col_name[mode] if cn in header_map]

    missing_rows = []

    for row in range(table_widget.rowCount()):
        name_item = table_widget.item(row, name_col)
        if not name_item:
            continue
        name_text = name_item.text().strip()

        # ✅ 常规：带 * 的参数检查 修改！！！！
        if "*" in name_text:
            # 特殊项：进、出口压力差* 只检查管程数值
            if name_text == "进、出口压力差*":
                col = header_map.get("管程数值")
                if col is not None:
                    val_item = table_widget.item(row, col)
                    if not val_item or not val_item.text().strip():
                        missing_rows.append((row, f"{name_text}（管程）"))
            else:
                for col in required_cols:
                    val_item = table_widget.item(row, col)
                    if not val_item or not val_item.text().strip():
                        missing_rows.append((row, name_text))
                        break

        # ✅ 强制补充项：进、出口压力差 的“管程数值”必须填写
        if mode == "设计数据" and name_text == "进、出口压力差":
            col = header_map.get("管程数值")
            if col is not None:
                val_item = table_widget.item(row, col)
                if not val_item or not val_item.text().strip():
                    missing_rows.append((row, name_text + "（管程）"))

    return len(missing_rows) > 0, missing_rows

"""高亮未填项"""
def highlight_missing_required_rows(table_widget: QTableWidget, missing_info: list):
    """
    高亮缺失值的行（浅蓝色），并恢复非缺失行为交替背景色。
    使用 Qt.UserRole+1 标记缺失行。
    """
    for row in range(table_widget.rowCount()):
        for col in range(table_widget.columnCount()):
            item = table_widget.item(row, col)
            if item:
                # 清除旧标记
                item.setData(Qt.UserRole + 1, None)

                # 恢复交替颜色
                if row % 2 == 0:
                    item.setBackground(QColor("#ffffff"))
                else:
                    item.setBackground(QColor("#f0f0f0"))

    # 设置缺失行背景并添加标记
    for row_idx, _ in missing_info:
        for col in range(table_widget.columnCount()):
            item = table_widget.item(row_idx, col)
            if item:
                item.setBackground(QColor("#90d7ec"))  # 浅蓝色
                item.setData(Qt.UserRole + 1, "missing")  # ✅ 标记为缺失

"""参数值类型限制，关联限制"""
def safe_set_text_and_color(widget, text, color=None):
    if hasattr(widget, "setText"):
        widget.setText(text)
        if hasattr(widget, "setToolTip"):
            widget.setToolTip(text)  # ✅ 加这一行
    if isinstance(widget, QWidget) and color:
        widget.setStyleSheet(f"color: {color};")

def validate_design_table_cell(param_name: str, column_name: str, value: str, line_edit_widget, table_widget=None, col_index=None) -> bool:
    """
    主入口函数，负责分派规则函数
    - 返回值：校验结果等级 "ok" / "warn" / "error"
    """

    param_name = param_name.strip()
    column_name = column_name.strip()
    key = (param_name, column_name)

    # ✅ 用户主动清空时，允许为空（后续由“是否必填”统一校验）
    if value.strip() == "":
        safe_set_text_and_color(line_edit_widget, "", "black")
        return "ok"

    try:
        # ✅ 自定义规则表（check_xxx）
        custom_rules = {
            ("公称直径*", "壳程数值"): check_dn,
            ("公称直径*", "管程数值"): check_dn,
            ("工作压力", "壳程数值"): check_work_pressure,
            ("工作压力", "管程数值"): check_work_pressure,
            ("工作温度（入口）", "壳程数值"): check_work_temp_in,
            ("工作温度（入口）", "管程数值"): check_work_temp_in,
            ("工作温度（出口）", "壳程数值"): check_work_temp_out,
            ("工作温度（出口）", "管程数值"): check_work_temp_out,
            ("最高允许工作压力", "壳程数值"): check_work_pressure_max,
            ("最高允许工作压力", "管程数值"): check_work_pressure_max,
            ("管板设计压差", "壳程数值"): check_tubeplate_design_pressure_gap,
            ("管板设计压差", "管程数值"): check_tubeplate_design_pressure_gap,
            ("设计压力*", "壳程数值"): check_design_pressure,
            ("设计压力*", "管程数值"): check_design_pressure,
            ("设计温度（最高）*", "壳程数值"): check_design_temp_max,
            ("设计温度（最高）*", "管程数值"): check_design_temp_max,
            ("最低设计温度", "壳程数值"): check_design_temp_min,
            ("最低设计温度", "管程数值"): check_design_temp_min,
            ("进、出口压力差*", "壳程数值"): check_in_out_pressure_gap,
            ("进、出口压力差*", "管程数值"): check_in_out_pressure_gap,
            ("自定义耐压试验压力（卧）", "壳程数值"): check_def_trail_stand_pressure_lying,
            ("自定义耐压试验压力（卧）", "管程数值"): check_def_trail_stand_pressure_lying,
            ("自定义耐压试验压力（立）", "壳程数值"): check_def_trail_stand_pressure_stand,
            ("自定义耐压试验压力（立）", "管程数值"): check_def_trail_stand_pressure_stand,
            ("耐压试验介质密度", "壳程数值"): check_trail_stand_pressure_medium_density,
            ("耐压试验介质密度", "管程数值"): check_trail_stand_pressure_medium_density,
            ("绝热层厚度", "壳程数值"): check_insulation_layer_thickness,
            ("绝热层厚度", "管程数值"): check_insulation_layer_thickness,
            ("绝热材料密度", "壳程数值"): check_insulation_material_density,
            ("绝热材料密度", "管程数值"): check_insulation_material_density,
            ("耐压试验类型*", "壳程数值"): check_trail_stand_pressure_type,
            ("耐压试验类型*", "管程数值"): check_trail_stand_pressure_type,
            ("耐压试验温度", "壳程数值"): check_pressure_test_temp,
            ("耐压试验温度", "管程数值"): check_pressure_test_temp,
            # ("沿长度平均的换热管金属温度*", "壳程数值"): check_avg_tube_metal_temp,
            ("沿长度平均的换热管金属温度*", "管程数值"): check_avg_tube_metal_temp,
            ("沿长度平均的壳程圆筒金属温度*", "壳程数值"): check_avg_shell_metal_temp
            # ("沿长度平均的壳程圆筒金属温度*", "管程数值"): check_avg_shell_metal_temp

        }

        # ✅ 通用规则（基础类型/范围检查）
        base_rules = {
            ("介质密度", "壳程数值"): ("float", (0, 1e10), "介质密度的参数值不能为负，请核对后输入"),
            ("介质密度", "管程数值"): ("float", (0, 1e10), "介质密度的参数值不能为负，请核对后输入"),
            ("介质入口流速", "壳程数值"): ("float", (0, 1e10), "介质入口流速的参数值不能为负，请核对后输入"),
            ("介质入口流速", "管程数值"): ("float", (0, 1e10), "介质入口流速的参数值不能为负，请核对后输入"),
            ("液柱静压力", "壳程数值"): ("float", (0, 1e10), "液柱静压力的参数值不能为负，请核对后输入"),
            ("液柱静压力", "管程数值"): ("float", (0, 1e10), "液柱静压力的参数值不能为负，请核对后输入"),
            ("腐蚀裕量*", "壳程数值"): ("float", (0, 1e10), "腐蚀裕量的参数值不能为负，请核对后输入"),
            ("腐蚀裕量*", "管程数值"): ("float", (0, 1e10), "腐蚀裕量的参数值不能为负，请核对后输入")
        }

        print(f"[校验函数] param={param_name}, col={column_name}, value='{value}'")

        if key in custom_rules:
            result, msg = custom_rules[key](value, line_edit_widget, param_name, column_name, table_widget, col_index)
            if result == "ok":
                safe_set_text_and_color(line_edit_widget, "", "black")
            elif result == "warn":
                safe_set_text_and_color(line_edit_widget, msg, "orange")
            elif result == "error":
                safe_set_text_and_color(line_edit_widget, msg, "red")
            return result

        if key in base_rules:
            try:
                dtype, limits, msg = base_rules[key]
                if dtype == "int":
                    num = int(value)
                elif dtype == "float":
                    num = float(value)
                else:
                    safe_set_text_and_color(line_edit_widget, "输入数据类型有误，请确认后输入", "red")
                    return "error"
                if limits:
                    min_v, max_v = limits
                    if not (min_v <= num <= max_v):
                        safe_set_text_and_color(line_edit_widget, msg, "red")
                        return "error"
                safe_set_text_and_color(line_edit_widget, "", "black")
                return "ok"
            except Exception:
                safe_set_text_and_color(line_edit_widget, "校验异常，请确认输入", "red")
                return "error"

        return "ok"

    except Exception:
        safe_set_text_and_color(line_edit_widget, "校验异常，请确认输入", "red")
        return "error"

def validate_general_table_cell(param_name: str, value: str, line_edit_widget, table_widget=None) -> str:
    """
    通用数据表 校验入口函数
    - param_name: 参数名称
    - value: 用户输入的参数值（字符串）
    - line_edit_widget: QLineEdit 显示提示
    - 返回值: 校验等级 "ok" / "warn" / "error"
    """

    param_name = param_name.strip()

    # ✅ 主动清空，允许通过
    if value.strip() == "":
        safe_set_text_and_color(line_edit_widget, "", "black")  # ✅ 正确
        return "ok"

    try:
        # ✅ 自定义规则（check_xxx 通常联动或复杂校验）
        custom_rules = {
            # ("参数名称",): check_xxx,
        }

        # ✅ 通用规则（类型 + 范围判断）
        base_rules = {
            ("设计使用年限*",): ("int", (0, 1e10), "设计使用年限不能为负，请核对后输入"),
            ("基本风压",): ("float", (0, 1e10), "基本风压值不能为负，请核对后输入"),
            ("雪压值",): ("float", (0, None), "雪压值不能为负，请核对后输入"),
            # ... 继续补充更多通用项
        }

        key = (param_name,)

        # ✅ 优先匹配自定义规则
        if key in custom_rules:
            result, msg = custom_rules[key](value, line_edit_widget, param_name, table_widget)
            if result == "ok":
                safe_set_text_and_color(line_edit_widget, "", "black")
            elif result == "warn":
                safe_set_text_and_color(line_edit_widget, msg, "orange")
            elif result == "error":
                safe_set_text_and_color(line_edit_widget, msg, "red")
            return result  # "ok" / "warn" / "error"

        # ✅ 通用处理
        if key in base_rules:
            dtype, limits, msg = base_rules[key]

            # 🧠 第一层：手动类型转换错误提示
            try:
                if dtype == "int":
                    num = int(value)
                elif dtype == "float":
                    num = float(value)
                else:
                    safe_set_text_and_color(line_edit_widget, "输入数据类型有误，请确认后输入", "red")
                    return "error"
            except ValueError:
                safe_set_text_and_color(line_edit_widget, "输入数据类型有误，请确认后输入", "red")
                return "error"

            # 🧠 第二层：其他逻辑错误
            try:
                if limits:
                    min_v, max_v = limits
                    if (min_v is not None and num < min_v) or (max_v is not None and num > max_v):
                        safe_set_text_and_color(line_edit_widget, msg, "red")
                        return "error"

                safe_set_text_and_color(line_edit_widget, "", "black")
                return "ok"

            except Exception:
                safe_set_text_and_color(line_edit_widget, "校验异常，请确认输入", "red")
                return "error"

        return "ok"  # 无匹配项默认通过

    except Exception as e:
        safe_set_text_and_color(line_edit_widget, "校验异常，请确认输入", "red")
        return "error"

def validate_trail_table_cell(column_name: str, value: str, tip_widget, table_widget=None, row_index: int = None) -> str:
    """
    检测数据表 - 通用列校验器（仅对“检测比例”做范围检查）
    特殊规则：
    - 若接头类型=T（管头），检测方法=R.T./P.T.，则壳程三列仅允许“/”
    """
    val = value.strip()

    # === 特殊规则优先处理 ===
    if table_widget and row_index is not None:
        try:
            # ✅ 从 UserRole 取接头种类
            jt_type = ""
            if table_widget.item(row_index, 2):  # 任意一列取 UserRole
                jt_type = table_widget.item(row_index, 2).data(Qt.UserRole) or ""

            method_item = table_widget.item(row_index, 1)
            method = method_item.text().strip() if method_item else ""

            if jt_type == "T（管头）" and method in ["R.T.", "P.T."] \
               and (column_name.startswith("壳程_技术等级")
                    or column_name.startswith("壳程_合格级别")
                    or column_name.startswith("壳程_检测比例")):
                if val == "/":
                    safe_set_text_and_color(tip_widget, "", "black")
                    return "ok"
                else:
                    safe_set_text_and_color(tip_widget, "此处仅允许 '/'", "red")
                    return "error"
        except Exception as e:
            print(f"[validate_trail_table_cell][DEBUG] 特殊规则异常: {e}")

    # === 通用规则 ===
    if val == "":
        safe_set_text_and_color(tip_widget, "", "black")
        return "ok"

    if not re.search(r"检测比例[%]?$", column_name):
        return "ok"

    pattern = r"^(≥|>)?\d{1,3}$"
    if not re.match(pattern, val):
        safe_set_text_and_color(tip_widget, "请输入合法格式，如 50，≥30 或 >20", "red")
        return "error"

    try:
        num_part = int(re.sub(r"[^\d]", "", val))
        if not (0 <= num_part <= 100):
            safe_set_text_and_color(tip_widget, "检测比例应在 0 ~ 100 之间，请核对后输入", "red")
            return "error"
    except Exception:
        safe_set_text_and_color(tip_widget, "检测比例格式异常", "red")
        return "error"

    safe_set_text_and_color(tip_widget, "", "black")
    return "ok"

def validate_coating_table_cell(column_name: str, value: str, tip_widget, table_widget=None) -> str:
    """
    涂漆数据表 校验器
    - 针对：干膜厚度（μm）、涂漆面积 两列进行校验
    """
    if value.strip() == "":
        safe_set_text_and_color(tip_widget, "", "black")
        return "ok"

    val = value.strip()

    # ✅ 如果列名像“列5”，说明未传入真实逻辑列头 → 尝试自己查
    if column_name.startswith("列") and table_widget and hasattr(table_widget, "logical_headers"):
        try:
            col_index = int(column_name.replace("列", ""))
            column_name = table_widget.logical_headers[col_index]
        except Exception:
            # 万一列号非法，直接跳过
            return "ok"

    col = column_name.strip()

    if col not in ["干膜厚度（μm）", "涂漆面积"]:
        return "ok"  # 其他列无需校验

    try:
        num = float(val)
    except ValueError:
        safe_set_text_and_color(tip_widget, "输入数据类型有误，请确认后输入", "red")
        return "error"

    if num <= 0:
        safe_set_text_and_color(tip_widget, f"{col}必须为正数，请核对后输入", "red")
        return "error"

    safe_set_text_and_color(tip_widget, "", "black")
    return "ok"

def dispatch_cell_validation(viewer, table, row, col, param_name, column_name, value, *args, **kwargs):
    print(f"[调试] dispatch_cell_validation: col={column_name}, value={value}")

    mode = getattr(table, "validation_mode", "design")

    if value.strip() == "":
        safe_set_text_and_color(viewer.line_tip, "", "black")
        return "ok"

    if mode == "design":
        return validate_design_table_cell(param_name, column_name, value, viewer.line_tip, table, col)

    elif mode == "general":
        if column_name != "数值":
            safe_set_text_and_color(viewer.line_tip, "", "black")
            return "ok"
        return validate_general_table_cell(param_name, value, viewer.line_tip, table)


    elif mode == "trail":
        result = validate_trail_table_cell(column_name, value, viewer.line_tip, table, row_index=row)

        if result == "error":
            return result

        item = table.item(row, col)
        if item:
            default_val = item.data(Qt.UserRole + 2)
            if default_val:
                if column_name.endswith("技术等级") and is_grade_lower(value, default_val):
                    msg = "技术等级不能低于默认值，请核对后输入"
                    safe_set_text_and_color(viewer.line_tip, msg, "red")
                    if hasattr(viewer, "import_tip_list"):
                        viewer.import_tip_list.append(f"[检测数据] 第{row - 1}行 - {column_name}: ❌ {msg}")

                    QTimer.singleShot(0, lambda: table.item(row, col).setText(""))
                    return "error"
                elif column_name.endswith("合格级别") and is_qualify_lower(value, default_val):
                    msg = "合格级别不能低于默认值，请核对后输入"
                    safe_set_text_and_color(viewer.line_tip, msg, "red")
                    if hasattr(viewer, "import_tip_list"):
                        viewer.import_tip_list.append(f"[检测数据] 第{row - 1}行 - {column_name}: ❌ {msg}")
                    QTimer.singleShot(0, lambda: table.item(row, col).setText(""))
                    return "error"

        safe_set_text_and_color(viewer.line_tip, "", "black")
        return result

    elif mode == "coating":
        return validate_coating_table_cell(column_name, value, viewer.line_tip, table)

    return "ok"


"""参考数据导入相关函数"""

def get_ref_data_excel_path(product_id: int) -> str:
    """
    给定产品ID，查询并返回对应的 条件输入数据表.xlsx 完整路径
    """
    serial = ""
    # ✅ 遍历 product_table_row_status，用 product_id 匹配行
    for row, status in bianl.product_table_row_status.items():
        if isinstance(status, dict):
            if str(status.get("product_id")) == str(product_id):
                serial = status.get("old_serial", "") or f"{row+1:03d}"
                break

    try:
        # 第一步：连接产品需求库，查产品需求表
        connection = get_connection(**db_config_3)
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT `项目ID`, `产品编号`, `产品名称`, `设备位号`
                FROM `产品需求表`
                WHERE `产品ID` = %s
                LIMIT 1
            """, (product_id,))
            product_row = cursor.fetchone()
        connection.close()

        if not product_row:
            raise ValueError(f"未找到产品ID {product_id} 的产品需求信息。")

        project_id = product_row['项目ID']
        product_code = product_row['产品编号']
        product_name = product_row['产品名称']
        device_loc_id = product_row['设备位号']

        # 第二步：连接项目需求库，查项目需求表
        connection = get_connection(**db_config_4)
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT `项目保存路径`,`项目名称`,`业主名称`
                FROM `项目需求表`
                WHERE `项目ID` = %s
                LIMIT 1
            """, (project_id,))
            project_row = cursor.fetchone()
        connection.close()

        if not project_row:
            raise ValueError(f"未找到项目ID {project_id} 的项目信息。")

        project_save_path = project_row['项目保存路径']
        project_path = project_row['项目名称']
        yezhu_path = project_row['业主名称']
        pinjie_path = f"{yezhu_path}_{project_path}"

        # ✅ 拼接文件夹名：序号_产品名称_产品编号_设备位号（自动跳过空值）
        parts = [serial, product_name, device_loc_id, product_code]
        folder_name = "_".join([str(p).strip() for p in parts if p and str(p).strip()])

        full_path = os.path.join(project_save_path, pinjie_path, folder_name, "条件输入数据表.xlsx")

        if not os.path.exists(full_path):
            raise FileNotFoundError(f"未找到文件：{full_path}")

        return full_path

    except Exception as e:
        print(f"[ERROR] get_ref_data_excel_path 出错: {e}")
        raise



def get_user_selected_excel_path(parent_widget=None) -> str:
    """
    弹出文件选择框，获取用户选择的Excel路径
    """
    file_path, _ = QFileDialog.getOpenFileName(
        parent_widget,
        "选择条件输入数据表",
        "",
        "Excel Files (*.xlsx);;All Files (*)"
    )
    if not file_path:
        raise FileNotFoundError("用户未选择文件")
    return file_path

def update_product_standard_table_from_excel(excel_path: str, table_widget):
    """
    从Excel中读取‘产品标准’Sheet，按规范/标准名称匹配，更新界面表格中的‘规范/标准代号’列
    """
    try:
        df = pd.read_excel(excel_path, sheet_name="产品标准", dtype=str)
        df.fillna("", inplace=True)

        # ✅ 构建映射表：规范/标准名称 -> 规范/标准代号
        std_map = {str(k).strip(): str(v).strip() for k, v in zip(df.iloc[:, 1], df.iloc[:, 2])}
        # 注意这里用的是第1列（B列，“规范/标准名称”），不是序号列了！

        for row in range(table_widget.rowCount()):
            name_item = table_widget.item(row, 1)  # 第1列是规范/标准名称
            target_item = table_widget.item(row, 2)  # 第2列是规范/标准代号

            if not name_item or not target_item:
                continue

            name = str(name_item.text()).strip()
            if name in std_map:
                target_item.setText(std_map[name])

    except Exception as e:
        raise RuntimeError(f"导入产品标准失败：{str(e)}")

def update_design_data_table_from_excel(excel_path: str, table_widget):
    """
    从Excel中读取‘设计数据’Sheet，按参数名称匹配，更新‘壳程数值’和‘管程数值’
    如果本地界面中“绝热层类型”是“无”，则跳过对应侧的绝热材料、厚度、密度的导入
    """
    try:
        import pandas as pd
        df = pd.read_excel(excel_path, sheet_name="设计数据", dtype=str)
        df.fillna("", inplace=True)

        # Excel 中构建映射表
        data_map = {
            str(row[1]).strip(): (str(row[3]).strip(), str(row[4]).strip())
            for _, row in df.iterrows()
        }

        # ✅ 获取界面当前的“绝热层类型”值
        insulation_type_shell = ""
        insulation_type_tube = ""
        for row in range(table_widget.rowCount()):
            name_item = table_widget.item(row, 1)
            if name_item and name_item.text().strip() == "绝热类型":
                shell_item = table_widget.item(row, 3)
                tube_item = table_widget.item(row, 4)
                insulation_type_shell = shell_item.text().strip() if shell_item else ""
                insulation_type_tube = tube_item.text().strip() if tube_item else ""
                break

        skip_shell = insulation_type_shell == "无"
        skip_tube = insulation_type_tube == "无"

        print(f"[导入判定] 绝热类型: 壳程={insulation_type_shell}, 管程={insulation_type_tube} | skip_shell={skip_shell}, skip_tube={skip_tube}")

        for row in range(table_widget.rowCount()):
            name_item = table_widget.item(row, 1)
            if not name_item:
                continue

            name = name_item.text().strip()
            if name not in data_map:
                continue

            shell_val, tube_val = data_map[name]

            # 判断是否为绝热项且需要跳过
            if name in {"绝热材料", "绝热层厚度", "绝热材料密度"}:
                if skip_shell:
                    shell_val = ""  # 不导入壳程
                if skip_tube:
                    tube_val = ""  # 不导入管程

            # 更新壳程
            shell_item = table_widget.item(row, 3)
            if shell_item:
                shell_item.setText(shell_val)

            # 更新管程
            tube_item = table_widget.item(row, 4)
            if tube_item:
                tube_item.setText(tube_val)

    except Exception as e:
        raise RuntimeError(f"导入设计数据失败：{str(e)}")

# 已改
def import_multi_conditions_from_excel(excel_path: str, product_id: int, viewer: QWidget):
    """
    导入Excel中的工况2/3小表到数据库，并做校核。
    - 工况2：I/J列（壳程/管程）
    - 工况3：K/L列（壳程/管程）
    从第3行开始
    """

    df = pd.read_excel(excel_path, sheet_name="设计数据", dtype=str, header=None)
    df.fillna("", inplace=True)

    gongkuang_cols = {
        2: (8, 9),   # Excel I=9, J=10 → df 索引=8,9
        3: (10, 11)  # Excel K=11, L=12 → df 索引=10,11
    }

    conn = get_connection(**db_config_2)
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT MAX(设计数据参数ID) AS max_sn FROM 产品设计活动表_设计数据表 WHERE 产品ID=%s", (product_id,))
            max_sn = cur.fetchone()["max_sn"] or 31
            seq = max_sn

            for gk_no, (col_kc, col_gc) in gongkuang_cols.items():
                # === 创建临时弹窗表格用于校核 ===
                from modules.condition_input.funcs.multi_conditions_dialog import MultiConditionsDialog
                dlg = MultiConditionsDialog(parent=viewer, product_id=product_id)
                table = dlg.tableWidget

                for idx in range(2, len(df)):  # 从第3行开始
                    pname_base = str(df.iloc[idx, 6]).strip()  # G列：参数名称
                    if not pname_base:
                        continue

                    kc_val = str(df.iloc[idx, col_kc]).strip()
                    gc_val = str(df.iloc[idx, col_gc]).strip()

                    # === 把 Excel 值写入临时弹窗表格 ===
                    for r in range(table.rowCount()):
                        header_item = table.verticalHeaderItem(r)
                        if header_item and header_item.text().strip() == pname_base:
                            if kc_val:
                                table.setItem(r, 1, QTableWidgetItem(kc_val))
                            if gc_val:
                                table.setItem(r, 2, QTableWidgetItem(gc_val))
                            break

                    # === 校核（调用 dispatch_cell_validation） ===
                    # 注意：多工况弹窗表格列结构为：0=参数单位，1=壳程数值，2=管程数值
                    # 这里传入的 col_idx 必须对应 1/2，避免读取到错误列（会导致联动校验引用到壳/管程相反列）
                    for side, (col_idx, val) in [("壳程数值", (1, kc_val)), ("管程数值", (2, gc_val))]:
                        if not val:
                            continue
                        print(f"[导入校核DEBUG] pname={pname_base}, gongkuang={gk_no}, side={side}, val={val}")
                        result = dispatch_cell_validation(
                            viewer, table, idx, col_idx,
                            pname_base, side, val
                        )
                        if result == "error":
                            if viewer and hasattr(viewer, "import_tip_list"):
                                viewer.import_tip_list.append(
                                    f"[多工况] {pname_base}[工况{gk_no}] - {side}: ❌ 非法值“{val}”，已清空"
                                )
                            if side == "壳程数值":
                                kc_val = ""
                            else:
                                gc_val = ""

                    db_field = f"{pname_base}[工况{gk_no}]"  # ✅ 无空格版本

                    # === 查数据库是否已存在 ===
                    cur.execute("""
                        SELECT COUNT(*) AS cnt FROM 产品设计活动表_设计数据表
                        WHERE 产品ID=%s AND 参数名称=%s
                    """, (product_id, db_field))
                    exists = cur.fetchone()["cnt"] > 0

                    # === 根据规则处理 ===
                    if exists:
                        cur.execute("""
                            UPDATE 产品设计活动表_设计数据表
                            SET 壳程数值=%s, 管程数值=%s
                            WHERE 产品ID=%s AND 参数名称=%s
                        """, (kc_val, gc_val, product_id, db_field))
                    else:
                        if kc_val or gc_val:
                            seq += 1
                            cur.execute("""
                                INSERT INTO 产品设计活动表_设计数据表
                                (设计数据参数ID, 产品ID, 参数名称, 壳程数值, 管程数值)
                                VALUES (%s, %s, %s, %s, %s)
                            """, (seq, product_id, db_field, kc_val, gc_val))

                print(f"[多工况导入] 工况{gk_no} 覆盖完成")

        conn.commit()
    finally:
        conn.close()

def update_general_data_table_from_excel(excel_path: str, table_widget, viewer=None):
    """
    从Excel中读取‘通用数据’Sheet，按参数名称匹配，更新‘参数值’。
    多选项字段将自动识别并标准化为“；”分隔格式。
    """
    try:
        df = pd.read_excel(excel_path, sheet_name="通用数据", dtype=str)
        df.fillna("", inplace=True)

        # 构建映射表：参数名称 -> 参数值
        data_map = {
            str(row[1]).strip(): str(row[3]).strip()
            for _, row in df.iterrows()
        }

        # 外径系列、外径的导入值在下方按「是否以外径为基准」统一处理后再记录/使用
        for row in range(table_widget.rowCount()):
            name_item = table_widget.item(row, 1)  # 第1列是参数名称
            value_item = table_widget.item(row, 3)  # 第3列是参数值

            if not name_item or not value_item:
                continue

            name = name_item.text().strip()
            if name not in data_map:
                continue

            # 外径系列、外径：不直接复制导入数据，以配置库+公称直径联动为准
            if name in ("外径系列", "外径"):
                continue

            raw_val = data_map[name]
            config = GENERAL_PARAM_CONFIG.get(name)

            # 不做修改，保留原始值，等待后续 validate_all_tables_after_import() 中统一处理
            value_item.setText(raw_val)

        # === 导入校验：以「是否以外径为基准」为准，处理外径系列、外径 ===
        base_val = (
            data_map.get("是否以外径为基准*") or data_map.get("是否已外径为基准") or ""
        ).strip()
        if viewer is not None:
            table = table_widget
            if base_val == "否":
                # 不论导入的外径系列/外径或配置库是什么，外径系列和外径均置为 "/"
                imported_series = data_map.get("外径系列", "").strip()
                imported_diameter = data_map.get("外径", "").strip()
                need_tip = (imported_series != "/" or imported_diameter != "/")
                for r in range(table.rowCount()):
                    name_item = table.item(r, 1)
                    if not name_item:
                        continue
                    param_name = name_item.text().strip()
                    if param_name in ("外径系列", "外径"):
                        val_item = table.item(r, 3)
                        if val_item is None:
                            val_item = QTableWidgetItem()
                            table.setItem(r, 3, val_item)
                        val_item.setText("/")
                # 仅当导入数据中外径系列或外径不全是"/"时再提示
                if need_tip:
                    tip_list = getattr(viewer, "import_tip_list", None)
                    if tip_list is not None:
                        tip_list.append("已根据「是否以外径为基准」为否，将外径系列、外径置为「/」。")
            elif base_val == "是":
                # base_val 为“是”：
                #   - 若导入的外径系列为“英制系列”或“公制系列”，直接使用导入值；
                #   - 否则按配置库外径系列填充，并给出提示。
                imported_series = data_map.get("外径系列", "").strip()
                # 记录导入值（仅用于后续可能的提示或调试）
                viewer._imported_outer_series = imported_series

                if _is_shell_and_tube(viewer):
                    allowed_series = {"英制系列", "公制系列"}
                    if imported_series in allowed_series:
                        # 导入值是合法系列：以用户填写为准
                        try:
                            _set_general_outer_diameter_series(viewer, imported_series)
                        except Exception as e:
                            raise RuntimeError(f"设置导入外径系列失败：{str(e)}")
                    else:
                        # 导入值非法或为空：按配置库判定系列，并提示用户
                        determined_series = _determine_diameter_series()
                        if determined_series:
                            try:
                                _set_general_outer_diameter_series(viewer, determined_series)
                            except Exception as e:
                                raise RuntimeError(f"设置配置库外径系列失败：{str(e)}")
                            tip_list = getattr(viewer, "import_tip_list", None)
                            if tip_list is not None:
                                tip_list.append(
                                    f"⚠️ 导入外径系列为“{imported_series or '空'}”，无效，已按配置库外径系列“{determined_series}”处理。"
                                )

    except Exception as e:
        raise RuntimeError(f"导入通用数据失败：{str(e)}")

def update_trail_data_table_from_excel(excel_path: str, table_widget):
    """
    从Excel中读取‘检测数据’Sheet，只更新壳程/管程字段，
    行对齐从界面row=2开始，Excel从第3行开始（跳过两级表头）
    """
    try:
        df = pd.read_excel(excel_path, sheet_name="检测数据", header=None, skiprows=2, dtype=str)
        df.fillna("", inplace=True)

        field_to_col = {
            "壳程_技术等级": 2,
            "壳程_检测比例": 3,
            "壳程_合格级别": 4,
            "管程_技术等级": 5,
            "管程_检测比例": 6,
            "管程_合格级别": 7
        }

        current_row = 2  # ✅ 第2行是界面第一个数据行
        for _, row in df.iterrows():
            if current_row >= table_widget.rowCount():
                break

            values = {
                "壳程_技术等级": str(row[2]).strip(),
                "壳程_检测比例": str(row[3]).strip(),
                "壳程_合格级别": str(row[4]).strip(),
                "管程_技术等级": str(row[5]).strip(),
                "管程_检测比例": str(row[6]).strip(),
                "管程_合格级别": str(row[7]).strip()
            }

            # ✅ 获取当前行检测方法
            method_item = table_widget.item(current_row, 1)
            method = method_item.text().strip() if method_item else ""

            for field, col in field_to_col.items():
                val = values.get(field, "")

                # 先取界面原值
                item = table_widget.item(current_row, col)
                cur_val = item.text().strip() if item else ""

                # === 如果界面里原本是 "/"，就不覆盖 ===
                if cur_val == "/":
                    print(f"[检测数据导入][DEBUG] row={current_row}, col={col}, 原值='/' → 跳过覆盖 (Excel值={val})")
                    continue

                if not item:
                    item = QTableWidgetItem()
                    table_widget.setItem(current_row, col, item)

                item.setText(val)

                # 手动触发校验
                from modules.condition_input.funcs.funcs_cdt_input import dispatch_cell_validation
                viewer = getattr(table_widget, "viewer", None)
                if viewer:
                    header_item = table_widget.horizontalHeaderItem(col)
                    column_name = header_item.text().strip() if header_item else ""
                    dispatch_cell_validation(viewer, table_widget, current_row, col, "", column_name, val)

            # 自动补全逻辑保持不变
            if method:
                for side in ["壳程", "管程"]:
                    tech_col = field_to_col.get(f"{side}_技术等级")
                    qualify_col = field_to_col.get(f"{side}_合格级别")

                    tech_val = table_widget.item(current_row, tech_col).text().strip() if table_widget.item(current_row, tech_col) else ""
                    qualify_val = table_widget.item(current_row, qualify_col).text().strip() if table_widget.item(current_row, qualify_col) else ""

                    if not tech_val and not qualify_val:
                        from .funcs_cdt_input import autofill_trail_test_grade
                        autofill_trail_test_grade(table_widget, current_row, side,
                                                  getattr(table_widget, "undo_stack", None))

            current_row += 1

    except Exception as e:
        raise RuntimeError(f"导入检测数据失败：{str(e)}")


def update_coating_data_table_from_excel(excel_path: str, coating_table_widget, product_std_table_widget):
    """
    从Excel中读取‘涂漆数据’Sheet，更新执行标准和每组涂层数据，
    执行标准统一从产品标准表中的“涂漆标准”获取。
    """
    try:
        df = pd.read_excel(excel_path, sheet_name="涂漆数据", dtype=str, header=None)
        df.fillna("", inplace=True)

        # ✅ 从产品标准表中获取“涂漆标准”的规范代号
        coating_std_value = ""
        for row in range(product_std_table_widget.rowCount()):
            name_item = product_std_table_widget.item(row, 1)
            value_item = product_std_table_widget.item(row, 2)
            if name_item and name_item.text().strip() == "涂漆标准" and value_item:
                coating_std_value = value_item.text().strip()
                break

        # ✅ 设置到涂漆数据表第0行第2列（执行标准/规范）
        std_item = coating_table_widget.item(0, 2)
        if std_item:
            std_item.setText(coating_std_value)

        # ✅ 涂层数据从第3行开始（即df的第2行索引）
        excel_rows = []
        current_usage = ""

        for idx in range(2, len(df)):
            row = df.iloc[idx]
            usage = str(row[0]).strip()
            if usage:
                current_usage = usage

            excel_rows.append({
                "用途": current_usage,
                "细类": str(row[1]).strip(),
                "油漆类别": str(row[2]).strip(),
                "颜色": str(row[3]).strip(),
                "干膜厚度（μm）": str(row[4]).strip(),
                "涂漆面积": str(row[5]).strip(),
                "备注": str(row[6]).strip()
            })

        # ✅ 写入界面表格
        current_row = 2
        last_usage = None

        while current_row < coating_table_widget.rowCount() and excel_rows:
            excel_row = excel_rows.pop(0)
            usage = excel_row["用途"]

            for col_idx, field in enumerate([
                "用途", "细类", "油漆类别", "颜色", "干膜厚度（μm）", "涂漆面积", "备注"
            ]):
                if col_idx in (0, 1):
                    continue  # 用途、细类列不更新

                item = coating_table_widget.item(current_row, col_idx)
                if not item:
                    continue

                val = excel_row.get(field, "")
                if field in ("涂漆面积", "备注"):
                    if usage != last_usage:
                        item.setText(val)
                else:
                    item.setText(val)

            last_usage = usage
            current_row += 1

    except Exception as e:
        raise RuntimeError(f"导入涂漆数据失败：{str(e)}")

def import_all_reference_data(excel_path: str, viewer: QWidget):
    """
    给定Excel路径和界面viewer对象，一次性导入所有参考数据并更新到界面
    """
    viewer.import_tip_list = []  # ✅ 存储 dispatch 校验中捕获的错误提示

    update_product_standard_table_from_excel(excel_path, viewer.tableWidget_product_std)
    update_design_data_table_from_excel(excel_path, viewer.tableWidget_design_data)
    import_multi_conditions_from_excel(excel_path, viewer.product_id, viewer)
    update_general_data_table_from_excel(excel_path, viewer.tableWidget_general_data, viewer)
    update_trail_data_table_from_excel(excel_path, viewer.tableWidget_trail_data)
    update_coating_data_table_from_excel(
        excel_path,
        viewer.tableWidget_coating_data,
        viewer.tableWidget_product_std
    )

    trigger_all_cross_table_relations(viewer)
    validate_all_tables_after_import(viewer)

"""导入参考数据对应的检查"""
def validate_all_tables_after_import(viewer: QWidget):
    tip_list = []

    # ✅ 设计数据表（新增：校验下拉值）
    product_id = getattr(viewer, "product_id", "")
    design_dropdown_config = apply_design_data_dropdowns(viewer=viewer, product_id=product_id)

    table = viewer.tableWidget_design_data
    for row in range(table.rowCount()):
        param_item = table.item(row, 1)
        if not param_item or not param_item.text():
            continue
        param_name = param_item.text().strip()

        for col_index, col_name in [(3, "壳程数值"), (4, "管程数值")]:
            cell_item = table.item(row, col_index)
            if not cell_item or not cell_item.text():
                continue
            val = cell_item.text().strip()

            conf = design_dropdown_config.get(param_name)
            if conf and not conf.get("editable", False):
                allowed = conf.get("options", [])
                if val not in allowed:
                    cell_item.setText("")
                    tip_list.append(f"[设计数据] {param_name} - {col_name}: ❌ 非法下拉值“{val}”，已清空")
                    continue

            result = validate_design_table_cell(param_name, col_name, val, QTableWidgetItem(), table, col_index)
            if result == "error":
                cell_item.setText("")
                tip_list.append(f"[设计数据] {param_name} - {col_name}: ❌ 非法值，已清空")
            elif result == "warn":
                tip_list.append(f"[设计数据] {param_name} - {col_name}: ⚠️ 可疑值")

    # ✅ 通用数据表
    table = viewer.tableWidget_general_data
    for row in range(table.rowCount()):
        param_item = table.item(row, 1)
        value_item = table.item(row, 3)
        if not param_item or not value_item or not param_item.text() or not value_item.text():
            continue
        param_name = param_item.text().strip()
        val = value_item.text().strip()

        conf = GENERAL_PARAM_CONFIG.get(param_name)
        if conf and not conf.get("editable", False):  # ✅ 仅校验不可编辑字段
            corrected_val, msg = validate_dropdown_value(param_name, val, GENERAL_PARAM_CONFIG)
            value_item.setText(corrected_val)
            if msg:
                tip_list.append(f"[通用数据] {param_name}: {msg}")
            continue

        # ✅ 再做常规校验
        result = validate_general_table_cell(param_name, val, QTableWidgetItem(), table)
        if result == "error":
            value_item.setText("")
            tip_list.append(f"[通用数据] {param_name}: ❌ 非法值，已清空")
        elif result == "warn":
            tip_list.append(f"[通用数据] {param_name}: ⚠️ 可疑值")

    # ✅ 检测数据表：检测比例列已有校验，这里扩展对委托配置列校验（技术等级/合格级别）
    trail_config = apply_trail_data_dropdowns()
    table = viewer.tableWidget_trail_data

    for row in range(2, table.rowCount()):
        # 检测方法
        method_item = table.item(row, 1)
        method = method_item.text().strip() if method_item else ""
        conf = trail_config.get(method)

        # 接头种类：优先用 UserRole
        jt_type = ""
        probe_item = table.item(row, 2)  # 用第2列(壳程_技术等级)来探测
        if probe_item:
            jt_type = probe_item.data(Qt.UserRole) or ""

        # 技术等级/合格级别列校验
        for col_index in [2, 4, 5, 7]:
            item = table.item(row, col_index)
            if not item or not item.text() or not conf:
                continue

            val = item.text().strip()
            valid_options = []
            for cols, opts in conf.items():
                if col_index in cols:
                    valid_options = opts
                    break

            # === 新增：特殊规则放行 ===
            if jt_type == "T（管头）" and val == "/":
                print(f"[DEBUG] 放行 /: row={row}, col={col_index}, jt_type={jt_type}, method={method}")
                continue

            if valid_options and val not in valid_options:
                print(f"[DEBUG] 清空非法值: row={row}, col={col_index}, jt_type={jt_type}, method={method}, val={val}")
                item.setText("")
                tip_list.append(
                    f"[检测数据] 第{row + 1}行 - 列{col_index + 1}: ❌ 非法下拉值“{val}”，已清空"
                )

        # 检测比例列校验
        for col_index in [3, 6]:
            item = table.item(row, col_index)
            if not item or not item.text():
                continue
            val = item.text().strip()
            header_item = table.horizontalHeaderItem(col_index)
            header = header_item.text() if header_item else f"列{col_index}"

            # === 新增：特殊规则放行 ===
            if jt_type == "T（管头）" and val == "/":
                print(f"[DEBUG] 放行检测比例 /: row={row}, col={col_index}, jt_type={jt_type}, method={method}")
                continue

            result = validate_trail_table_cell(header, val, QTableWidgetItem(), table, row_index=row)
            if result == "error":
                item.setText("")
                tip_list.append(f"[检测数据] 第{row + 1}行 - {header}: ❌ 非法值，已清空")
            elif result == "warn":
                tip_list.append(f"[检测数据] 第{row + 1}行 - {header}: ⚠️ 可疑值")

    # ✅ 涂漆数据表
    table = viewer.tableWidget_coating_data
    for row in range(2, table.rowCount()):
        for col_index in [4, 5]:
            item = table.item(row, col_index)
            if not item or not item.text():
                continue
            val = item.text().strip()
            # ✅ 优先从 logical_headers 获取列名
            if hasattr(table, "logical_headers") and col_index < len(table.logical_headers):
                header = table.logical_headers[col_index]
            else:
                header_item = table.horizontalHeaderItem(col_index)
                header = header_item.text().strip() if header_item and header_item.text() else f"列{col_index}"
            result = validate_coating_table_cell(header, val, QTableWidgetItem(), table)
            print(f"Validating column: {header}, value: {val}, result: {result}")
            if result == "error":
                item.setText("")
                tip_list.append(f"[涂漆数据] 第{row+1}行 - {header}: ❌ 非法值，已清空")
            elif result == "warn":
                tip_list.append(f"[涂漆数据] 第{row+1}行 - {header}: ⚠️ 可疑值")

    # ✅ 合并导入校验过程中记录的提示（包括通用数据导入阶段关于外径系列的提示）
    if hasattr(viewer, "import_tip_list"):
        tip_list.extend(viewer.import_tip_list)

    # 导入完成后触发外径自动填充（若开启"以外径为基准"），按配置库+公称直径计算外径
    try:
        autofill_outer_diameter(viewer)
    except Exception:
        pass

    # ✅ 显示提示：主显示 + tooltip 显示完整内容
    tip_message = "\n".join(tip_list) if tip_list else "✅ 所有导入数据校验通过。"
    viewer.line_tip.setText(tip_message[:80].replace("\n", " | "))
    viewer.line_tip.setToolTip(tip_message)
    viewer.line_tip.setStyleSheet("color: black;")  # ✅ 强制黑色字体

def trigger_all_cross_table_relations(viewer: QWidget):
    """
    仅触发“绝热层类型”联动，避免影响焊接接头等其他联动逻辑。
    用于导入参考数据时确保绝热项锁定状态正确。
    """
    table = viewer.tableWidget_design_data
    for row in range(table.rowCount()):
        param_item = table.item(row, 1)
        if not param_item:
            continue
        param_name = param_item.text().strip()

        if "绝热类型" == param_name:
            for col in [3, 4]:  # 壳程和管程列
                item = table.item(row, col)
                if item and item.text().strip():
                    handle_cross_table_triggers(viewer, table, row, col)

def validate_dropdown_value(param_name: str, value: str, config: dict) -> (str, str):
    """
    检查并返回合法的下拉框值，非法则返回 ("", msg)。
    - param_name: 参数名称
    - value: 原始值
    - config: 对应的下拉配置（如 GENERAL_PARAM_CONFIG）
    """
    val = value.strip()
    conf = config.get(param_name)
    if not conf:
        return val, ""

    allowed = conf.get("options", [])
    typ = conf.get("type", "single")

    if typ == "single":
        if val not in allowed:
            return "", f"❌ 非法下拉值“{val}”，已清空"

    elif typ == "multi":
        clean_text = re.sub(r"[;；,，\s]+", "", val)

        matched = [opt for opt in allowed if opt in clean_text]

        if not matched:
            return "", f"❌ 非法选项“{value}”，已清空"

        corrected = "；".join(matched)
        return corrected, ""

    return val, ""

"""保存至本地条件输入数据表"""
def is_file_locked(filepath: str) -> bool:
    """
    判断文件是否被占用（即是否可写）
    """
    import tempfile
    import os

    if not os.path.exists(filepath):
        return False

    try:
        # 尝试以追加方式打开，如果失败说明文件被占用
        with open(filepath, 'a'):
            return False
    except IOError:
        return True

def save_local_condition_file(product_id: int, viewer: QWidget) -> bool:
    """
    保存界面数据到本地 Excel，如果文件被占用则提示并返回 False。
    —— 改动：写出时使用“默认顺序”的行索引，确保导出的 Excel 始终是固定顺序。
    """
    local_path = get_ref_data_excel_path(product_id)
    print(f"{local_path}")
    if is_file_locked(local_path):
        QMessageBox.warning(viewer, "文件占用", f"请先关闭本地文件：\n{local_path}\n然后重试保存。")
        return False  # 阻止继续
    try:
        wb = load_workbook(local_path)
    except FileNotFoundError:
        print(f"未找到本地条件数据文件：{local_path}")
        return False
    # === 关键：获取每张表的“默认写出顺序”索引 ===
    order_std     = get_row_index_order_for_default_write(viewer.tableWidget_product_std)
    order_design  = get_row_index_order_for_default_write(viewer.tableWidget_design_data)
    order_general = get_row_index_order_for_default_write(viewer.tableWidget_general_data)
    # 检测/涂漆没有“参数ID默认顺序”的诉求，仍按当前显示顺序写
    order_trail   = None
    order_coating = None
    update_sheet_from_table(
        wb["产品标准"], viewer.tableWidget_product_std,
        col_start=1, col_end=3, excel_col_offset=2, excel_row_offset=2,
        row_index_order=order_std
    )
    # 这部分代码的参数，我们后续会用到
    design_col_start = 1
    design_col_end = 5
    design_excel_col_offset = 2
    update_sheet_from_table(
        wb["设计数据"], viewer.tableWidget_design_data,
        col_start=1, col_end=5, excel_col_offset=2, excel_row_offset=2,
        row_index_order=order_design
    )
    fill_multi_conditions(wb["设计数据"], product_id, viewer.tableWidget_design_data, order_design)

    # 1. 定义模板中的最大参数行数（以NEN为准）
    TEMPLATE_MAX_ROWS = 34
    # 2. 获取当前UI界面上实际的数据行数
    current_data_rows = viewer.tableWidget_design_data.rowCount()

    # 3. 如果当前产品的数据行数小于模板的最大行数，则删除多余的行
    if current_data_rows < TEMPLATE_MAX_ROWS:
        row_to_start_delete = 2 + current_data_rows
        num_rows_to_delete = TEMPLATE_MAX_ROWS - current_data_rows
        from openpyxl.styles import Border, Side
        if num_rows_to_delete > 0:
            wb["设计数据"].delete_rows(row_to_start_delete, num_rows_to_delete)
            print(f"检测到行数不匹配：已从“设计数据”表中删除了 {num_rows_to_delete} 个多余的格式化行。")
            # 1. 定义一个标准的黑色细线边框样式
            thin_border_side = Side(style='thin', color='000000')

            # 2. 获取最后一行数据所在的Excel行号
            last_data_row_num = 2 + current_data_rows - 1

            # 3. 遍历最后一行的所有单元格，为它们“补上”下边框
            #    列的范围应该与写入数据时保持一致
            for col_idx in range(design_excel_col_offset,
                                 design_excel_col_offset + (design_col_end - design_col_start)):
                cell = wb["设计数据"].cell(row=last_data_row_num, column=col_idx)
                # 复制现有边框，只修改底部
                existing_border = cell.border
                new_border = Border(left=existing_border.left,
                                    right=existing_border.right,
                                    top=existing_border.top,
                                    bottom=thin_border_side)  # <--- 设置底部边框
                cell.border = new_border

            print(f"已成功为第 {last_data_row_num} 行数据修复底部边框。")


    update_sheet_from_table(
        wb["通用数据"], viewer.tableWidget_general_data,
        col_start=1, col_end=4, excel_col_offset=2, excel_row_offset=2,
        row_index_order=order_general
    )
    update_sheet_from_table(
        wb["检测数据"], viewer.tableWidget_trail_data,
        col_start=2, col_end=8, excel_col_offset=3, excel_row_offset=1,
        row_index_order=order_trail
    )
    update_sheet_from_table(
        wb["涂漆数据"], viewer.tableWidget_coating_data,
        col_start=2, col_end=7, excel_col_offset=3, excel_row_offset=1,
        row_index_order=order_coating
    )

    wb.save(local_path)
    print(f"✅ 本地条件数据表已成功保存到: {local_path}")
    return True

def update_sheet_from_table(sheet, table_widget, col_start=0, col_end=None,
                            excel_col_offset=1, excel_row_offset=2,
                            row_index_order=None):
    """
    将 table_widget 的指定列范围写入到 sheet 中，跳过 MergedCell，支持 Excel 起始列和起始行偏移
    —— 改动：支持 row_index_order（行索引列表）；若为 None 则使用 0..n-1。
    """
    rows = table_widget.rowCount()
    total_cols = table_widget.columnCount()
    col_end = col_end if col_end is not None else total_cols

    row_indices = row_index_order if row_index_order is not None else list(range(rows))

    for logical_row_idx, row in enumerate(row_indices):
        for col in range(col_start, col_end):
            item = table_widget.item(row, col)
            value = item.text() if item else ""

            excel_row = logical_row_idx + excel_row_offset
            excel_col = excel_col_offset + (col - col_start)

            cell = sheet.cell(row=excel_row, column=excel_col)
            if isinstance(cell, MergedCell):
                continue
            cell.value = value

def fill_multi_conditions(sheet, product_id, table_widget=None, row_index_order=None):
    """
    导出工况2/3的数据到Excel，不做参数名称匹配，直接按顺序填充。
    工况2固定写到 I/J 列，工况3固定写到 K/L 列，从第3行开始。
    """
    import re

    gongkuang_data = {}  # {工况号: [(kc, gc), (kc, gc), ...]}
    conn = get_connection(**db_config_2)
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT 参数名称, 壳程数值, 管程数值
                FROM 产品设计活动表_设计数据表
                WHERE 产品ID = %s
                ORDER BY 设计数据参数ID
            """, (product_id,))
            for row in cur.fetchall():
                pname = row["参数名称"].strip()
                kc = row.get("壳程数值") or ""
                gc = row.get("管程数值") or ""
                m = re.match(r"(.+)\s*\[工况(\d+)\]", pname)
                if m:
                    gk_no = int(m.group(2))
                    gongkuang_data.setdefault(gk_no, []).append((kc, gc))
    finally:
        conn.close()

    if not gongkuang_data:
        print("[多工况导出] 没有发现工况2/3数据，跳过")
        return

    align_center = Alignment(horizontal="center", vertical="center")

    # === 填数据（从第3行开始，直接顺序写） ===
    for gk_no, values in gongkuang_data.items():
        for idx, (kc_val, gc_val) in enumerate(values):
            excel_row = idx + 3
            if gk_no == 2:
                c1 = sheet.cell(row=excel_row, column=9)   # I列
                c2 = sheet.cell(row=excel_row, column=10)  # J列
                c1.value, c2.value = kc_val, gc_val
                c1.alignment = c2.alignment = align_center
                print(f"[多工况导出][DEBUG] 工况2 -> row={excel_row}, kc={kc_val}, gc={gc_val}")
            elif gk_no == 3:
                c1 = sheet.cell(row=excel_row, column=11)  # K列
                c2 = sheet.cell(row=excel_row, column=12)  # L列
                c1.value, c2.value = kc_val, gc_val
                c1.alignment = c2.alignment = align_center
                print(f"[多工况导出][DEBUG] 工况3 -> row={excel_row}, kc={kc_val}, gc={gc_val}")




"""跨表联动逻辑函数"""
def show_info_tip(viewer: QWidget, message: str):
    viewer.line_tip.setText(message)
    viewer.line_tip.setToolTip(message)

def handle_cross_table_triggers(viewer: QWidget, changed_table: QTableWidget, row: int, col: int):
    undo_stack = getattr(viewer, "undo_stack", None)

    # ✅ 涂漆标准 → 执行标准/规范联动
    if changed_table == viewer.tableWidget_product_std:

        name_item = changed_table.item(row, 1)
        value_item = changed_table.item(row, 2)
        if name_item and value_item and name_item.text().strip() == "涂漆标准":
            std_value = value_item.text().strip()
            target_table = viewer.tableWidget_coating_data
            std_cell = target_table.item(0, 2)

            if std_cell is None:
                std_cell = QTableWidgetItem()
                std_cell.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                std_cell.setTextAlignment(Qt.AlignCenter)
                target_table.setItem(0, 2, std_cell)

            old_val = std_cell.text()
            if std_value != old_val and undo_stack:
                from modules.condition_input.funcs.undo_command import CellEditCommand
                cmd = CellEditCommand(target_table, 0, 2, old_val, std_value)
                undo_stack.push(cmd)
                cmd.redo()

            show_info_tip(viewer, "[涂漆数据]执行标准/规范已自动刷新。")

    # ✅ 焊接接头系数* → 检测数据（仅壳程或管程）
    elif changed_table == viewer.tableWidget_design_data:
        name_item = changed_table.item(row, 1)
        if not name_item:
            return

        param_name = name_item.text().strip()

        # ✅ 焊接接头系数联动检测数据
        if "焊接接头系数*" in param_name:
            if col == 3:
                shell_val = changed_table.item(row, 3).text().strip()
                update_trail_table_side_only(viewer.tableWidget_trail_data, "壳程", shell_val, undo_stack)
                show_info_tip(viewer, "[检测数据]壳程检测比例及合格级别已自动刷新。")
            elif col == 4:
                tube_val = changed_table.item(row, 4).text().strip()
                update_trail_table_side_only(viewer.tableWidget_trail_data, "管程", tube_val, undo_stack)
                show_info_tip(viewer, "[检测数据]管程检测比例及合格级别已自动刷新。")

        # ✅ 绝热层类型联动
        elif param_name == "绝热类型":
            side = "壳程" if col == 3 else "管程" if col == 4 else None
            if not side:
                return

            cell = changed_table.item(row, col)
            val_text = cell.text().strip() if cell else ""
            prev_val = getattr(cell, "_prev_val", "") if cell else ""
            cell._prev_val = val_text  # 记录当前为下次使用

            is_none_now = val_text == "无"
            is_none_prev = prev_val == "无"

            # ✅ 仅当从“无”↔其他值之间变化时联动
            if is_none_now == is_none_prev:
                print("跳过绝热类型联动（状态未变化）")
                return

            make_fields_editable = not is_none_now
            param_names = {"绝热材料", "绝热层厚度", "绝热材料密度"}

            for r in range(changed_table.rowCount()):
                sub_item = changed_table.item(r, 1)
                if not sub_item or sub_item.text().strip() not in param_names:
                    continue

                target_col = 3 if side == "壳程" else 4
                target_item = changed_table.item(r, target_col)
                if target_item is None:
                    target_item = QTableWidgetItem()
                    changed_table.setItem(r, target_col, target_item)

                if not make_fields_editable:
                    target_item.setText("")
                    target_item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                else:
                    target_item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled | Qt.ItemIsEditable)

            show_info_tip(viewer, f"[设计数据]{side}绝热项状态已更新")

        # 1206新修改-外径、外径系列、是否已外径为基准、公称直径联动
        # ✅ 公称直径* 变化 → 触发外径自动填充（仅当“是否以外径为基准*”为“是”时）
        elif param_name == "公称直径*" and col in (3, 4):
            if _is_shell_and_tube(viewer) and _is_outer_by_diameter_enabled(viewer):
                try:
                    autofill_outer_diameter(viewer)
                    # show_info_tip(viewer, "[通用数据]外径已根据公称直径与外径系列自动更新。")
                except Exception:
                    pass

    # ✅ 检测比例 → 联动补齐 技术等级 和 合格级别（仅当为空）
    # ✅ 新增：清空其中任一字段 → 自动清空其余两个字段
    elif changed_table == viewer.tableWidget_trail_data:
        header_item = changed_table.horizontalHeaderItem(col)
        col_name = header_item.text().strip() if header_item else ""
        side = None

        if "壳程" in col_name:
            side = "壳程"
        elif "管程" in col_name:
            side = "管程"

        # 自动补齐技术等级与合格级别
        if col_name in [f"{side}_检测比例"] and side:
            did_fill = autofill_trail_test_grade(changed_table, row, side, undo_stack)
            if did_fill:
                show_info_tip(viewer, f"[检测数据]{side}检测比例已自动联动更新技术等级与合格级别")

        # 清空联动逻辑
        if side and col_name in [f"{side}_技术等级", f"{side}_检测比例", f"{side}_合格级别"]:
            item = changed_table.item(row, col)
            if item and item.text().strip() == "":
                related_cols = {
                    f"{side}_技术等级": [f"{side}_检测比例", f"{side}_合格级别"],
                    f"{side}_检测比例": [f"{side}_技术等级", f"{side}_合格级别"],
                    f"{side}_合格级别": [f"{side}_技术等级", f"{side}_检测比例"]
                }
                for other_col_name in related_cols.get(col_name, []):
                    col_idx = next((i for i in range(changed_table.columnCount())
                                    if changed_table.horizontalHeaderItem(i).text().strip() == other_col_name), None)
                    if col_idx is not None:
                        target_item = changed_table.item(row, col_idx)
                        if target_item and target_item.text().strip():
                            old_val = target_item.text()
                            target_item.setText("")
                            if undo_stack:
                                from modules.condition_input.funcs.undo_command import CellEditCommand
                                undo_stack.push(CellEditCommand(changed_table, row, col_idx, old_val, ""))

    # 1206新修改-外径、外径系列、是否已外径为基准、公称直径联动
    # ✅ 通用数据：外径系列 变化 → 触发外径自动填充
    elif changed_table == viewer.tableWidget_general_data:
        name_item = changed_table.item(row, 1)
        if not name_item:
            return
        param_name = name_item.text().strip()
        # 外径系列变更
        if param_name == "外径系列" and col == 3:
            # 防抖：避免代理与 itemChanged 双路径导致短时间重复触发
            try:
                import time
                last_ts = getattr(viewer, "_outer_series_ts", 0)
                now_ts = int(time.time() * 1000)
                if now_ts - last_ts < 800:
                    return
                setattr(viewer, "_outer_series_ts", now_ts)
            except Exception:
                pass
            try:
                autofill_outer_diameter(viewer)
                # show_info_tip(viewer, "[通用数据]外径已根据公称直径与外径系列自动更新。")
            except Exception:
                pass
        # 基准开关切换的情况由 view.update_general_diameter_linkage 统一处理（避免重复触发）

def update_trail_table_side_only(table: QTableWidget, side: str, factor_val: str, undo_stack=None):
    """
    根据焊接接头系数，联动更新检测数据表指定侧（壳程或管程）的：
    - 技术等级
    - 检测比例
    - 合格级别
    ✅ 同时设置默认值（UserRole+2）用于后续校验。
    """
    factor_map = {
        "1":    ("AB", "100", "Ⅱ"),
        "1.0":  ("AB", "100", "Ⅱ"),
        "0.9":  ("AB", "100", "Ⅱ"),
        "0.85": ("AB", "≥20", "Ⅲ"),
        "0.8":  ("AB", "≥20", "Ⅲ")
    }

    if factor_val not in factor_map:
        print(f"❎ 跳过无效系数: {factor_val}")
        return

    row = 2  # 固定行（第一行数据）
    col_map = {
        "壳程": {"等级": 2, "比例": 3, "合格": 4},
        "管程": {"等级": 5, "比例": 6, "合格": 7}
    }

    if side not in col_map:
        return

    grade_val, ratio_val, qualify_val = factor_map[factor_val]
    values_to_set = {
        "等级": grade_val,
        "比例": ratio_val,
        "合格": qualify_val
    }

    for field, new_val in values_to_set.items():
        col = col_map[side][field]
        item = table.item(row, col)
        if not item:
            item = QTableWidgetItem()
            table.setItem(row, col, item)

        old_val = item.text()
        item.setText(new_val)
        item.setData(Qt.UserRole + 2, new_val)  # ✅ 设置默认值以供后续校验使用

        if undo_stack and old_val != new_val:
            from modules.condition_input.funcs.undo_command import CellEditCommand
            undo_stack.push(CellEditCommand(table, row, col, old_val, new_val))

    print(f"✅ {side}联动成功: 系数={factor_val} → 等级={grade_val}, 比例={ratio_val}, 合格={qualify_val}")

def autofill_trail_test_grade(trail_table: QTableWidget, row: int, side: str, undo_stack: QUndoStack) -> bool:
    """
    自动推导 技术等级 / 合格级别（无论是否为空，强制写入）：
    - side: "壳程" / "管程"
    - 返回值：是否发生写入
    """
    headers = {trail_table.horizontalHeaderItem(c).text().strip(): c
               for c in range(trail_table.columnCount()) if trail_table.horizontalHeaderItem(c)}

    method_item = trail_table.item(row, headers.get("检测方法"))
    ratio_item = trail_table.item(row, headers.get(f"{side}_检测比例"))
    if not method_item or not ratio_item:
        return False

    method = method_item.text().strip()
    ratio = ratio_item.text().strip()
    if not method or not ratio:
        return False

    if validate_trail_table_cell(f"{side}_检测比例", ratio, None, trail_table) != "ok":
        return False

    import re
    try:
        ratio_num = float(re.sub(r"[^\d.]", "", ratio))
    except ValueError:
        return False

    match_table = {
        "R.T.":  [("100", "AB", "Ⅱ"), ("≥20", "AB", "Ⅲ")],
        "D.R.":  [("100", "AB", "Ⅱ"), ("≥20", "AB", "Ⅲ")],
        "C.R.":  [("100", "AB", "Ⅱ"), ("≥20", "AB", "Ⅲ")],
        "U.T.":  [("100", "B",  "Ⅰ"), ("≥20", "B",  "Ⅱ")],
        "U.I.T.": [("100", "B",  "Ⅰ"), ("≥20", "B",  "Ⅱ")],
        "TOFD": [("100", "B",  "Ⅰ"), ("≥20", "B",  "Ⅱ")],
        "PAUT": [("100", "B",  "Ⅰ"), ("≥20", "B",  "Ⅱ")],
        "M.T.": [("100", "/",  "Ⅰ")],
        "P.T.": [("100", "/",  "Ⅰ")],
        "M.T.[FB]": [("100", "/", "Ⅰ")]
    }

    candidates = match_table.get(method)
    if not candidates:
        return False

    selected_grade = ""
    selected_qualify = ""
    for limit_str, grade, qualify in candidates:
        if ratio_num >= float(re.sub(r"[^\d.]", "", limit_str)):
            selected_grade = grade
            selected_qualify = qualify
            break

    def force_update_cell(col_name: str, new_val: str) -> bool:
        col = headers.get(col_name)
        if col is None:
            return False
        old_item = trail_table.item(row, col)
        old_val = old_item.text().strip() if old_item else ""
        if not old_item:
            old_item = QTableWidgetItem()
            trail_table.setItem(row, col, old_item)

        old_item.setText(new_val)
        old_item.setData(Qt.UserRole + 2, new_val)
        if undo_stack:
            undo_stack.push(CellEditCommand(trail_table, row, col, old_val, new_val))
        return old_val != new_val

    did_fill1 = force_update_cell(f"{side}_技术等级", selected_grade)
    did_fill2 = force_update_cell(f"{side}_合格级别", selected_qualify)
    return did_fill1 or did_fill2

def compute_trail_default_grade(method: str, ratio_str: str, field_type: str) -> str:
    """
    根据检测方法和检测比例，返回默认 技术等级 或 合格级别。
    - method: 检测方法，如 "R.T."
    - ratio_str: 比例字段，如 "100" 或 "≥20"
    - field_type: "技术等级" 或 "合格级别"
    """
    match_table = {
        "R.T.":  [("100", "AB", "Ⅱ"), ("≥20", "AB", "Ⅲ")],
        "D.R.":  [("100", "AB", "Ⅱ"), ("≥20", "AB", "Ⅲ")],
        "C.R.":  [("100", "AB", "Ⅱ"), ("≥20", "AB", "Ⅲ")],
        "U.T.":  [("100", "B",  "Ⅰ"), ("≥20", "B",  "Ⅱ")],
        "U.I.T.":[("100", "B",  "Ⅰ"), ("≥20", "B",  "Ⅱ")],
        "TOFD": [("100", "B",  "Ⅰ"), ("≥20", "B",  "Ⅱ")],
        "PAUT": [("100", "B",  "Ⅰ"), ("≥20", "B",  "Ⅱ")],
        "M.T.": [("100", "/",  "Ⅰ")],
        "P.T.": [("100", "/",  "Ⅰ")],
        "M.T.[FB]": [("100", "/", "Ⅰ")]
    }

    import re
    def extract_num(s):
        try:
            return float(re.sub(r"[^\d.]", "", s))
        except:
            return -1

    ratio_num = extract_num(ratio_str)
    candidates = match_table.get(method, [])

    for limit_str, tech, qualify in candidates:
        if ratio_num >= extract_num(limit_str):
            return tech if field_type == "技术等级" else qualify
    return ""

"""技术等级和合格级别不能低于默认值"""
GRADE_ORDER = {"AB": 1, "B": 2, "C": 3}
QUALIFY_ORDER = {"Ⅲ": 1, "Ⅱ": 2, "Ⅰ": 3}

def is_grade_lower(user_val: str, default_val: str) -> bool:
    return GRADE_ORDER.get(user_val, 0) < GRADE_ORDER.get(default_val, 0)

def is_qualify_lower(user_val: str, default_val: str) -> bool:
    return QUALIFY_ORDER.get(user_val, 0) < QUALIFY_ORDER.get(default_val, 0)


"""下拉框定义"""
class MultiParamComboDelegate(QStyledItemDelegate):
    def __init__(self, config: dict, parent=None, viewer=None, undo_stack=None):
        super().__init__(parent)
        self.config = config  # {参数名: {"type": "single"|"multi", "options": [...], "editable": bool}}
        self.viewer = viewer
        self.undo_stack = undo_stack

    def _get_config(self, index):
        row, col = index.row(), index.column()
        param_item = self.parent().item(row, 1)
        if not param_item:
            return None, None
        param_name = param_item.text().strip()
        return self.config.get(param_name), param_name

    def createEditor(self, parent, option, index):
        conf, _ = self._get_config(index)
        if not conf:
            return super().createEditor(parent, option, index)

        if conf["type"] == "multi":
            editor = CheckableComboBox(conf["options"], parent)
            return editor
        else:
            combo = QComboBox(parent)
            combo.addItems(conf["options"])
            combo.setEditable(conf.get("editable", False))
            return combo

    def setEditorData(self, editor, index):
        conf, _ = self._get_config(index)
        if not conf:
            return super().setEditorData(editor, index)
        val = index.data()

        if conf["type"] == "multi":
            values = [v.strip() for v in val.split("；") if v.strip()]
            editor.setCheckedItems(values)
        else:
            i = editor.findText(val)
            editor.setCurrentIndex(i if i >= 0 else 0)

    def setModelData(self, editor, model, index):
        conf, param_name = self._get_config(index)
        if not conf:
            return super().setModelData(editor, model, index)

        old_val = index.data()

        if conf["type"] == "multi":
            new_val = "；".join(editor.checkedItems())
        else:
            new_val = editor.currentText()

        model.setData(index, new_val)

        if old_val != new_val and self.undo_stack:
            cmd = CellEditCommand(self.parent(), index.row(), index.column(), old_val, new_val)
            self.undo_stack.push(cmd)

        # 校验 & 联动
        if self.viewer:
            row, col = index.row(), index.column()
            table = self.parent()
            param_item = table.item(row, 1)
            param_name = param_item.text().strip() if param_item else ""

            if hasattr(table, "logical_headers"):
                column_name = table.logical_headers[col]
            else:
                header_item = table.horizontalHeaderItem(col)
                column_name = header_item.text().strip() if header_item else ""

            # ✅ 调用统一校验分发
            dispatch_cell_validation(self.viewer, table, row, col, param_name, column_name, new_val)

            handle_cross_table_triggers(self.viewer, table, row, col)

#创建自定义 QComboBox 带 checkbox
class CheckableComboBox(QComboBox):
    def __init__(self, options, parent=None):
        super().__init__(parent)
        self.setEditable(True)
        self.setInsertPolicy(QComboBox.NoInsert)
        self.view().setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.setModel(QStandardItemModel(self))
        self._options = options
        self._init_items(options)
        self.lineEdit().setReadOnly(True)
        self.lineEdit().setText("")

    def _init_items(self, options):
        for text in options:
            item = QStandardItem(text)
            item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsUserCheckable)
            item.setData(Qt.Unchecked, Qt.CheckStateRole)
            self.model().appendRow(item)

    def hidePopup(self):
        selected = []
        for i in range(self.model().rowCount()):
            item = self.model().item(i)
            if item.checkState() == Qt.Checked:
                selected.append(item.text())
        self.lineEdit().setText("；".join(selected))
        super().hidePopup()

    def setCheckedItems(self, values: list):
        # 合并为一个原始字符串，用于乱序/无分隔判断
        raw_text = "".join(values)

        selected = []
        for i in range(self.model().rowCount()):
            item = self.model().item(i)
            option_text = item.text()
            # 若 option_text 在任何原始片段中出现（哪怕没分号），也视为勾选
            if any(option_text in v for v in values) or option_text in raw_text:
                item.setCheckState(Qt.Checked)
                selected.append(option_text)
            else:
                item.setCheckState(Qt.Unchecked)

        self.lineEdit().setText("；".join(selected))

    def checkedItems(self) -> list:
        return [self.model().item(i).text()
                for i in range(self.model().rowCount())
                if self.model().item(i).checkState() == Qt.Checked]

"""添加各表格下拉框"""

#勿删有用！！！
def _get_config(self, index):
    try:
        row, col = index.row(), index.column()
        param_item = self.parent().item(row, 1)
        if not param_item:
            return None, None
        param_name = param_item.text().strip()
        return self.config.get(param_name), param_name
    except Exception as e:
        print(f"[下拉框配置错误] 无法获取参数名: {e}")
        return None, None

#设计数据下拉框
def fetch_design_dropdown_config(product_id):
    """
    从数据库读取所有下拉字段配置，返回 config 字典
    """
    config = {}
    conn = get_connection(**db_config_1)
    try:
        with conn.cursor() as cursor:
            cursor.execute("""
                SELECT 参数名称, type, editable, options
                FROM 设计数据选项模板
            """)
            rows = cursor.fetchall()
            for row in rows:
                param = row["参数名称"]
                typ = row["type"]
                editable = str(row["editable"]).lower() in ("true", "1", "是")
                try:
                    options = ast.literal_eval(row["options"])
                except Exception as e:
                    print(f"⚠️ 参数 {param} 的选项解析失败：{e}")
                    options = []

                config[param] = {
                    "type": typ,
                    "editable": editable,
                    "options": options
                }
    finally:
        conn.close()

    return config
def apply_design_data_dropdowns(table_widget=None, product_id=None, viewer=None, undo_stack=None):
    config = fetch_design_dropdown_config(product_id)

    # ⚠️ 特殊逻辑：耐压试验类型，根据产品类型删减选项
    if product_id:
        prod_type = get_product_type_from_db(product_id)
        if prod_type == "管壳式热交换器":
            if "耐压试验类型*" in config:
                config["耐压试验类型*"]["options"] = ["液压试验", "气压试验"]

    return config
def get_product_type_from_db(product_id):
    from modules.condition_input.funcs.db_cnt import get_connection
    conn = get_connection(**db_config_3)
    try:
        with conn.cursor() as cursor:
            cursor.execute("SELECT 产品类型 FROM 产品需求表 WHERE 产品ID = %s", (product_id,))
            result = cursor.fetchone()
            return result.get("产品类型") if result else ""
    finally:
        conn.close()

#通用数据下拉框
def fetch_general_dropdown_config():
    """
    从数据库读取通用数据表的下拉字段配置
    注意：外径系列已从下拉框配置中移除，改为只读显示，值由user_config决定
    """
    config = {}
    conn = get_connection(**db_config_1)
    try:
        with conn.cursor() as cursor:
            cursor.execute("""
                SELECT 参数名称, type, editable, options
                FROM 通用数据选项模板
            """)
            rows = cursor.fetchall()
            for row in rows:
                name = row["参数名称"]
                # # === 移除外径系列的下拉框配置 ===
                # if name == "外径系列":
                #     continue  # 跳过外径系列，不添加到下拉框配置中
                
                typ = row["type"]
                editable = str(row["editable"]).lower() in ("true", "1", "是")
                try:
                    options = ast.literal_eval(row["options"])
                except Exception as e:
                    print(f"⚠️ 参数 {name} 的 options 无法解析：{e}")
                    options = []

                config[name] = {
                    "type": typ.strip(),
                    "editable": editable,
                    "options": options
                }
    finally:
        conn.close()
    return config
def apply_general_data_dropdowns():
    return fetch_general_dropdown_config()
#勿删
GENERAL_PARAM_CONFIG = fetch_general_dropdown_config()

def fetch_trail_dropdown_config():
    """
    从数据库读取“检测数据”下拉选项配置，返回结构如：
    {
        "R.T.": {
            (2,5): [...],
            (4,7): [...]
        },
        ...
    }
    """
    config = {}
    conn = get_connection(**db_config_1)
    try:
        with conn.cursor() as cursor:
            cursor.execute("SELECT `检测方法`, `column`, `options` FROM 无损检测数据选项模板")
            for row in cursor.fetchall():
                method = row["检测方法"]
                column_str = row["column"]  # 例如 "2,5"
                try:
                    columns = tuple(int(c.strip()) for c in column_str.split(","))
                    options = ast.literal_eval(row["options"])
                except Exception as e:
                    print(f"❌ 检测数据选项解析失败: {method}-{column_str}: {e}")
                    continue

                if method not in config:
                    config[method] = {}
                config[method][columns] = options
    finally:
        conn.close()
    return config
class TrailTableComboDelegate(QStyledItemDelegate):
    def __init__(self, config=None, parent=None):
        super().__init__(parent)
        self.config = config or {}

    def createEditor(self, parent, option, index):
        method_item = index.sibling(index.row(), 1)
        method_name = method_item.data().strip() if method_item and method_item.data() else ""

        col = index.column()
        options = []
        method_conf = self.config.get(method_name)
        if method_conf:
            for key_cols, vals in method_conf.items():
                if col in key_cols:
                    options = vals
                    break

        if not options:
            return super().createEditor(parent, option, index)

        combo = QComboBox(parent)
        combo.addItems(options)
        combo.setEditable(False)

        # ✅ 添加 Delete / Backspace 快捷键
        for key in (Qt.Key_Delete, Qt.Key_Backspace):
            shortcut = QShortcut(QKeySequence(key), combo)
            shortcut.activated.connect(lambda idx=index: self._clear_current_combo_cell(idx))

        return combo

    def setModelData(self, editor, model, index):
        method_item = index.sibling(index.row(), 1)
        method_name = method_item.data().strip() if method_item and method_item.data() else ""

        col = index.column()
        new_val = editor.currentText()
        old_val = index.data()
        model.setData(index, new_val)

        # ✅ 撤销记录
        table = self.parent()
        undo_stack = getattr(table, "undo_stack", None)
        if undo_stack and old_val != new_val:
            from modules.condition_input.funcs.undo_command import CellEditCommand
            undo_stack.push(CellEditCommand(table, index.row(), index.column(), old_val, new_val))

        # ✅ 调用校验 & 联动
        viewer = getattr(table, "viewer", None)
        if viewer:
            row = index.row()
            header_item = table.horizontalHeaderItem(col)
            column_name = header_item.text().strip() if header_item else ""
            from modules.condition_input.funcs.funcs_cdt_input import dispatch_cell_validation, handle_cross_table_triggers

            dispatch_cell_validation(viewer, table, row, col, "", column_name, new_val)
            QTimer.singleShot(0, lambda: handle_cross_table_triggers(viewer, table, row, col))

        # 自动提示等级选项改变也能触发联动

    #已修改
    def is_dropdown_cell(self, index):
        col = index.column()
        row = index.row()

        # ✅ 跳过前2行（表头）或越界行
        if row < 2 or row >= self.parent().rowCount():
            return False

        method_item = index.sibling(row, 1)
        if not method_item:
            return False

        method_data = method_item.data()
        if not isinstance(method_data, str):
            return False

        method_name = method_data.strip()
        method_conf = self.config.get(method_name, {})

        for key_cols in method_conf.keys():
            if col in key_cols:
                return True

        return False

    def _clear_current_combo_cell(self, index):
        table = self.parent()
        row, col = index.row(), index.column()
        item = table.item(row, col)
        if item:
            old_val = item.text()
            item.setText("")
            undo_stack = getattr(table, "undo_stack", None)
            if undo_stack:
                from .undo_command import CellEditCommand
                undo_stack.push(CellEditCommand(table, row, col, old_val, ""))

            # ✅ 校验 & 联动
            viewer = getattr(table, "viewer", None)
            if viewer:
                header_item = table.horizontalHeaderItem(col)
                column_name = header_item.text().strip() if header_item else ""
                from .funcs_cdt_input import dispatch_cell_validation, handle_cross_table_triggers
                dispatch_cell_validation(viewer, table, row, col, "", column_name, "")
                QTimer.singleShot(0, lambda: handle_cross_table_triggers(viewer, table, row, col))

            table.closePersistentEditor(item)


def apply_trail_data_dropdowns():
    return fetch_trail_dropdown_config()


