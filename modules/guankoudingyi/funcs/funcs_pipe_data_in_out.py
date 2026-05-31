import math
import os
from collections import OrderedDict
from datetime import datetime

import openpyxl
from PyQt5.QtWidgets import QTableWidgetItem, QMessageBox, QFileDialog
from PyQt5.QtCore import Qt


from PyQt5.QtWidgets import QFileDialog, QMessageBox, QTableWidgetItem
from PyQt5.QtCore import Qt
from openpyxl import load_workbook
from openpyxl.cell import MergedCell
from openpyxl.styles import Border, Side, Font
from openpyxl.utils import get_column_letter
import pymysql
from modules.guankoudingyi.funcs.funcs_pipe_comboBox_value import get_component_nominal_size_od, get_nominal_diameter
from modules.guankoudingyi.db_cnt import get_connection, db_config_2, db_config_1
from modules.guankoudingyi.funcs.funcs_pipe_table import check_last_row_and_add_new, is_duplicate_port_code, delete_selected_pipe_rows

# —— 需要写入模板的字段（界面 -> 模板中文名）——
# 管口定义界面的映射：左边是界面列的中文名（表格里用的），右边是模板里“参数中文名”所在单元格的文字。
# —— 需要写入模板的字段（界面 -> 模板中文名）——
# 管口定义界面的映射：左边是界面列的中文名（表格里用的），右边是模板里“参数中文名”所在单元格的文字。
FIELD_MAP = OrderedDict([
    ("管口代号", "管口代号"),
    ("管口功能", "管口功能"),
    ("管口用途", "管口用途"),
    ("公称尺寸", "管口公称尺寸"),
    ("管口所属元件", "管口所属元件"),
    ("所属元件焊接接头系数", "所属元件焊接接头系数"),
    ("轴向定位基准", "管口轴向定位基准"),
    ("轴向定位距离", "管口轴向定位距离"),
    ("接管与壳体连接结构形式", "接管与壳体连接结构形式"),
    ("轴向夹角（°）", "管口轴向夹角"),
    ("周向方位（°）", "管口周向方位"),
    ("偏心距", "管口偏心距"),

    # ——法兰参数——
    ("法兰标准", "接管法兰标准"),
    ("压力等级", "接管法兰压力等级"),
    ("法兰型式", "接管法兰型式"),
    ("密封面型式", "接管法兰密封面型式"),
    ("焊端规格", "接管法兰焊端规格"),
])

# ——从《元件计算结果表》写入“接管材料”列需要的映射（DB Name -> 模板行名）——
DBNAME_TO_TEMPLATE_ROW = {
    "开孔元件内径": "所属元件内径",
    "开孔元件有效厚度": "所属元件壁厚",
    "开孔元件焊接接头系数": "所属元件焊接接头系数",
    "接管焊接接头系数": "接管焊接接头系数",
    "接管与壳体连接结构型式": "接管与壳体连接结构形式",
    "接管名义内直径": "接管内径",
    "接管大端外径": "接管外径",
    "接管实际外伸长度": "接管实际外伸高度",
    "接管实际内伸长度": "接管实际内伸高度",
    "开孔元件腐蚀裕量": "接管腐蚀裕量",
    "接管覆层厚度": "接管覆层厚度",
    "接管小端壁厚": "接管小端壁厚",
    "接管大端壁厚": "接管大端壁厚",
    "接管锥段角度": "接管锥段角度",
    "接管材料类型": "接管材料类型",
    "接管材料牌号": "接管材料牌号",
    "接管材料试验温度下许用应力": "接管材料试验温度下许用应力",
    "接管材料设计温度下许用应力": "接管材料设计温度下许用应力",
    "接管材料试验温度下屈服强度": "接管材料设计温度下屈服强度",  # 以你的模板行名为准
    "接管重量": "接管重量",
}

# ——界面列名到列号的映射（与你渲染时的顺序一致；序号列=0，从1开始对应下方）——
UI_COL_INDEX = {
    "管口代号": 1,
    "管口功能": 2,
    "管口用途": 3,
    "公称尺寸": 4,
    "法兰标准": 5,
    "压力等级": 6,
    "法兰型式": 7,
    "密封面型式": 8,
    "焊端规格": 9,
    "管口所属元件": 10,
    "轴向定位基准": 11,
    "轴向定位距离": 12,
    "轴向夹角（°）": 13,
    "周向方位（°）": 14,
    "偏心距": 15,
    "外伸高度": 16,
    "管口附件": 17,
    "管口载荷": 18,
}


def _collect_nozzle_rows_from_ui(stats_widget):
    """从界面 tableWidget_pipe 收集每个管口的一列数据（最后空白行忽略）"""
    table = stats_widget.tableWidget_pipe
    rows = []

    # 过滤掉最后新增空白行
    last = table.rowCount() - 1
    if last < 0:
        return rows

    def cell(r, c):
        item = table.item(r, c)
        return "" if (item is None or item.text() == "None") else item.text().strip()

    for r in range(0, last):  # 不包含最后空白行
        code = cell(r, UI_COL_INDEX["管口代号"])
        if not code:
            # 没有管口代号的直接跳过
            continue
        one = {"管口代号": code}
        for k in UI_COL_INDEX:
            if k == "管口代号":
                continue
            one[k] = cell(r, UI_COL_INDEX[k])
        rows.append(one)
    return rows


def _find_column_indexes_to_hide(ws):
    """
    找到需要隐藏的列（第二行表头单元格含"隐藏列"字样的列）
    注意：模板里第2行表头中包含"隐藏列"的那些列需要在导出时删除。
    """
    to_hide = set()
    for col in range(1, ws.max_column + 1):
        v = (ws.cell(row=2, column=col).value or "")
        if isinstance(v, str) and "隐藏列" in v:
            to_hide.add(col)
    return to_hide


def _build_row_index_by_param_name(ws):
    """
    在模板中定位“参数名所在行”（扫描的是excel表格）。
    我们就扫描整表：把单元格文本作为 key，行号作为 value。
    """
    name2row = {}
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip():
                name2row.setdefault(v.strip(), r)
    return name2row


def export_nozzle_listing(stats_widget, template_rel_dir="guankoudingyi/table_template",
                          template_name="管口导出模板.xlsx",
                          out_dir_rel="exports"):
    """
    导出主方法：读取模板 → 填值 → 隐藏“隐藏列” → 另存
    返回导出的绝对路径
    """
    # 1) 找模板
    proj_root = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))  # 你的模块在 modules.guankoudingyi 下
    template_path = os.path.join(proj_root, template_rel_dir, template_name)
    # print("[调试] 当前模板路径 =", template_path)
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"未找到导出模板：{template_path}")

    # 2) 收集界面数据
    nozzle_cols = _collect_nozzle_rows_from_ui(stats_widget)
    if not nozzle_cols:
        raise RuntimeError("没有可导出的管口数据（请先输入管口代号等信息）")

    # 3) 打开模板
    wb = load_workbook(template_path)
    ws = wb.active

    # [修改] 记录并解除模板里所有既有合并单元格，避免随后的 delete_cols/insert_cols
    #       破坏 <mergeCells> 结构，导致 Excel 打开时“发现问题，是否修复？”
    existing_merged = list(ws.merged_cells.ranges)
    if existing_merged:
        for rng in existing_merged:
            ws.unmerge_cells(str(rng))

    # 4) 找“参数中文名”所在行：用全表扫描建立索引
    row_index = _build_row_index_by_param_name(ws)

    # 5) 先删除模板里第2行表头带"隐藏列"的列，避免影响后续列号计算
    hide_cols = _find_column_indexes_to_hide(ws)
    # 按倒序删除列，避免删除后列号变化的问题
    for col in sorted(hide_cols, reverse=True):
        ws.delete_cols(col)

    # 6) 找模板里“值”列的起点：按你的模板，一般有一个标注为“值”的列（或直接指定第一个值列）
    start_col = None
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=2, column=c).value  # excel的第二行
        if isinstance(v, str) and v.strip() == "值":
            start_col = c
            # print("(1)起始列为:", start_col)
            break
    if start_col is None:
        start_col = 5
    # print("(2)起始列为:", start_col)

    # 7) 写表头：
    # 在第二行覆盖"值"列的表头为"管口1"
    # 其余“管口2、3...”逐个插入到“值”列右边（始终在“单位、备注”前面）
    if nozzle_cols:
        # 覆盖“值”列➡管口1
        ws.cell(row=2, column=start_col, value="管口1").font = Font(bold=True)
        # 后续管口➡每次在start+1 位置插入新列
        for i in range(1, len(nozzle_cols)):
            insert_at = start_col + i  # 插入点（始终把“单位、备注”推到右边）
            ws.insert_cols(insert_at)
            header_cell_2 = ws.cell(row=2, column=insert_at, value=f"管口{i + 1}")
            header_cell_2.font = Font(bold=True)  # 设置表头加粗

    # 8) 写值：按 FIELD_MAP 把每个"接管材料"行 × 各管口列 写入
    # ===== 一次性构建 代号->元件名称 映射，导出循环复用 =====
    product_id = getattr(stats_widget, "product_id", None)
    nozzle_elem_map = build_nozzlecode_to_element_map(product_id) if product_id else {}

    for i, col_data in enumerate(nozzle_cols):
        out_col = start_col + i

        # 8.1 写 UI 字段
        for ui_key, tpl_cn_name in FIELD_MAP.items():
            row = row_index.get(tpl_cn_name)
            if not row:
                continue

            # ——所属元件焊接接头系数 动态计算——
            if ui_key == "所属元件焊接接头系数":
                pipe_belong = (col_data.get("管口所属元件") or "").strip()
                if not product_id or not pipe_belong:
                    val = ""  # 没有产品ID或未选所属元件 → 留空
                else:
                    ok, v = _get_weld_joint_efficiency(product_id, pipe_belong)
                    # 失败时留空
                    val = v if ok else ""
                _set_cell_safely(ws, row, out_col, val)
                continue

            _set_cell_safely(ws, row, out_col, col_data.get(ui_key, ""))

        # 8.2 由“管口代号->元件名称”，批量写《元件计算结果表》的值
        nozzle_code = (col_data.get("管口代号") or "").strip()
        elem_name = (nozzle_elem_map.get(nozzle_code) or "").strip()
        if not elem_name and product_id and nozzle_code:
            # 兜底单查一次（可选）
            elem_name = get_element_name_by_product_and_nozzlecode(product_id, nozzle_code)

        if product_id and elem_name:
            # 批量读取该元件需要的所有 Name -> Value
            name2val = query_element_calc_values(product_id, elem_name)
            if name2val:
                for db_name, tpl_row_name in DBNAME_TO_TEMPLATE_ROW.items():
                    row = row_index.get(tpl_row_name)
                    if not row:
                        continue
                    _set_cell_safely(ws, row, out_col, name2val.get(db_name, ""))

    # 9) 美化表格
    # ---- ① 调整列宽 ----
    for i in range(len(nozzle_cols)):
        out_col = start_col + i
        col_letter = get_column_letter(out_col)
        ws.column_dimensions[col_letter].width = 17  # 可根据需要调整宽度
    # 调整 ”接管材料“ 所在的列宽
    ws.column_dimensions["B"].width = 25

    # ---- ② 给整个表格加边框 ----
    # 边框样式
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # 应用到当前表的所有单元格
    for r in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in r:
            cell.border = thin_border

    # ③ 重新合并第一行标题：从第1列到最后一列（包含单位和备注）
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
    title_cell = ws.cell(row=1, column=1)
    title_cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")

    # 10) 另存为：让用户选择保存路径和文件名（而不是固定到项目/exports）
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    suggested_name = f"NOZZLE_LISTING_导出_{ts}.xlsx"

    # 弹出“另存为”对话框
    out_path, _ = QFileDialog.getSaveFileName(
        stats_widget,  # 用你的窗口/控件作为父级
        "另存为",
        suggested_name,  # 默认文件名
        "Excel 工作簿 (*.xlsx)"  # 过滤器
    )

    # 用户取消
    if not out_path:
        return None  # 或者 raise RuntimeError("用户取消保存")

    # 补全后缀
    if not out_path.lower().endswith(".xlsx"):
        out_path += ".xlsx"

    # 保存
    try:
        wb.save(out_path)
    except PermissionError:
        QMessageBox.warning(stats_widget, "保存失败", "文件可能正在被占用，请关闭后重试。")
        return None

    return out_path


def _set_cell_safely(ws, row, col, value):
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        # 把值写到该合并区域的左上角
        for rng in ws.merged_cells.ranges:
            if cell.coordinate in rng:
                ws.cell(rng.min_row, rng.min_col, value)
                return
    else:
        cell.value = value


def _resolve_value_field_by_belong(pipe_belong: str):
    """
    根据“管口所属元件”判断取值列：
    - 包含“管箱” → 管程数值
    - 包含“壳体”或“外头盖” → 壳程数值
    其余情况返回 None
    """
    if not pipe_belong:
        return None
    if "管箱" in pipe_belong:
        return "管程数值"
    if ("壳体" in pipe_belong) or ("外头盖" in pipe_belong):
        return "壳程数值"
    return None


def _get_weld_joint_efficiency(product_id: str, pipe_belong: str):
    """
    读取产品设计活动库“产品设计活动表_设计数据表”中的参数“焊接接头系数*”。
    会根据“管口所属元件”自动选择 管程数值/壳程数值。
    返回 (ok: bool, value_or_msg: float|str)
    """
    value_field = _resolve_value_field_by_belong(pipe_belong)
    if not value_field:
        return False, "无效的管口所属元件"

    conn = None
    cursor = None
    try:
        conn = get_connection(**db_config_2)
        cursor = conn.cursor(pymysql.cursors.DictCursor)
        cursor.execute(f"""
            SELECT `{value_field}` AS v
            FROM 产品设计活动表_设计数据表
            WHERE 产品ID = %s AND 参数名称 = '焊接接头系数*'
            LIMIT 1
        """, (product_id,))
        row = cursor.fetchone()
        if not row or row.get("v") is None:
            return False, "未获取到焊接接头系数*"

        # 兜底转为 float
        try:
            return True, float(row["v"])
        except (TypeError, ValueError):
            return True, float(str(row["v"]).strip())
    except Exception as e:
        return False, f"数据库错误: {e}"
    finally:
        cursor and cursor.close()
        conn and conn.close()


def build_nozzlecode_to_element_map(product_id: str) -> dict:
    """
    为指定产品构建 {管口代号(value) -> 元件名称} 映射。
    数据来源：产品设计活动表_计算提交表
      - 条件：产品ID = product_id AND key = '管口表序号'
      - 取该行的 value 作为“管口代号”，同一行的 元件名称 作为映射值
    """
    if not product_id:
        return {}

    conn = None
    cur = None
    mapping = {}
    try:
        conn = get_connection(**db_config_2)
        cur = conn.cursor(pymysql.cursors.DictCursor)
        sql = """
            SELECT 元件名称, value
            FROM 产品设计活动表_计算提交表
            WHERE 产品ID = %s AND `key` = '管口表序号' AND value IS NOT NULL
        """
        cur.execute(sql, (product_id,))
        for row in cur.fetchall() or []:
            nozzle_code = (row.get("value") or "").strip()
            elem_name = (row.get("元件名称") or "").strip()
            if nozzle_code:
                mapping[nozzle_code] = elem_name
        return mapping
    except Exception as e:
        print(f"[导出] 构建代号→元件名称映射失败: {e}")
        return {}
    finally:
        cur and cur.close()
        conn and conn.close()


def get_element_name_by_product_and_nozzlecode(product_id: str, nozzle_code: str) -> str:
    """
    单次查询：由（产品ID, 管口代号）取“元件名称”。
    本质等价于在 产品设计活动表_计算提交表 中查找
      WHERE 产品ID=... AND key='管口表序号' AND value = nozzle_code
    """
    if not (product_id and nozzle_code):
        return ""
    conn = None
    cur = None
    try:
        conn = get_connection(**db_config_2)
        cur = conn.cursor(pymysql.cursors.DictCursor)
        sql = """
            SELECT 元件名称
            FROM 产品设计活动表_计算提交表
            WHERE 产品ID = %s AND `key` = '管口表序号' AND value = %s
            LIMIT 1
        """
        cur.execute(sql, (product_id, nozzle_code))
        row = cur.fetchone()
        return (row.get("元件名称") or "").strip() if row else ""
    except Exception as e:
        print(f"[导出] 代号→元件名称查询失败: {e}")
        return ""
    finally:
        cur and cur.close()
        conn and conn.close()


def query_element_calc_values(product_id: str, element_name: str, name_list=None) -> dict:
    """
    从《产品设计活动表_元件计算结果表》读取指定元件的多项计算结果。
    :param product_id: 产品ID
    :param element_name: 元件名称（如：壳程入口接管）
    :param name_list: 要取的 Name 列表；None=取 DBNAME_TO_TEMPLATE_ROW 的全部键
    :return: {Name: Value}；查不到的键不会出现在字典里
    """
    if not (product_id and element_name):
        return {}
    if name_list is None:
        name_list = list(DBNAME_TO_TEMPLATE_ROW.keys())

    conn = None
    cur = None
    out = {}
    try:
        conn = get_connection(**db_config_2)
        cur = conn.cursor(pymysql.cursors.DictCursor)
        # 用 IN 批量取，避免一条条查
        placeholders = ",".join(["%s"] * len(name_list))
        sql = f"""
            SELECT `Name`, `Value`
            FROM 产品设计活动表_元件计算结果表
            WHERE 产品ID = %s
              AND 元件名称 = %s
              AND `Name` IN ({placeholders})
        """
        cur.execute(sql, (product_id, element_name, *name_list))
        for row in cur.fetchall() or []:
            n = (row.get("Name") or "").strip()
            v = row.get("Value")
            out[n] = v if v is not None else ""
    except Exception as e:
        print(f"[导出] 读取《元件计算结果表》失败: {e}")
    finally:
        cur and cur.close()
        conn and conn.close()
    return out


# =========================
# 通用导入：table_template 下的"导入excel模板"
# =========================
"""清除当前界面数据"""
def _clear_pipe_table_except_last_row(stats_widget):
    """
    清空管口表格，保留最后一行空白行
    同时清除界面隐藏的管口ID映射，并删除数据库中该产品ID对应的管口数据
    """
    table = stats_widget.tableWidget_pipe
    # 获取当前行数
    current_rows = table.rowCount()

    # 如果只有一行（空白行），直接返回
    if current_rows <= 1:
        return

    # 1. 清除界面隐藏的管口ID映射
    try:
        from modules.guankoudingyi.funcs.funcs_pipe_table import ensure_hidden_maps
        ensure_hidden_maps(stats_widget)

        # 清除隐藏管口ID映射
        if hasattr(stats_widget, 'row_hidden_pipe_id'):
            stats_widget.row_hidden_pipe_id.clear()


        # 清除待删除管口ID集合
        if hasattr(stats_widget, 'deleted_pipe_ids'):
            stats_widget.deleted_pipe_ids.clear()


    except Exception as e:
        print(f"[WARNING] 清除隐藏管口ID映射失败: {str(e)}")

    # 2. 删除数据库中该产品ID对应的管口数据
    try:
        product_id = getattr(stats_widget, 'product_id', None)
        if product_id:
            conn = get_connection(**db_config_2)
            cursor = conn.cursor()

            try:
                # 删除产品设计活动表_管口表中的数据
                cursor.execute("DELETE FROM 产品设计活动表_管口表 WHERE 产品ID = %s", (product_id,))
                pipe_deleted_count = cursor.rowcount

                # 删除产品设计活动表_管口类别表中的数据
                cursor.execute("DELETE FROM 产品设计活动表_管口类别表 WHERE 产品ID = %s", (product_id,))
                category_deleted_count = cursor.rowcount

                # 删除产品设计活动表_管口类别表中的数据
                cursor.execute("DELETE FROM 产品设计活动表_管口类型选择表 WHERE 产品ID = %s", (product_id,))
                category_deleted_count = cursor.rowcount

                # 删除产品设计活动表_管口载荷表中的数据
                cursor.execute("DELETE FROM 产品设计活动表_管口载荷表 WHERE 产品ID = %s", (product_id,))
                category_deleted_count = cursor.rowcount

                # 提交事务
                conn.commit()


            except Exception as e:
                conn.rollback()

            finally:
                cursor.close()
                conn.close()
        else:
            print("[WARNING] 未获取到产品ID，跳过数据库删除操作")

    except Exception as e:
        print(f"[WARNING] 数据库删除操作失败: {str(e)}")

    # 3. 删除除最后一行外的所有行
    for row in range(current_rows - 2, -1, -1):  # 从倒数第二行开始，到第0行结束
        table.removeRow(row)

    # 4. 刷新序号
    if hasattr(stats_widget, "refresh_pipe_table_sequence"):
        stats_widget.refresh_pipe_table_sequence()


"""模板格式校验函数"""
def validate_excel_template_format(worksheet):
    """
    校验Excel模板格式是否正确
    检查标题行是否包含预期的字段（不要求完全匹配）

    :param worksheet: Excel工作表对象
    :return: (is_valid: bool, error_messages: list)
    """
    error_messages = []

    try:
        print(f"[DEBUG] 开始校验模板格式，最大行数: {worksheet.max_row}, 最大列数: {worksheet.max_column}")

        # 检查是否有足够的列
        if worksheet.max_column < 19:
            error_messages.append(f"模板列数不足，期望至少19列，实际{worksheet.max_column}列")
            print(f"[DEBUG] 列数检查失败: 期望19列，实际{worksheet.max_column}列")

        # 检查是否有标题行（至少2行）
        if worksheet.max_row < 2:
            error_messages.append("模板缺少标题行，至少需要2行标题")
            print(f"[DEBUG] 标题行检查失败: 期望至少2行，实际{worksheet.max_row}行")

        # 预期的列标题关键词（不要求完全匹配，只要包含这些关键词即可）
        expected_keywords = {
            # 第1行标题关键词
            1: {
                2: ["管口代号"],
                3: ["管口功能"],
                4: ["管口用途"],
                5: ["公称尺寸"],
                18: ["管口附件"],
                19: ["管口载荷"]
            },
            # 第2行标题关键词
            2: {
                6: ["法兰标准"],
                7: ["压力等级"],
                8: ["法兰型式"],
                9: ["密封面型式"],
                10: ["焊端规格"],
                11: ["管口所属元件"],
                12: ["轴向定位基准"],
                13: ["轴向定位距离"],
                14:["轴向夹角"],
                15: ["周向方位"],
                16: ["偏心距"],
                17: ["外伸高度"],

            }
        }

        # 检查第1行标题
        for col, keywords in expected_keywords[1].items():
            if col <= worksheet.max_column:  # 只检查存在的列
                cell_value = worksheet.cell(row=1, column=col).value
                actual_header = str(cell_value).strip() if cell_value else ""

                # 检查是否包含预期的关键词
                found_keyword = False
                for keyword in keywords:
                    if keyword in actual_header:
                        found_keyword = True
                        break

                if not found_keyword:
                    error_messages.append(f"第1行第{col}列标题不正确：期望包含'{keywords[0]}'，实际'{actual_header}'")
                    print(f"[DEBUG] 第1行第{col}列标题检查失败: 期望包含'{keywords[0]}'，实际'{actual_header}'")

        # 检查第2行标题
        for col, keywords in expected_keywords[2].items():
            if col <= worksheet.max_column:  # 只检查存在的列
                cell_value = worksheet.cell(row=2, column=col).value
                actual_header = str(cell_value).strip() if cell_value else ""

                # 检查是否包含预期的关键词
                found_keyword = False
                for keyword in keywords:
                    if keyword in actual_header:
                        found_keyword = True
                        break

                if not found_keyword:
                    error_messages.append(f"第2行第{col}列标题不正确：期望包含'{keywords[0]}'，实际'{actual_header}'")
                    print(f"[DEBUG] 第2行第{col}列标题检查失败: 期望包含'{keywords[0]}'，实际'{actual_header}'")

        is_valid = len(error_messages) == 0
        print(f"[DEBUG] 模板格式校验结果: {'通过' if is_valid else '失败'}, 错误数量: {len(error_messages)}")

    except Exception as e:
        error_messages.append(f"模板格式校验过程中发生错误：{str(e)}")
        print(f"[DEBUG] 模板格式校验异常: {str(e)}")
        is_valid = False

    return is_valid, error_messages


"""从错误信息中提取行号数字，用于排序"""
def _extract_row_number_for_sort(error_msg):
    """
    从错误信息字符串中提取行号数字，用于按数字大小排序
    :param error_msg: 错误信息字符串，格式如 "第1行数据不合法" 或 "第10行，请先在条件输入界面填写公称直径"
    :return: 行号数字，如果无法提取则返回0
    """
    import re
    match = re.search(r'第(\d+)行', error_msg)
    if match:
        return int(match.group(1))
    return 0  # 如果无法提取，返回0作为默认值


"""验证管口附件（第18列）"""
def validate_pipe_attachment(attachment_value, row):
    """
    验证管口附件是否合法：
    允许的值：接管法兰配对法兰、接管拉筋、防冲挡板、破涡器
    支持多选，多个附件用分号";"分隔，如："接管法兰配对法兰;接管拉筋;破涡器"
    :param attachment_value: 原始附件文本
    :param row: 行号（用于错误信息）
    :return: (验证后的值, 错误信息列表)
    """
    errors = []
    if not attachment_value:
        return attachment_value, errors

    allowed_attachments = {"接管法兰配对法兰", "接管拉筋", "防冲挡板", "破涡器"}
    attachment_text = str(attachment_value).strip()
    
    # 如果为空字符串，直接返回
    if not attachment_text:
        return attachment_text, errors
    
    # 支持多选：用分号";"分隔多个附件
    # 先按分号拆分，去除每个附件的前后空格
    attachment_list = [item.strip() for item in attachment_text.split(";") if item.strip()]
    
    # 如果没有有效的附件，返回错误
    if not attachment_list:
        errors.append(f"管口附件列，第{row}行数据不合法")
        return "", errors
    
    # 验证每个附件是否在允许的列表中
    invalid_attachments = []
    for attachment in attachment_list:
        if attachment not in allowed_attachments:
            invalid_attachments.append(attachment)
    
    # 如果有不合法的附件，返回错误
    if invalid_attachments:
        errors.append(f"管口附件列，第{row}行数据不合法（包含不支持的附件：{', '.join(invalid_attachments)}）")
        return "", errors
    
    # 所有附件都合法，返回用分号连接的值（去除重复，保持顺序）
    # 使用 OrderedDict 保持顺序并去重
    seen = set()
    unique_attachments = []
    for attachment in attachment_list:
        if attachment not in seen:
            seen.add(attachment)
            unique_attachments.append(attachment)
    
    return ";".join(unique_attachments), errors


"""从Excel模板导入管口数据"""
def import_nozzle_from_excel(stats_widget):
    """
    从Excel模板导入管口数据
    使用table_template下的导入模板.xlsx作为模板
    """
    try:
        # 1. 获取模板文件路径
        proj_root = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
        template_path = os.path.join(proj_root, "guankoudingyi", "table_template", "导入模板.xlsx")

        # if not os.path.exists(template_path):
        #     QMessageBox.warning(stats_widget, "模板文件不存在", f"未找到导入模板文件：\n{template_path}")
        #     return False

        # 2. 弹出文件选择对话框
        file_path, _ = QFileDialog.getOpenFileName(
            stats_widget,
            "选择要导入的Excel文件",
            "",
            "Excel文件 (*.xlsx *.xls);;所有文件 (*)"
        )

        if not file_path:
            return False

        # 3. 读取Excel文件
        try:
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active
        except Exception as e:
            QMessageBox.critical(stats_widget, "文件读取失败", f"无法读取Excel文件：\n{str(e)}")
            return False

        # 4. 校验导入文件格式
        is_valid_format, format_errors = validate_excel_template_format(ws)
        if not is_valid_format:
            QMessageBox.warning(stats_widget, "导入失败", "导入模板失败")
            return False

        # 5. 先解析Excel数据，检查是否有有效的管口数据
        try:
            product_id = getattr(stats_widget, 'product_id', None)
            imported_data, template_duplicates, validation_errors = _parse_excel_data(ws, product_id)
        except Exception as e:
            QMessageBox.critical(stats_widget, "数据解析失败", f"解析Excel数据时发生错误：\n{str(e)}")
            return False

        # 6. 如果没有有效的管口数据，直接返回
        if not imported_data:
            QMessageBox.warning(stats_widget, "导入失败", "Excel文件中没有找到有效的管口数据")
            return False

        # 7. 检查并切换单位类型（如果需要）
        try:
            _check_and_switch_unit_types(stats_widget, ws)
        except Exception as e:
            print(f"单位类型检查失败: {e}")

        # 8. 显示验证错误和重复信息（合并显示）
        if validation_errors or template_duplicates:
            message_parts = []

            # 添加重复信息
            if template_duplicates:
                duplicate_text = "\n".join(template_duplicates)
                message_parts.append(f"管口代号列：发现以下重复的管口代号，已跳过重复项：\n{duplicate_text}")

            # 添加验证错误信息
            if validation_errors:

                # 按列分组错误信息，使用集合避免重复
                nominal_size_errors = set()
                weld_end_spec_errors = set()
                pipe_belong_axial_errors = set()  # 合并管口所属元件和轴向定位基准错误
                axial_distance_errors = set()  # 轴向定位距离错误
                axial_angle_errors = set()  # 轴向夹角错误
                circumferential_position_errors = set()  # 周向方位错误
                eccentricity_errors = set()  # 偏心距错误
                extension_height_errors = set()  # 外伸高度错误
                flange_standard_errors = set()  # 法兰标准错误
                pressure_level_errors = set()  # 压力等级错误
                flange_form_errors = set()  # 法兰型式错误
                sealing_face_form_errors = set()  # 密封面型式错误

                for error in validation_errors:
                    if "公称尺寸" in error:
                        # 提取行号并减2（Excel第3行对应界面第1行）
                        row_num = int(error.split("第")[1].split("行")[0]) - 2
                        nominal_size_errors.add(f"第{row_num}行数据不合法")
                    elif "法兰标准" in error:
                        # 提取行号并减2
                        row_num = int(error.split("第")[1].split("行")[0]) - 2
                        flange_standard_errors.add(f"第{row_num}行数据不合法")
                    elif "压力等级" in error:
                        # 提取行号并减2
                        row_num = int(error.split("第")[1].split("行")[0]) - 2
                        pressure_level_errors.add(f"第{row_num}行数据不合法")
                    elif "法兰型式" in error:
                        # 提取行号并减2
                        row_num = int(error.split("第")[1].split("行")[0]) - 2
                        flange_form_errors.add(f"第{row_num}行数据不合法")
                    elif "密封面型式" in error:
                        # 提取行号并减2
                        row_num = int(error.split("第")[1].split("行")[0]) - 2
                        sealing_face_form_errors.add(f"第{row_num}行数据不合法")
                    elif "焊端规格" in error:
                        # 提取行号并减2
                        row_num = int(error.split("第")[1].split("行")[0]) - 2
                        weld_end_spec_errors.add(f"第{row_num}行数据不合法")
                    elif "管口所属元件" in error or "轴向定位基准" in error:
                        # 提取行号并减2
                        row_num = int(error.split("第")[1].split("行")[0]) - 2
                        pipe_belong_axial_errors.add(f"第{row_num}行数据不合法")
                    elif "轴向定位距离" in error:
                        # 提取行号并减2
                        row_num = int(error.split("第")[1].split("行")[0]) - 2
                        axial_distance_errors.add(f"第{row_num}行数据不合法")
                    elif "轴向夹角" in error:
                        # 提取行号并减2
                        row_num = int(error.split("第")[1].split("行")[0]) - 2
                        axial_angle_errors.add(f"第{row_num}行数据不合法")
                    elif "周向方位" in error:
                        # 提取行号并减2
                        row_num = int(error.split("第")[1].split("行")[0]) - 2
                        circumferential_position_errors.add(f"第{row_num}行数据不合法")
                    elif "偏心距" in error:
                        # 提取行号并减2
                        row_num = int(error.split("第")[1].split("行")[0]) - 2
                        # 检查是否是特殊提示信息
                        if "请先在条件输入界面填写公称直径" in error:
                            eccentricity_errors.add(f"第{row_num}行，请先在条件输入界面填写公称直径")
                        elif "偏心距与夹角不能同时填写非零值" in error:
                            eccentricity_errors.add(f"第{row_num}行，轴向夹角和偏心距不能同时赋值")
                        else:
                            eccentricity_errors.add(f"第{row_num}行数据不合法")
                    elif "外伸高度" in error:
                        # 提取行号并减2
                        row_num = int(error.split("第")[1].split("行")[0]) - 2
                        # 检查是否是特殊提示信息
                        if "请先在条件输入界面填写公称直径" in error:
                            extension_height_errors.add(f"第{row_num}行，请先在条件输入界面填写公称直径")
                        else:
                            extension_height_errors.add(f"第{row_num}行数据不合法")


                # 添加公称尺寸错误信息
                if nominal_size_errors:
                    message_parts.append(f"公称尺寸列：{', '.join(sorted(nominal_size_errors, key=_extract_row_number_for_sort))}")

                if flange_standard_errors:
                    message_parts.append(f"法兰标准列：{', '.join(sorted(flange_standard_errors, key=_extract_row_number_for_sort))}")

                    # 添加压力等级错误信息
                if pressure_level_errors:
                    message_parts.append(f"压力等级列：{', '.join(sorted(pressure_level_errors, key=_extract_row_number_for_sort))}")

                    # 添加法兰型式错误信息
                if flange_form_errors:
                    message_parts.append(f"法兰型式列：{', '.join(sorted(flange_form_errors, key=_extract_row_number_for_sort))}")

                    # 添加密封面型式错误信息
                if sealing_face_form_errors:
                    message_parts.append(f"密封面型式列：{', '.join(sorted(sealing_face_form_errors, key=_extract_row_number_for_sort))}")

                # 添加焊端规格错误信息
                if weld_end_spec_errors:
                    message_parts.append(f"焊端规格列：{', '.join(sorted(weld_end_spec_errors, key=_extract_row_number_for_sort))}")

                # 添加管口所属元件和轴向定位基准错误信息（合并显示）
                if pipe_belong_axial_errors:
                    message_parts.append(f"管口所属元件和轴向定位基准列：{', '.join(sorted(pipe_belong_axial_errors, key=_extract_row_number_for_sort))}")

                # 添加轴向定位距离错误信息
                if axial_distance_errors:
                    message_parts.append(f"轴向定位距离列：{', '.join(sorted(axial_distance_errors, key=_extract_row_number_for_sort))}")


                # 添加轴向夹角错误信息
                if axial_angle_errors:
                    message_parts.append(f"轴向夹角列：{', '.join(sorted(axial_angle_errors, key=_extract_row_number_for_sort))}")

                # 添加周向方位错误信息
                if circumferential_position_errors:
                    message_parts.append(f"周向方位列：{', '.join(sorted(circumferential_position_errors, key=_extract_row_number_for_sort))}")

                # 添加偏心距错误信息
                if eccentricity_errors:
                    print(f"调试：偏心距错误集合内容: {eccentricity_errors}")
                    message_parts.append(f"偏心距列：{', '.join(sorted(eccentricity_errors, key=_extract_row_number_for_sort))}")

                # 添加外伸高度错误信息
                if extension_height_errors:
                    message_parts.append(f"外伸高度列：{', '.join(sorted(extension_height_errors, key=_extract_row_number_for_sort))}")

            # 合并显示
            if message_parts and hasattr(stats_widget, 'line_tip'):
                combined_message = "\n".join(message_parts)
                stats_widget.line_tip.setText(combined_message)
                stats_widget.line_tip.setStyleSheet("color: orange;")
                stats_widget.line_tip.setToolTip(f"{combined_message}")

        # 9. 清除当前界面数据（保留最后一行空白行）
        # 在清空界面前，将 cannot_be_deleted 设为 False，允许删除所有管口
        try:
            stats_widget.cannot_be_deleted = False
            _clear_pipe_table_except_last_row(stats_widget)
        except Exception as e:
            QMessageBox.critical(stats_widget, "界面清除失败", f"清除界面数据时发生错误：\n{str(e)}")
            # 即使出错也要恢复 cannot_be_deleted 标志
            stats_widget.cannot_be_deleted = True
            return False

        # 10. 将数据填充到界面
        try:
            _fill_data_to_ui(stats_widget, imported_data)
        except Exception as e:
            QMessageBox.critical(stats_widget, "数据填充失败", f"填充数据到界面时发生错误：\n{str(e)}")
            return False

        # 11. 设置管口功能列只读状态
        # try:
        #     from modules.guankoudingyi.funcs.funcs_pipe_table import set_pipe_function_column_readonly
        #     set_pipe_function_column_readonly(stats_widget)
        # except Exception as e:
        #     print(f"[WARNING] 设置管口功能列只读状态失败: {str(e)}")

        # 11.5. 导入完成后，重新设置 cannot_be_deleted 为 True，并设置默认管口不可删除
        try:
            from modules.guankoudingyi.funcs.funcs_pipe_table import set_default_pipe_cannot_be_deleted
            set_default_pipe_cannot_be_deleted(stats_widget)
        except Exception as e:
            print(f"[WARNING] 设置默认管口不可删除状态失败: {str(e)}")
            # 即使出错也要恢复 cannot_be_deleted 标志
            stats_widget.cannot_be_deleted = True

        # 12. 显示导入成功信息
        QMessageBox.information(stats_widget, "导入成功", f"成功导入 {len(imported_data)} 条管口数据")
        return True

    except Exception as e:
        QMessageBox.critical(stats_widget, "导入失败", f"导入过程中发生错误：\n{str(e)}")
        return False


"""解析Excel工作表数据，提取管口信息"""
def _parse_excel_data(worksheet, product_id=None):
    """
    解析Excel工作表数据，提取管口信息
    根据模板结构：
    - 第2-5列，第1行：管口代号、管口功能、管口用途、公称尺寸
    - 第6-17列，第2行：法兰标准到外伸高度（12个字段）
    - 第18-19列，第1行：管口附件、管口载荷
    返回管口数据列表、重复信息和验证错误信息
    """
    imported_data = []
    pipe_codes = set()  # 用于检查管口代号重复
    duplicate_info = []  # 记录重复信息
    validation_errors = []  # 记录验证错误信息

    try:
        # 查找数据开始行（从第3行开始，前两行是标题）
        data_start_row = 3
        max_row = worksheet.max_row

        # 预先计算，获取整表“公称尺寸”列（第5列）位于管箱圆筒/外头盖圆筒上的管口的最大 公称尺寸对应的接管实际外径od 的值，用于“管箱圆筒/外头盖圆筒”的上限校验
        max_nominal_size_od = None
        try:
            unit_type_col5 = get_column_unit_type(worksheet, 5)
            if unit_type_col5 in ["DN", "NPS"]:
                # 延迟导入：从选项模块将公称尺寸转为接管实际外径od
                from modules.guankoudingyi.funcs.funcs_pipe_comboBox_value import get_component_nominal_size_od
                for r_cal in range(3, max_row + 1):
                    # 先获取该行的管口所属元件（第11列）
                    pipe_belong_raw = _get_cell_value(worksheet, r_cal, 11)
                    if not pipe_belong_raw:
                        continue

                    pipe_belong = str(pipe_belong_raw).strip()
                    # 只处理管箱圆筒和外头盖圆筒的管口
                    if "管箱圆筒" not in pipe_belong and "外头盖圆筒" not in pipe_belong:
                        continue

                    # 获取公称尺寸并转换为接管实际外径od
                    v = _get_cell_value(worksheet, r_cal, 5)
                    if not v:
                        continue
                    od_val = get_component_nominal_size_od(v, product_id=product_id, stats_widget=None, size_type_override=unit_type_col5)
                    if isinstance(od_val, (int, float)):
                        if (max_nominal_size_od is None) or (od_val > max_nominal_size_od):
                            max_nominal_size_od = od_val
        except Exception:
            max_nominal_size_od = None

        # 遍历数据行
        for row in range(data_start_row, max_row + 1):
            # 读取管口代号（第2列，第1行标题）
            pipe_code_cell = worksheet.cell(row=row, column=2)
            pipe_code = str(pipe_code_cell.value).strip() if pipe_code_cell.value else ""

            # 如果管口代号为空，跳过这一行
            if not pipe_code or pipe_code == "None":
                continue

            # 检查管口代号是否在模板内重复
            if pipe_code in pipe_codes:
                duplicate_info.append(f"第{row - 2}行：'{pipe_code}'")
                continue  # 跳过重复的管口代号

            # 添加到已存在的代号集合中
            pipe_codes.add(pipe_code)

            # 获取法兰标准并验证
            flange_standard_raw = _get_cell_value(worksheet, row, 6)  # 第6列
            flange_standard_validated = ""

            # 获取压力等级列的单位类型（用于验证法兰标准）
            pressure_unit_type = get_column_unit_type(worksheet, 7)  # 第7列

            # 验证法兰标准（传入压力类型）
            flange_standard_validated, flange_standard_error_messages = validate_flange_standard_with_error_info(
                flange_standard_raw, row, pressure_unit_type
            )
            validation_errors.extend(flange_standard_error_messages)

            # 检查法兰标准是否验证通过
            flange_standard_valid = bool(flange_standard_validated and flange_standard_validated.strip())

            # 获取压力等级并验证
            pressure_level_raw = _get_cell_value(worksheet, row, 7)  # 第7列
            pressure_level_validated = ""

            # 如果法兰标准验证通过，才验证压力等级
            if flange_standard_valid and pressure_level_raw:
                # 根据法兰标准确定单位类型
                if flange_standard_validated in ["HG/T 20615-2009", "HG/T 20623-2009(A)", "HG/T 20623-2009(B)","SH/T 3406-2022","SH/T 3406-2022(A)","SH/T 3406-2022(B)"]:
                    # 这些标准要求使用 Class 单位
                    required_unit_type = "Class"
                elif flange_standard_validated == "HG/T 20592-2009":
                    # HG/T 20592-2009 标准要求使用 PN 单位
                    required_unit_type = "PN"
                else:
                    # 其他情况，使用空字符串
                    pressure_level_validated = ""
                    required_unit_type = None

                if required_unit_type:
                    # 获取压力等级列的单位类型
                    actual_unit_type = pressure_unit_type

                    # 检查单位类型是否匹配
                    if actual_unit_type == required_unit_type:
                        # 单位类型匹配，进行压力等级值验证（传入法兰标准）
                        pressure_level_validated, pressure_level_error_messages = validate_pressure_level_with_error_info(
                            pressure_level_raw, required_unit_type, row, flange_standard_validated
                        )
                        validation_errors.extend(pressure_level_error_messages)
                    else:
                        # 单位类型不匹配，设为空并记录错误
                        pressure_level_validated = ""
                        validation_errors.append(f"压力等级列，第{row}行类型不合法")
            else:
                # 法兰标准验证失败时，压力等级设为空
                pressure_level_validated = ""

            # 获取法兰型式并验证
            flange_form_raw = _get_cell_value(worksheet, row, 8)  # 第8列
            flange_form_validated = ""

            # 验证法兰型式
            if flange_standard_valid and pressure_level_validated and flange_form_raw:
                # 获取压力等级列的单位类型
                pressure_level_type = get_column_unit_type(worksheet, 7)
                if pressure_level_type in ["Class", "PN"]:
                    flange_form_validated, flange_form_error_messages = validate_flange_form_by_database(
                        flange_form_raw, flange_standard_validated, pressure_level_validated, pressure_level_type, row
                    )
                    validation_errors.extend(flange_form_error_messages)
                else:
                    flange_form_validated = ""
            else:
                # 法兰标准或压力等级验证失败时，法兰型式设为空
                flange_form_validated = ""

            # 获取密封面型式并验证
            sealing_face_form_raw = _get_cell_value(worksheet, row, 9)  # 第9列
            sealing_face_form_validated = ""

            # 验证密封面型式
            if flange_standard_valid and pressure_level_validated and flange_form_validated and sealing_face_form_raw:
                # 获取压力等级列的单位类型
                pressure_level_type = get_column_unit_type(worksheet, 7)
                if pressure_level_type in ["Class", "PN"]:
                    sealing_face_form_validated, sealing_face_form_error_messages = validate_sealing_face_form_by_database(
                        sealing_face_form_raw, flange_standard_validated, pressure_level_validated, pressure_level_type, flange_form_validated, row
                    )
                    validation_errors.extend(sealing_face_form_error_messages)
                else:
                    sealing_face_form_validated = ""
            else:
                # 前置条件不满足时，密封面型式设为空
                sealing_face_form_validated = ""

            # 获取公称尺寸并验证
            nominal_size_raw = _get_cell_value(worksheet, row, 5)  # 第5列
            nominal_size_validated = ""

            if nominal_size_raw and product_id:
                # 获取公称尺寸列的单位类型
                unit_type = get_column_unit_type(worksheet, 5)
                if unit_type in ["DN", "NPS"]:
                    # 验证公称尺寸是否在数据库中存在，根据法兰标准进行筛选
                    nominal_size_validated = validate_nominal_size_by_unit(nominal_size_raw, unit_type, product_id, flange_standard_validated)
                    # 如果验证后为空，说明数据不合法，记录错误信息
                    if not nominal_size_validated and nominal_size_raw:
                        validation_errors.append(f"公称尺寸第{row}行数据不合法")
                else:
                    nominal_size_validated = nominal_size_raw
            else:
                nominal_size_validated = nominal_size_raw

            # 获取焊端规格并验证
            weld_end_spec_raw = _get_cell_value(worksheet, row, 10)  # 第10列
            weld_end_spec_validated = ""

            if weld_end_spec_raw and product_id:
                # 获取焊端规格列的单位类型
                unit_type = get_column_unit_type(worksheet, 10)
                if unit_type in ["Sch", "mm"]:
                    # 验证焊端规格是否合法
                    weld_end_spec_validated = validate_weld_end_spec_by_unit(weld_end_spec_raw, unit_type, product_id)
                    # 如果验证后为空，说明数据不合法，记录错误信息
                    if not weld_end_spec_validated and weld_end_spec_raw:
                        validation_errors.append(f"焊端规格列，第{row}行数据不合法")
                else:
                    weld_end_spec_validated = weld_end_spec_raw
            else:
                weld_end_spec_validated = weld_end_spec_raw

            # 获取管口所属元件并验证
            pipe_belong_raw = _get_cell_value(worksheet, row, 11)  # 第11列
            pipe_belong_validated = ""
            
              # 获取管口功能（用于互斥逻辑）
            pipe_function_raw = _get_cell_value(worksheet, row, 3)  # 第3列

            if pipe_belong_raw and product_id:
                # 验证管口所属元件是否合法
                pipe_belong_validated = validate_pipe_belong_by_product_type(pipe_belong_raw, product_id, pipe_function_raw)
                # 如果验证后为空，说明数据不合法，记录错误信息
                if not pipe_belong_validated and pipe_belong_raw:
                    validation_errors.append(f"管口所属元件列，第{row}行数据不合法")
            else:
                pipe_belong_validated = pipe_belong_raw

            # 锥壳：校验公称尺寸是否超出锥壳长度，超出则公称尺寸置空并按既有格式报错
            if pipe_belong_validated and ("锥壳" in pipe_belong_validated) and nominal_size_validated:
                current_od = get_component_nominal_size_od(
                    nominal_size_validated, product_id=product_id, stats_widget=None
                )
                tube_ok, tube_nominal_diameter = get_nominal_diameter(product_id, "管箱")
                shell_ok, shell_nominal_diameter = get_nominal_diameter(product_id, "壳体")
                if (not tube_ok) or (tube_nominal_diameter is None):
                    tube_nominal_diameter = 300
                if (not shell_ok) or (shell_nominal_diameter is None):
                    shell_nominal_diameter = 400
                cone_length = (shell_nominal_diameter - tube_nominal_diameter) / math.tan(math.radians(30))
                if cone_length < 0:
                    cone_length = 0

                if isinstance(current_od, (int, float)) and current_od > cone_length:
                    nominal_size_validated = ""
                    validation_errors.append(f"公称尺寸第{row}行数据不合法")

          

            # 获取轴向定位基准并验证
            axial_position_base_raw = _get_cell_value(worksheet, row, 12)  # 第12列
            axial_position_base_validated = ""

            # 如果管口所属元件未填入，轴向定位基准置空
            if not pipe_belong_validated:
                axial_position_base_validated = ""
            elif axial_position_base_raw and pipe_belong_validated:
                # 验证轴向定位基准是否合法（包含互斥逻辑）
                axial_position_base_validated = validate_axial_position_base(
                    axial_position_base_raw,
                    pipe_belong_validated,
                    pipe_function_raw,
                    len(imported_data),  # 当前行在imported_data中的索引
                    imported_data
                )
                # 如果验证后为空，说明数据不合法，记录错误信息
                if not axial_position_base_validated and axial_position_base_raw:
                    # 检查是否是因为互斥问题
                    if pipe_function_raw in ["壳程入口", "壳程出口"] and axial_position_base_raw in ["左基准线", "右基准线"]:
                        validation_errors.append(f"轴向定位基准列，第{row}行数据不合法（壳程入口和壳程出口的轴向定位基准必须互斥）")
                    else:
                        validation_errors.append(f"轴向定位基准列，第{row}行数据不合法")
            else:
                axial_position_base_validated = axial_position_base_raw

            # 获取轴向定位距离并验证
            axial_distance_raw = _get_cell_value(worksheet, row, 13)  # 第13列
            axial_distance_validated = ""

            # 验证轴向定位距离
            if axial_distance_raw:
                axial_distance_validated = validate_axial_position_distance(
                    axial_distance_raw,
                    nominal_size_validated,
                    pipe_belong_validated,
                    product_id=product_id,
                    stats_widget=None,
                    max_nominal_size_od=max_nominal_size_od
                )
                # 如果验证后为空，说明数据不合法，记录错误信息
                if not axial_distance_validated and axial_distance_raw:
                    validation_errors.append(f"轴向定位距离列，第{row}行数据不合法")
            else:
                axial_distance_validated = axial_distance_raw

            # 获取轴向夹角、周向方位、偏心距并验证
            axial_angle_raw = _get_cell_value(worksheet, row, 14)  # 第14列
            circumferential_position_raw = _get_cell_value(worksheet, row, 15)  # 第15列
            eccentricity_raw = _get_cell_value(worksheet, row, 16)  # 第16列

            axial_angle_validated = ""
            circumferential_position_validated = ""
            eccentricity_validated = ""

            # 当管口所属元件为管板类（固定管板/前端管板/后端管板）时，统一导入为“—”，方便界面侧直接识别为禁用状态
            if pipe_belong_validated in ["固定管板", "前端管板", "后端管板"]:
                axial_angle_validated = "—"
                circumferential_position_validated = "—"
                eccentricity_validated = "—"
            else:
                # 验证轴向夹角
                if axial_angle_raw:
                    axial_angle_validated = validate_axial_angle(axial_angle_raw)
                    # 如果验证后为空，说明数据不合法，记录错误信息
                    if not axial_angle_validated and axial_angle_raw:
                        validation_errors.append(f"轴向夹角列，第{row}行数据不合法")
                else:
                    axial_angle_validated = axial_angle_raw

                # 验证周向方位
                if circumferential_position_raw:
                    circumferential_position_validated = validate_circumferential_position(circumferential_position_raw)
                    # 如果验证后为空，说明数据不合法，记录错误信息
                    if not circumferential_position_validated and circumferential_position_raw:
                        validation_errors.append(f"周向方位列，第{row}行数据不合法")
                else:
                    circumferential_position_validated = circumferential_position_raw

                # 验证偏心距
                eccentricity_validated, eccentricity_error_messages = validate_eccentricity_with_error_info(
                    eccentricity_raw, pipe_belong_validated, product_id, row, axial_angle_raw
                )
                if eccentricity_error_messages:
                    validation_errors.extend(eccentricity_error_messages)

            # 获取外伸高度并验证
            extension_height_raw = _get_cell_value(worksheet, row, 17)  # 第17列
            extension_height_validated = ""

            # 验证外伸高度
            extension_height_validated, extension_height_error_messages = validate_extension_height_with_error_info(
                extension_height_raw, pipe_belong_validated, product_id, row
            )
            validation_errors.extend(extension_height_error_messages)

            # 获取管口附件并验证（第17列）
            pipe_attachment_raw = _get_cell_value(worksheet, row, 18)
            pipe_attachment_validated, attachment_errors = validate_pipe_attachment(pipe_attachment_raw, row)
            validation_errors.extend(attachment_errors)

            # 构建管口数据字典
            pipe_data = {
                # 第2-5列，第1行标题对应的数据
                "管口代号": pipe_code,
                "管口功能": _get_cell_value(worksheet, row, 3),  # 第3列
                "管口用途": _get_cell_value(worksheet, row, 4),  # 第4列
                "公称尺寸": nominal_size_validated,  # 验证后的公称尺寸

                # 第6-17列，第2行标题对应的数据（法兰相关参数）
                "法兰标准": flange_standard_validated,  # 验证后的法兰标准
                "压力等级": pressure_level_validated,  # 验证后的压力等级
                "法兰型式": flange_form_validated,  # 验证后的法兰型式
                "密封面型式": sealing_face_form_validated,  # 验证后的密封面型式
                "焊端规格": weld_end_spec_validated,  # 验证后的焊端规格
                "管口所属元件": pipe_belong_validated,  # 验证后的管口所属元件
                "轴向定位基准": axial_position_base_validated,  # 验证后的轴向定位基准
                "轴向定位距离": axial_distance_validated,  # 验证后的轴向定位距离
                "轴向夹角（°）": axial_angle_validated,  # 验证后的轴向夹角
                "周向方位（°）": circumferential_position_validated,  # 验证后的周向方位
                "偏心距": eccentricity_validated,  # 验证后的偏心距
                "外伸高度": extension_height_validated,  # 验证后的外伸高度

                # 第18-19列，第1行标题对应的数据
                "管口附件": pipe_attachment_validated,  # 验证后的管口附件（第18列）
                "管口载荷": _get_cell_value(worksheet, row, 19)  # 第19列
            }

            imported_data.append(pipe_data)

        return imported_data, duplicate_info, validation_errors

    except Exception as e:
        raise e  # 重新抛出异常，让调用者处理


"""获取单元格数据"""
def _get_cell_value(worksheet, row, col):
    """
    安全获取单元格值
    """
    try:
        cell = worksheet.cell(row=row, column=col)
        if cell.value is None:
            return ""
        return str(cell.value).strip()
    except:
        return ""


"""填充数据到界面"""
def _fill_data_to_ui(stats_widget, imported_data):
    """
    将导入的数据填充到界面表格中
    """
    try:
        table = stats_widget.tableWidget_pipe
        if not table:
            raise Exception("无法获取表格控件")

        # 确保有足够的行数
        current_rows = table.rowCount()
        needed_rows = len(imported_data)

        # 如果当前只有一行空白行，需要添加更多行
        if current_rows < needed_rows + 1:  # +1 是为了保留最后的空白行
            for _ in range(needed_rows + 1 - current_rows):
                table.insertRow(table.rowCount())

        # 临时断开信号，防止填充过程中触发事件
        try:
            table.blockSignals(True)

            # 填充数据
            for i, pipe_data in enumerate(imported_data):
                # 设置序号
                seq_item = QTableWidgetItem(str(i + 1))
                seq_item.setTextAlignment(Qt.AlignCenter)
                seq_item.setFlags(seq_item.flags() & ~Qt.ItemIsEditable)  # 序号列不可编辑
                table.setItem(i, 0, seq_item)

                # 填充各列数据
                for field_name, col_index in UI_COL_INDEX.items():
                    value = pipe_data.get(field_name, "")
                    item = QTableWidgetItem(str(value))
                    item.setTextAlignment(Qt.AlignCenter)
                    table.setItem(i, col_index, item)

            # 填充完所有数据后，检查是否需要添加新行
            if hasattr(stats_widget, 'check_last_row_and_add_new'):
                try:
                    stats_widget.check_last_row_and_add_new()
                except Exception as e:
                    pass

        finally:
            # 恢复信号连接
            table.blockSignals(False)

        # 设置最后一行空白行，确保序号列正确
        try:
            last_row = table.rowCount() - 1

            # 临时断开信号，防止设置空白行时触发事件
            table.blockSignals(True)

            # 设置最后一行序号列（自动填充且不可编辑）
            seq_item = QTableWidgetItem(str(last_row + 1))
            seq_item.setTextAlignment(Qt.AlignCenter)
            seq_item.setFlags(seq_item.flags() & ~Qt.ItemIsEditable)  # 序号列不可编辑
            table.setItem(last_row, 0, seq_item)

            # 设置其他列为空白
            for col in range(1, table.columnCount()):
                empty_item = QTableWidgetItem("")
                empty_item.setTextAlignment(Qt.AlignCenter)
                table.setItem(last_row, col, empty_item)

            # 确保最后一行空白行的其他列不可编辑（因为管口代号为空）
            from modules.guankoudingyi.funcs.funcs_pipe_table import control_last_row_editable_state
            control_last_row_editable_state(stats_widget, enable_editing=False)

        except Exception as e:
            pass
        finally:
            # 恢复信号连接
            table.blockSignals(False)

            # 处理管板元件的特殊逻辑
            # try:
            #     post_import_processing(stats_widget)
            # except Exception as e:
            #     print(f"后处理过程中发生错误: {e}")

    except Exception as e:
        raise e


"""获取指定列单位（公称尺寸、压力等级、焊端规格）"""
def get_column_unit_type(worksheet, column_index):
    """
    获取指定列的单位类型
    :param worksheet: Excel工作表对象
    :param column_index: 列索引（从1开始）
    :return: 单位类型字符串，如 'DN', 'NPS', 'Class', 'PN', 'Sch', 'mm'
    """
    try:
        # 检查第1行和第2行的表头
        header1 = worksheet.cell(row=1, column=column_index).value
        header2 = worksheet.cell(row=2, column=column_index).value

        # 将表头转换为字符串并去除空格
        header1_str = str(header1).strip() if header1 else ""
        header2_str = str(header2).strip() if header2 else ""

        # 根据表头内容判断单位类型
        if "DN" in header1_str or "DN" in header2_str:
            return "DN"
        elif "NPS" in header1_str or "NPS" in header2_str:
            return "NPS"
        elif "Class" in header1_str or "Class" in header2_str:
            return "Class"
        elif "PN" in header1_str or "PN" in header2_str:
            return "PN"
        elif "Sch" in header1_str or "Sch" in header2_str:
            return "Sch"
        elif "mm" in header1_str or "mm" in header2_str:
            return "mm"
        else:
            return None

    except Exception as e:
        return None


"""验证公称尺寸列"""
def validate_nominal_size_by_unit(nominal_size_value, unit_type, product_id, flange_standard=None):
    """
    根据单位类型和法兰标准验证公称尺寸是否在数据库中存在
    :param nominal_size_value: 公称尺寸值
    :param unit_type: 单位类型（DN或NPS）
    :param product_id: 产品ID
    :param flange_standard: 法兰标准，用于筛选公称尺寸范围
    :return: 如果数据库中存在则返回原值，否则返回空字符串
    """
    try:
        if not nominal_size_value or not unit_type:
            return ""

        from modules.guankoudingyi.db_cnt import get_connection, db_config_1
        import pymysql

        conn = get_connection(**db_config_1)
        cursor = conn.cursor(pymysql.cursors.DictCursor)

        # 根据单位类型选择查询列
        if unit_type == "DN":
            column_name = "DN"
        elif unit_type == "NPS":
            column_name = "NPS"
        else:
            return ""

        # 处理查询值
        if isinstance(nominal_size_value, float):
            query_value = str(int(nominal_size_value))
        else:
            query_value = str(nominal_size_value).strip()

        # 构建基础查询
        base_sql = f"""
            SELECT OD FROM 公称尺寸表
            WHERE `{column_name}` = %s
        """
        
        # 根据法兰标准添加筛选条件
        if flange_standard:
            if flange_standard in ["HG/T 20615-2009","SH/T 3406-2022"]:
                # DN≤600 或 NPS≤24
                if unit_type == "DN":
                    base_sql += " AND CAST(`DN` AS UNSIGNED) <= 600"
                elif unit_type == "NPS":
                    base_sql += " AND CAST(`NPS` AS UNSIGNED) <= 24"
            elif flange_standard in ["HG/T 20623-2009(A)", "HG/T 20623-2009(B)","SH/T 3406-2022(A)", "SH/T 3406-2022(B)"]:
                # DN>600 或 NPS>24
                if unit_type == "DN":
                    base_sql += " AND CAST(`DN` AS UNSIGNED) > 600"
                elif unit_type == "NPS":
                    base_sql += " AND CAST(`NPS` AS UNSIGNED) > 24"
        
        base_sql += " LIMIT 1"
        
        cursor.execute(base_sql, (query_value,))
        row = cursor.fetchone()

        cursor.close()
        conn.close()

        # 如果数据库中存在则返回原值，否则返回空字符串
        if row:
            return nominal_size_value
        else:
            return ""

    except Exception as e:
        return ""

"""验证焊端规格列"""
def validate_weld_end_spec_by_unit(weld_end_spec_value, unit_type, product_id):
    """
    根据单位类型验证焊端规格是否合法
    :param weld_end_spec_value: 焊端规格值
    :param unit_type: 单位类型（Sch或mm）
    :param product_id: 产品ID
    :return: 如果合法则返回原值，否则返回空字符串
    """
    try:
        if not weld_end_spec_value or not unit_type:
            return ""

        if unit_type == "Sch":
            # 查询焊端规格类型表
            from modules.guankoudingyi.db_cnt import get_connection, db_config_1
            import pymysql

            conn = get_connection(**db_config_1)
            cursor = conn.cursor(pymysql.cursors.DictCursor)

            # 查询焊端规格类型表
            query_value = str(weld_end_spec_value).strip()
            sql = """
                SELECT 焊端规格类型Sch FROM 焊端规格类型表
                WHERE 焊端规格类型Sch = %s
                LIMIT 1
            """
            cursor.execute(sql, (query_value,))
            row = cursor.fetchone()

            cursor.close()
            conn.close()

            # 如果数据库中存在则返回原值，否则返回空字符串
            if row:
                return weld_end_spec_value
            else:
                return ""

        elif unit_type == "mm":
            # mm单位：数据必须为数字
            try:
                if weld_end_spec_value.strip() == "程序推荐":
                    return weld_end_spec_value
                else:
                    # 尝试转换为数字
                    float(weld_end_spec_value)
                    return weld_end_spec_value

            except (ValueError, TypeError):
                return ""
        else:
            return ""

    except Exception as e:
        return ""


"""检查模板的单位类型与当前界面是否一致，如果不一致则切换界面单位类型"""
def _check_and_switch_unit_types(stats_widget, worksheet):
    """
    检查模板的单位类型与当前界面是否一致，如果不一致则切换界面单位类型
    :param stats_widget: Stats类实例
    :param worksheet: Excel工作表对象
    """
    try:
        from modules.guankoudingyi.funcs.pipe_get_units_types import get_current_unit_types_from_ui

        # 获取当前界面的单位类型
        current_unit_types = get_current_unit_types_from_ui(stats_widget)

        # 1. 检查并切换公称尺寸单位类型
        template_nominal_size_unit = get_column_unit_type(worksheet, 5)  # 公称尺寸列
        if template_nominal_size_unit and template_nominal_size_unit in ["DN", "NPS"]:
            current_nominal_size_type = current_unit_types.get("公称尺寸类型", "")
            if current_nominal_size_type != template_nominal_size_unit:
                _switch_combo_unit_type(stats_widget, 'combo_nominal_size_type',
                                        template_nominal_size_unit, "公称尺寸")

        # 2. 检查并切换压力等级单位类型
        template_pressure_level_unit = get_column_unit_type(worksheet, 7)  # 压力等级列
        if template_pressure_level_unit and template_pressure_level_unit in ["Class", "PN"]:
            current_pressure_level_type = current_unit_types.get("公称压力类型", "")
            if current_pressure_level_type != template_pressure_level_unit:
                _switch_combo_unit_type(stats_widget, 'combo_pressure_level_type',
                                        template_pressure_level_unit, "压力等级")

        # 3. 检查并切换焊端规格单位类型
        template_weld_end_spec_unit = get_column_unit_type(worksheet, 10)  # 焊端规格列
        if template_weld_end_spec_unit and template_weld_end_spec_unit in ["Sch", "mm"]:
            current_weld_end_spec_type = current_unit_types.get("焊端规格类型", "")
            if current_weld_end_spec_type != template_weld_end_spec_unit:
                _switch_combo_unit_type(stats_widget, 'combo_weld_end_spec_type',
                                        template_weld_end_spec_unit, "焊端规格")

    except Exception as e:
        print(f"切换单位类型时发生错误: {e}")


"""切换指定下拉框"""
def _switch_combo_unit_type(stats_widget, combo_attr_name, target_unit_type, unit_name):
    """
    切换指定下拉框的单位类型
    :param stats_widget: Stats类实例
    :param combo_attr_name: 下拉框属性名
    :param target_unit_type: 目标单位类型
    :param unit_name: 单位名称（用于日志）
    """
    try:
        if hasattr(stats_widget, combo_attr_name):
            combo = getattr(stats_widget, combo_attr_name)
            if combo:
                # 查找对应的索引
                for i in range(combo.count()):
                    if combo.itemText(i) == target_unit_type:
                        combo.setCurrentIndex(i)
                        print(f"已切换{unit_name}单位类型到 '{target_unit_type}'")
                        break
    except Exception as e:
        print(f"切换{unit_name}单位类型时发生错误: {e}")

"""验证管口所属元件"""
def validate_pipe_belong_by_product_type(pipe_belong_value, product_id, pipe_function_value=None):
    """
    根据产品类型验证管口所属元件是否合法
    :param pipe_belong_value: 管口所属元件值
    :param product_id: 产品ID
    :param pipe_function_value: 管口功能值（用于管程/壳程入口出口的特殊限制）
    :return: 如果合法则返回原值，否则返回空字符串
    """
    try:
        if not pipe_belong_value or not product_id:
            return ""

        from modules.guankoudingyi.obtain_product_type_version import get_product_type_and_version

        # 获取产品类型和型式
        product_type, product_version = get_product_type_and_version(product_id)
        print("product_type: ", product_type)
        print("product_version",product_version)


        if not product_type:
            return pipe_belong_value

        # 定义各产品类型允许的元件类型（通用场景）
        allowed_components = {
            "AEU": ["管箱圆筒", "管箱平盖", "壳体圆筒", "壳体封头","固定管板"],
            "BEU": ["管箱圆筒", "管箱封头", "壳体圆筒", "壳体封头","固定管板"],
            "AES": ["管箱圆筒", "管箱平盖", "壳体圆筒", "外头盖圆筒", "外头盖封头","固定管板"],
            "BES": ["管箱圆筒", "管箱封头", "壳体圆筒", "外头盖圆筒", "外头盖封头","固定管板"],
            "NEN": ["前端管箱圆筒", "后端管箱圆筒", "壳体圆筒", "前端管箱平盖", "后端管箱平盖","前端管板","后端管板"],
            "BEM": ["前端管箱圆筒", "后端管箱圆筒", "壳体圆筒", "前端管箱封头", "后端管箱封头","前端管板","后端管板"],
            "AEM": ["前端管箱圆筒", "后端管箱圆筒", "壳体圆筒", "前端管箱平盖", "后端管箱封头","前端管板","后端管板"],
            "AKU": ["管箱圆筒", "管箱平盖", "壳程大端圆筒", "锥壳", "壳程封头"],
            "BKU": ["管箱圆筒", "管箱封头", "壳程大端圆筒", "锥壳", "壳程封头"],
            "NEN(Head)": ["前端管箱圆筒", "后端管箱圆筒", "壳体圆筒", "前端管箱封头", "后端管箱封头","前端管板","后端管板"]

        }

        # 特殊场景
        if pipe_function_value in ["管程入口", "管程出口"]:
            tube_allowed = {
                "AEU": ["管箱圆筒", "管箱平盖"],
                "AES": ["管箱圆筒", "管箱平盖"],
                "BEU": ["管箱圆筒", "管箱封头"],
                "BES": ["管箱圆筒", "管箱封头"],
                "NEN": ["前端管箱圆筒", "后端管箱圆筒", "前端管箱平盖", "后端管箱平盖"],
                "BEM": ["前端管箱圆筒", "后端管箱圆筒", "前端管箱封头", "后端管箱封头"],
                "AEM": ["前端管箱圆筒", "后端管箱圆筒", "前端管箱平盖", "后端管箱封头"],
                "NEN(Head)": ["前端管箱圆筒", "后端管箱圆筒", "前端管箱封头", "后端管箱封头"],
                "AKU": ["管箱圆筒", "管箱平盖"],
                "BKU": ["管箱圆筒", "管箱封头"],
            }
            allowed_list = tube_allowed.get(product_version, [])
        elif pipe_function_value == "壳程入口" and product_version in ["AKU", "BKU"]:
            allowed_list = ["壳程大端圆筒","锥壳"]
        elif pipe_function_value in ["壳程入口", "壳程出口"]:
            if pipe_function_value == "壳程入口" and product_version in ["AKU", "BKU"]:
              allowed_list = ["壳程大端圆筒","锥壳"]
            else:
              allowed_list = ["壳体圆筒"]
        elif pipe_function_value in ["壳程气相出口", "壳程液相出口"] and product_version in ["AKU", "BKU"]:
            allowed_list = ["壳程大端圆筒"]
        elif pipe_function_value in ["壳程液位计1", "壳程液位计2","壳程温度计"] and product_version in ["AKU", "BKU"]:
            allowed_list = ["壳程大端圆筒","壳程封头"]

        else:
            # 其他保持原有逻辑
            allowed_list = allowed_components.get(product_version, [])

        # 检查管口所属元件是否在允许列表中
        if pipe_belong_value in allowed_list:
            return pipe_belong_value
        else:
            return ""

    except Exception as e:
        return ""

"""验证轴向定位基准"""
def validate_axial_position_base(axial_position_base_value, pipe_belong_value, pipe_function_value=None, current_row=None, imported_data=None):
    """
    根据管口所属元件验证轴向定位基准是否合法
    同时处理壳程入口和壳程出口的轴向定位基准互斥逻辑
    :param axial_position_base_value: 轴向定位基准值
    :param pipe_belong_value: 管口所属元件值
    :param pipe_function_value: 管口功能值（用于互斥逻辑）
    :param current_row: 当前行号（用于互斥逻辑）
    :param imported_data: 导入的数据列表（用于互斥逻辑）
    :return: 如果合法则返回原值，否则返回空字符串
    """
    try:
        if not axial_position_base_value or not pipe_belong_value:
            return ""

        # 定义各元件类型允许的轴向定位基准
        allowed_bases = {
            # 圆筒类元件：左基准线或右基准线
            "管箱圆筒": ["左基准线", "右基准线"],
            "壳体圆筒": ["左基准线", "右基准线"],
            "外头盖圆筒": ["左基准线", "右基准线"],
            "前端管箱圆筒": ["左基准线", "右基准线"],
            "后端管箱圆筒": ["左基准线", "右基准线"],
            "壳程大端圆筒": ["左基准线", "右基准线"],
            "锥壳": ["左基准线", "右基准线"],

            # 平盖类元件：平盖中心线
            "管箱平盖": ["平盖中心线"],
            "前端管箱平盖": ["平盖中心线"],
            "后端管箱平盖": ["平盖中心线"],

            # 封头类元件：封头中心线
            "管箱封头": ["封头中心线"],
            "壳体封头": ["封头中心线"],
            "壳程封头": ["封头中心线"],
            "外头盖封头": ["封头中心线"],
            "前端管箱封头": ["封头中心线"],
            "后端管箱封头": ["封头中心线"],

            #管板类元件：管程侧端面或壳程侧端面
            "固定管板":["管程侧端面","壳程侧端面"],
            "前端管板": ["管程侧端面", "壳程侧端面"],
            "后端管板": ["管程侧端面", "壳程侧端面"]

        }

        # 获取当前元件类型允许的轴向定位基准
        allowed_list = allowed_bases.get(pipe_belong_value, [])

        # 检查轴向定位基准是否在允许列表中
        if axial_position_base_value not in allowed_list:
            return ""

        # 处理壳程入口和壳程出口的轴向定位基准互斥逻辑（仅判断模板内）
        if (pipe_function_value in ["壳程入口", "壳程出口"] and
            axial_position_base_value in ["左基准线", "右基准线"] and
            imported_data is not None and current_row is not None):

            # 在模板内查找另一个壳程管口（壳程入口或壳程出口）
            target_function = "壳程出口" if pipe_function_value == "壳程入口" else "壳程入口"

            # 在当前行之前的数据中查找另一个壳程管口
            for i in range(current_row):
                data = imported_data[i]
                if data.get("管口功能") == target_function:
                    other_base = data.get("轴向定位基准", "")
                    if other_base in ["左基准线", "右基准线"] and other_base == axial_position_base_value:
                        # 不互斥，置空当前管口的轴向定位基准
                        return ""
                    break

        return axial_position_base_value

    except Exception as e:
        return ""

"""验证轴向定位距离"""
def validate_axial_position_distance(axial_distance_value, nominal_size_value, pipe_belong_value=None,
                                     product_id=None, stats_widget=None, max_nominal_size_od=None):
    """
    验证轴向定位距离是否合法（支持 程序推荐/居中/数字）。
    规则：
    1) 轴向定位距离为空 → 返回空字符串
    2) 必须先有公称尺寸与管口所属元件，否则返回空字符串
    3) 若为文本：
       - 管板类（固定管板/前端管板/后端管板）仅允许“居中”
       - 非管板类允许“程序推荐”或“居中”，原样返回
    4) 若为数值：
       - “管箱圆筒/外头盖圆筒”：min = 0.5*当前管口公称尺寸对应的接管外径od；
         max 若可获得公称尺寸最大的管口对应的接管实际外径od 则 = max_od*2.5 - 0.5*当前管口对应的接管实际外径od；
         无法获得最大值时，仅校验不小于 min。
       - “壳体圆筒”：min = 0.5*当前管口公称尺寸对应的接管外径od；max = 换热管长度 - 0.5*当前管口公称尺寸对应的接管外径od。
       - “固定管板”：min = 0.5*当前管口公称尺寸对应的接管外径od；max = 管板上最大管口外径的 50 倍。
       - “前端管板/后端管板”：min = 0；max = 管板上最大管口外径的 50 倍。
       - 其他所属元件：置空。
    :param axial_distance_value: 轴向定位距离值
    :param nominal_size_value: 公称尺寸值（DN/NPS）
    :param pipe_belong_value: 管口所属元件
    :param product_id: 产品ID
    :param stats_widget: 可选，用于在有界面时读取最大公称尺寸/管板最大公称尺寸
    :return: 合法返回原值，否则返回空字符串
    """
    try:
        # 1) 空值 → 置空
        if not axial_distance_value:
            return ""

        # 2) 必须先填公称尺寸与所属元件
        if not nominal_size_value:
            return ""
        if not pipe_belong_value:
            return ""

        # 3) 允许的文本值
        if axial_distance_value in ["程序推荐", "居中"]:
            pipe_belong_str = str(pipe_belong_value or "")
            # 管板类：仅允许“居中”
            if "管板" in pipe_belong_str:
                return axial_distance_value if axial_distance_value == "居中" else ""
            # 其他元件：程序推荐/居中均视为合法
            return axial_distance_value

        # 4) 数字解析
        try:
            distance_val = float(axial_distance_value)
        except (ValueError, TypeError):
            return ""

        # 5) 计算当前管口对应的接管实际外径od
        try:
            from modules.guankoudingyi.funcs.funcs_pipe_comboBox_value import (
                get_component_nominal_size_od,
                get_max_pipe_nominal_size_from_ui,
                get_heat_exchanger_tube_length,
                get_max_tubesheet_nominal_size_from_ui,
            )
        except Exception:
            return ""

        current_od = get_component_nominal_size_od(
            nominal_size_value, product_id=product_id, stats_widget=stats_widget
        )
        if current_od is None:
            return ""

        # 6) 按所属元件类型计算范围
        pipe_belong_str = str(pipe_belong_value or "")
        min_distance = None
        max_distance = None

        # 管板类（固定管板/前端管板/后端管板）
        if "管板" in pipe_belong_str:
            # 获取管板上最大公称尺寸对应的外径（仅在有界面时可用）
            max_tubesheet_od = None
            if stats_widget:
                try:
                    max_ts_nominal = get_max_tubesheet_nominal_size_from_ui(stats_widget)
                    if max_ts_nominal:
                        max_tubesheet_od = get_component_nominal_size_od(
                            max_ts_nominal, product_id=product_id, stats_widget=stats_widget
                        )
                except Exception:
                    max_tubesheet_od = None

            # 固定管板：min = 0.5 * 当前OD；前/后端管板：min = 0
            if "固定管板" in pipe_belong_str:
                min_distance = round(0.5 * current_od, 2)
            else:
                min_distance = 0.0

            # max = 管板上最大管口外径的 50 倍（若能获取）
            if isinstance(max_tubesheet_od, (int, float)) and max_tubesheet_od > 0:
                max_distance = round(50.0 * max_tubesheet_od, 2)

        # 管箱圆筒 / 外头盖圆筒
        elif ("管箱圆筒" in pipe_belong_str) or ("外头盖圆筒" in pipe_belong_str):
            min_distance = round(0.5 * current_od, 2)
            # 优先使用外部传入的最大公称尺寸对应的接管实际外径（od）；否则在有界面时从UI获取
            if isinstance(max_nominal_size_od, (int, float)) and max_nominal_size_od > 0:
                max_distance = round(max_nominal_size_od * 2.5 - 0.5 * current_od, 2)
            else:
                max_nominal_size = get_max_pipe_nominal_size_from_ui(stats_widget) if stats_widget else None
                if max_nominal_size:
                    max_od = get_component_nominal_size_od(
                        max_nominal_size, product_id=product_id, stats_widget=stats_widget
                    )
                    if isinstance(max_od, (int, float)):
                        max_distance = round(max_od * 2.5 - 0.5 * current_od, 2)

        # 壳体圆筒、壳程大端
        elif ("壳体圆筒" in pipe_belong_str)or("壳程大端圆筒"in pipe_belong_str):
            tube_ok, tube_nominal_diameter = get_nominal_diameter(product_id, "管箱")
            shell_ok, shell_nominal_diameter = get_nominal_diameter(product_id, "壳体")
            if (not tube_ok) or (tube_nominal_diameter is None):
                tube_nominal_diameter = 300
            if (not shell_ok) or (shell_nominal_diameter is None):
                shell_nominal_diameter = 400
            cone_length = (shell_nominal_diameter - tube_nominal_diameter) /math.tan(math.radians(30))
            if cone_length < 0:
                cone_length = 0
            if pipe_belong_value=="壳程大端圆筒":
                min_distance = round(0.5 * current_od, 2)
                tube_len = get_heat_exchanger_tube_length(product_id) if product_id else None
                if isinstance(tube_len, (int, float)):
                    max_distance = round(tube_len + 1 / 2 * shell_nominal_diameter - cone_length - 0.5 * current_od, 2)
            else:
                min_distance = round(0.5 * current_od, 2)
                tube_len = get_heat_exchanger_tube_length(product_id) if product_id else None
                if isinstance(tube_len, (int, float)):
                    max_distance = round(tube_len + 1 / 2 * shell_nominal_diameter  - 0.5 * current_od, 2)
        elif "锥壳"in pipe_belong_str:
            min_distance = round(0.5 * current_od, 2)
            tube_ok, tube_nominal_diameter = get_nominal_diameter(product_id, "管箱")
            shell_ok, shell_nominal_diameter = get_nominal_diameter(product_id, "壳体")
            if (not tube_ok) or (tube_nominal_diameter is None):
                tube_nominal_diameter = 300
            if (not shell_ok) or (shell_nominal_diameter is None):
                shell_nominal_diameter = 400
            cone_length = (shell_nominal_diameter - tube_nominal_diameter) / math.tan(math.radians(30))
            if cone_length < 0:
                cone_length = 0
            if isinstance(cone_length, (int, float)):
                max_distance = round(cone_length  - 0.5 * current_od, 2)






        else:
            # 其他类型暂不支持 → 置空
            return ""

        # 7) 校验范围
        if min_distance is None:
            return ""

        if max_distance is None:
            # 仅校验下限
            return axial_distance_value if distance_val >= min_distance else ""
        else:
            return axial_distance_value if (min_distance <= distance_val <= max_distance) else ""

    except Exception:
        return ""

"""验证轴向夹角"""
def validate_axial_angle(axial_angle_value):
    """
    验证轴向夹角是否合法
    :param axial_angle_value: 轴向夹角值
    :return: 如果合法则返回原值，否则返回空字符串
    """
    try:
        if not axial_angle_value:
            return ""

        # 检查是否为数值
        try:
            float_value = float(axial_angle_value)
            # 数值必须在-90到90之间
            if -90 <= float_value <= 90:
                return axial_angle_value
            else:
                return ""
        except (ValueError, TypeError):
            # 不是数值
            return ""

    except Exception as e:
        return ""

"""验证轴向方位"""
def validate_circumferential_position(circumferential_position_value):
    """
    验证周向方位是否合法
    :param circumferential_position_value: 周向方位值
    :return: 如果合法则返回原值，否则返回空字符串
    """
    try:
        if not circumferential_position_value:
            return ""

        # 检查是否为数值
        try:
            float_value = float(circumferential_position_value)
            # 数值必须在0到360之间
            if 0 <= float_value < 360:
                return circumferential_position_value
            else:
                return ""
        except (ValueError, TypeError):
            # 不是数值
            return ""

    except Exception as e:
        return ""

"""验证偏心距"""
# def validate_eccentricity_with_error_info(eccentricity_value, pipe_belong_value, product_id, row,
#                                           axial_angle_value=None):
#     """
#     验证偏心距是否合法，并返回验证结果和错误信息
#     :param eccentricity_value: 偏心距值
#     :param pipe_belong_value: 管口所属元件值
#     :param product_id: 产品ID
#     :param row: 行号
#     :param axial_angle_value: 轴向夹角值
#     :return: (验证后的值, 错误信息列表)
#     """
#     try:
#         if not eccentricity_value:
#             return eccentricity_value, []
#
#         # 1. 检查轴向夹角和偏心距是否同时赋值
#         if axial_angle_value and axial_angle_value.strip():
#             return "", [f"偏心距列，第{row}行数据不合法"]
#
#         # 2. 先检查管口所属元件是否填写
#         if not pipe_belong_value:
#             return "", [f"偏心距列，第{row}行数据不合法"]
#
#         # 3. 检查公称直径是否填写
#         try:
#             from modules.guankoudingyi.funcs.funcs_pipe_comboBox_value import get_nominal_diameter
#             success, result = get_nominal_diameter(product_id, pipe_belong_value)
#
#             if not success:
#                 return "", [f"偏心距列，第{row}行，请先在条件输入界面填写公称直径"]
#
#             # 4. 验证偏心距范围：-1/2*公称直径 ~ 1/2*公称直径
#             try:
#                 float_value = float(eccentricity_value)
#                 half_diameter = result / 2
#
#                 if -half_diameter <= float_value <= half_diameter:
#                     return eccentricity_value, []
#                 else:
#                     return "", [f"偏心距列，第{row}行数据不合法"]
#             except (ValueError, TypeError):
#                 return "", [f"偏心距列，第{row}行数据不合法"]
#
#         except Exception:
#             return "", [f"偏心距列，第{row}行数据不合法"]
#
#     except Exception as e:
#         return "", [f"偏心距列，第{row}行数据不合法"]
def validate_eccentricity_with_error_info(eccentricity_value, pipe_belong_value, product_id, row,
                                          axial_angle_value=None):
    """
    验证偏心距是否合法，并返回验证结果和错误信息
    :param eccentricity_value: 偏心距值
    :param pipe_belong_value: 管口所属元件值
    :param product_id: 产品ID
    :param row: 行号
    :param axial_angle_value: 轴向夹角值
    :return: (验证后的值, 错误信息列表)
    """
    def is_non_zero_value(value):
        """辅助函数：判断值是否为非空且非零的有效数值"""
        if not value:  # 空值、None等直接视为无效
            return False
        # 处理字符串类型的空白（如"  "）
        stripped_value = str(value).strip()
        if not stripped_value:
            return False
        # 尝试转换为数值，判断是否非零
        try:
            num = float(stripped_value)
            return num != 0
        except ValueError:  # 无法转换为数值的情况（如"abc"）视为无效
            return False

    try:
        if not eccentricity_value:
            return eccentricity_value, []

        # 1. 检查轴向夹角和偏心距是否同时为非零有效值（0不算有效值）
        ecc_non_zero = is_non_zero_value(eccentricity_value)
        axial_non_zero = is_non_zero_value(axial_angle_value)
        if ecc_non_zero and axial_non_zero:
            return "", [f"偏心距列，第{row}行：偏心距与夹角不能同时填写非零值"]

        # 2. 检查管口所属元件是否填写
        if not pipe_belong_value:
            return "", [f"偏心距列，第{row}行数据不合法"]

        # 3. 检查公称直径是否填写
        try:
            from modules.guankoudingyi.funcs.funcs_pipe_comboBox_value import get_nominal_diameter
            success, result = get_nominal_diameter(product_id, pipe_belong_value)

            if not success:
                return "", [f"偏心距列，第{row}行：请先在条件输入界面填写公称直径"]

            # 4. 验证偏心距范围：-1/2*公称直径 ~ 1/2*公称直径
            try:
                float_value = float(eccentricity_value)
                half_diameter = result / 2

                if -half_diameter <= float_value <= half_diameter:
                    return eccentricity_value, []
                else:
                    return "", [f"偏心距列，第{row}行：偏心距需在{-half_diameter}~{half_diameter}范围内"]
            except (ValueError, TypeError):
                return "", [f"偏心距列，第{row}行：偏心距需为有效的数值"]

        except Exception:
            return "", [f"偏心距列，第{row}行：获取公称直径失败，请检查参数"]

    except Exception as e:
        return "", [f"偏心距列，第{row}行：验证过程出错，{str(e)}"]

"""验证外伸高度"""
def validate_extension_height_with_error_info(extension_height_value, pipe_belong_value, product_id, row):
    """
    验证外伸高度是否合法，并返回验证结果和错误信息
    :param extension_height_value: 外伸高度值
    :param pipe_belong_value: 管口所属元件值
    :param product_id: 产品ID
    :param row: 行号
    :return: (验证后的值, 错误信息列表)
    """
    try:

        if not extension_height_value:
            return extension_height_value, []

        # 1. 检查管口所属元件是否填写
        if not pipe_belong_value:
            return "", [f"外伸高度列，第{row}行数据不合法"]

        # 2. 获取公称直径
        try:
            from modules.guankoudingyi.funcs.funcs_pipe_comboBox_value import get_nominal_diameter
            success, result = get_nominal_diameter(product_id, pipe_belong_value)

            if not success:
                return "", [f"外伸高度列，第{row}行，请先在条件输入界面填写公称直径"]

            # 3. 如果外伸高度是"程序推荐"，且管口所属元件和公称直径都已填写，则允许
            if extension_height_value.strip() == "程序推荐":
                return extension_height_value, []

            # 4. 验证外伸高度不能小于1/2公称直径
            try:
                float_value = float(extension_height_value)
                half_diameter = result / 2
                print("half_diameter", half_diameter)

                if float_value >= half_diameter:
                    return extension_height_value, []
                else:
                    return "", [f"外伸高度列，第{row}行数据不合法"]
            except (ValueError, TypeError):
                return "", [f"外伸高度列，第{row}行数据不合法"]

        except Exception:
            return "", [f"外伸高度列，第{row}行数据不合法"]

    except Exception as e:
        return "", [f"外伸高度列，第{row}行数据不合法"]

"""验证法兰标准"""
def validate_flange_standard_with_error_info(flange_standard_value, row, pressure_unit_type=None):
    """
    验证法兰标准是否合法，并返回验证结果和错误信息
    根据压力类型验证法兰标准：
    - Class压力类型：允许 HG/T 20615-2009, HG/T 20623-2009(A), HG/T 20623-2009(B)
    - PN压力类型：允许 HG/T 20592-2009
    :param flange_standard_value: 法兰标准值
    :param row: 行号
    :param pressure_unit_type: 压力等级的单位类型（Class或PN）
    :return: (验证后的值, 错误信息列表)
    """
    try:
        if not flange_standard_value:
            return flange_standard_value, []

        # 根据压力类型定义允许的法兰标准值
        if pressure_unit_type == "Class":
            # Class压力类型允许的法兰标准
            allowed_standards = ["HG/T 20615-2009", "HG/T 20623-2009(A)", "HG/T 20623-2009(B)","SH/T 3406-2022","SH/T 3406-2022(A)","SH/T 3406-2022(B)"]
        elif pressure_unit_type == "PN":
            # PN压力类型允许的法兰标准
            allowed_standards = ["HG/T 20592-2009(A)","HG/T 20592-2009(B)"]
        else:
            # 如果没有指定压力类型，允许所有标准（兼容旧逻辑）
            allowed_standards = ["HG/T 20592-2009", "HG/T 20615-2009", "HG/T 20623-2009(A)", "HG/T 20623-2009(B)","SH/T 3406-2022","SH/T 3406-2022(A)","SH/T 3406-2022(B)"]

        # 检查是否为允许的法兰标准
        if flange_standard_value in allowed_standards:
            return flange_standard_value, []
        else:
            # 根据压力类型生成更详细的错误信息
            if pressure_unit_type == "Class":
                error_msg = f"法兰标准列，第{row}行数据不合法（Class压力类型仅允许：HG/T 20615-2009、HG/T 20623-2009(A)、HG/T 20623-2009(B),SH/T 3406-2022,SH/T 3406-2022(A),SH/T 3406-2022(B)）"
            elif pressure_unit_type == "PN":
                error_msg = f"法兰标准列，第{row}行数据不合法（PN压力类型仅允许：HG/T 20592-2009）"
            else:
                error_msg = f"法兰标准列，第{row}行数据不合法"
            return "", [error_msg]

    except Exception as e:
        return "", [f"法兰标准列，第{row}行数据不合法"]

"""验证压力等级"""
def validate_pressure_level_with_error_info(pressure_level_value, unit_type, row, flange_standard=None):
    """
    验证压力等级是否合法，并返回验证结果和错误信息
    根据法兰标准限制Class压力等级的允许值：
    - HG/T 20623-2009(A) 和 HG/T 20623-2009(B)：Class允许 150, 300, 600, 900
    - HG/T 20615-2009：Class允许 150, 300, 600, 900, 1500, 2500
    - PN类型：允许 2.5, 6, 10, 16, 25, 40, 63, 100, 160
    :param pressure_level_value: 压力等级值
    :param unit_type: 单位类型（Class或PN）
    :param row: 行号
    :param flange_standard: 法兰标准（用于限制Class压力等级范围）
    :return: (验证后的值, 错误信息列表)
    """
    try:
        if not pressure_level_value:
            return pressure_level_value, []

        # 定义各单位类型允许的压力等级值
        if unit_type == "Class":
            # 根据法兰标准确定Class允许的压力等级
            if flange_standard in ["HG/T 20623-2009(A)", "HG/T 20623-2009(B)"]:
                # HG/T 20623-2009(A)和(B)只允许150, 300, 600, 900
                allowed_values = {"Class": ["150", "300", "600", "900"]}
            elif flange_standard == "HG/T 20615-2009":
                # HG/T 20615-2009允许全部Class压力等级
                allowed_values = {"Class": ["150", "300", "600", "900","1500", "2500"]}
            elif flange_standard == "SH/T 3406-2022":
                allowed_values = {"Class": ["150", "300", "400","600","900","1500", "2500"]}
            elif flange_standard in ["SH/T 3406-2022(A)" ,"SH/T 3406-2022(B)"] :
                allowed_values = {"Class": ["75","150", "300","600", "900"]}
            else:
                # 默认允许全部Class压力等级（兼容旧逻辑）
                allowed_values = {"Class": ["150", "300", "600", "900", "1500", "2500"]}
        elif unit_type == "PN":
            # PN类型的压力等级
            allowed_values = {"PN": ["2.5", "6", "10", "16", "25", "40", "63", "100", "160"]}
        else:
            return "", [f"压力等级列，第{row}行数据不合法"]

        # 检查单位类型是否支持
        if unit_type not in allowed_values:
            return "", [f"压力等级列，第{row}行数据不合法"]

        # 将输入值转换为字符串进行比较
        input_value = str(pressure_level_value).strip()

        # 检查是否为允许的压力等级值
        if input_value in allowed_values[unit_type]:
            return pressure_level_value, []
        else:
            # 生成更详细的错误提示
            if unit_type == "Class" and flange_standard:
                if flange_standard in ["HG/T 20623-2009(A)", "HG/T 20623-2009(B)"]:
                    error_msg = f"压力等级列，第{row}行数据不合法（{flange_standard}标准下Class压力等级仅允许：150、300、600、900）"
                elif flange_standard == "HG/T 20615-2009":
                    error_msg = f"压力等级列，第{row}行数据不合法（{flange_standard}标准下Class压力等级仅允许：150、300、600、900、1500、2500）"
                elif flange_standard == "SH/T 3406-2022":
                    error_msg = f"压力等级列，第{row}行数据不合法（{flange_standard}标准下Class压力等级仅允许：150、300、400、600、900、1500、2500）"
                elif flange_standard in ["SH/T 3406-2022(A)", "SH/T 3406-2022(B)"]:
                    error_msg = f"压力等级列，第{row}行数据不合法（{flange_standard}标准下Class压力等级仅允许：150、300、600、900）"
                else:
                    error_msg = f"压力等级列，第{row}行数据不合法"
            elif unit_type == "PN":
                error_msg = f"压力等级列，第{row}行数据不合法（PN压力等级仅允许：2.5、6、10、16、25、40、63、100、160）"
            else:
                error_msg = f"压力等级列，第{row}行数据不合法"
            return "", [error_msg]

    except Exception as e:
        return "", [f"压力等级列，第{row}行数据不合法"]

"""验证法兰型式"""
def validate_flange_form_by_database(flange_form_value, flange_standard, pressure_level, pressure_level_type, row):
    """
    通过查询管口关系对应表验证法兰型式是否合法
    :param flange_form_value: 法兰型式值
    :param flange_standard: 法兰标准
    :param pressure_level: 压力等级
    :param pressure_level_type: 压力等级类型（Class或PN）
    :param row: 行号
    :return: (验证后的值, 错误信息列表)
    """
    try:
        if not flange_form_value:
            return flange_form_value, []

        # 导入数据库连接配置
        from modules.guankoudingyi.funcs.funcs_pipe_comboBox_value import get_connection, db_config_1

        conn = get_connection(**db_config_1)
        cursor = conn.cursor()

        # 查询管口关系对应表
        sql = """
            SELECT 法兰型式 FROM 管口关系对应表
            WHERE 法兰标准 = %s AND 公称压力 = %s AND 公称压力类型 = %s AND 法兰型式 = %s
            LIMIT 1
        """

        cursor.execute(sql, (flange_standard, pressure_level, pressure_level_type, flange_form_value))
        result = cursor.fetchone()

        cursor.close()
        conn.close()

        if result:
            # 在表中找到匹配的记录，保留原值
            return flange_form_value, []
        else:
            # 在表中未找到匹配的记录，置空
            return "", [f"法兰型式列，第{row}行数据不合法"]

    except Exception as e:
        print(f"查询管口关系对应表失败: {e}")
        return "", [f"法兰型式列，第{row}行数据不合法"]

"""验证密封面型式"""
def validate_sealing_face_form_by_database(sealing_face_form_value, flange_standard, pressure_level, pressure_level_type, flange_form, row):
    """
    通过查询管口关系对应表验证密封面型式是否合法
    :param sealing_face_form_value: 密封面型式值
    :param flange_standard: 法兰标准
    :param pressure_level: 压力等级
    :param pressure_level_type: 压力等级类型（Class或PN）
    :param flange_form: 法兰型式
    :param row: 行号
    :return: (验证后的值, 错误信息列表)
    """
    try:
        if not sealing_face_form_value:
            return sealing_face_form_value, []

        # 导入数据库连接配置
        from modules.guankoudingyi.funcs.funcs_pipe_comboBox_value import get_connection, db_config_1

        conn = get_connection(**db_config_1)
        cursor = conn.cursor()

        # 查询管口关系对应表
        sql = """
            SELECT 密封面型式 FROM 管口关系对应表
            WHERE 法兰标准 = %s AND 公称压力 = %s AND 公称压力类型 = %s AND 法兰型式 = %s AND 密封面型式 = %s
            LIMIT 1
        """

        cursor.execute(sql, (flange_standard, pressure_level, pressure_level_type, flange_form, sealing_face_form_value))
        result = cursor.fetchone()

        cursor.close()
        conn.close()

        if result:
            # 在表中找到匹配的记录，保留原值
            return sealing_face_form_value, []
        else:
            # 在表中未找到匹配的记录，置空
            return "", [f"密封面型式列，第{row}行数据不合法"]

    except Exception as e:
        print(f"查询管口关系对应表失败: {e}")
        return "", [f"密封面型式列，第{row}行数据不合法"]


